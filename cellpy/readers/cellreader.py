# -*- coding: utf-8 -*-
"""Datareader for cell testers and potentiostats.

This module is used for loading data and databases created by different cell
testers and exporing them in a common hdf5-format.

Example:
    >>> c = cellpy.get(["super_battery_run_01.res", "super_battery_run_02.res"]) # loads and merges the runs
    >>> voltage_curves = c.get_cap()
    >>> c.save("super_battery_run.hdf")

"""

import collections
import copy
import csv
import functools
import itertools
import logging
import numbers
import os
import sys
import time
import datetime
import warnings
from pathlib import Path
from typing import Union, Sequence, List, Optional, Iterable
from dataclasses import asdict

import numpy as np
import openpyxl
import pandas as pd
from pandas.errors import PerformanceWarning
from pint.errors import DimensionalityError
from pint import Quantity
from scipy import interpolate

from cellpy.exceptions import (
    DeprecatedFeature,
    NullData,
    WrongFileVersion,
    NoDataFound,
    UnderDefined,
)
from cellpy.parameters import prms
from cellpy.parameters.legacy.update_headers import (
    rename_summary_columns,
    rename_raw_columns,
    rename_fid_columns,
    rename_step_columns,
)
from cellpy.parameters.internal_settings import (
    get_cellpy_units,
    get_headers_normal,
    get_headers_step_table,
    get_headers_summary,
    headers_normal,
    headers_step_table,
    headers_summary,
    get_default_raw_units,
    get_default_output_units,
    CELLPY_FILE_VERSION,
    MINIMUM_CELLPY_FILE_VERSION,
    PICKLE_PROTOCOL,
    CellpyUnits,
    CellpyMetaCommon,
    CellpyMetaIndividualTest,
)

from cellpy.readers.core import (
    Data,
    FileID,
    identify_last_data_point,
    interpolate_y_on_x,
    pickle_protocol,
    xldate_as_datetime,
    generate_default_factory,
    Q,
    convert_from_simple_unit_label_to_string_unit_label,
)
from cellpy.internals.core import OtherPath

DIGITS_C_RATE = 5

HEADERS_NORMAL = get_headers_normal()  # TODO @jepe refactor this (not needed)
HEADERS_SUMMARY = get_headers_summary()  # TODO @jepe refactor this (not needed)
HEADERS_STEP_TABLE = get_headers_step_table()  # TODO @jepe refactor this (not needed)

# TODO: @jepe - new feature - method for assigning new cycle numbers and step numbers
#   - Sometimes the user forgets to increment the cycle number and it would be good
#   to have a method so that its possible to set new cycle numbers manually
#   - Some testers merges different steps into one (e.g CC-CV), it would be nice to have
#   a method for "splitting that up"

# TODO: @jepe - performance warnings - mixed types within cols (pytables)

warnings.filterwarnings("ignore", category=pd.io.pytables.PerformanceWarning)
pd.set_option("mode.chained_assignment", None)  # "raise", "warn", None

module_logger = logging.getLogger(__name__)


class CellpyCell:
    """Main class for working and storing data.

    This class is the main work-horse for cellpy where all the functions for
    reading, selecting, and tweaking your data is located. It also contains the
    header definitions, both for the cellpy hdf5 format, and for the various
    cell-tester file-formats that can be read. The class can contain
    several cell-tests and each test is stored in a list. If you see what I mean...

    Attributes:
        # TODO v.1.0.1: update this
        data: cellpy.Data object
        cellpy_units: cellpy.units object
        cellpy_datadir: path to cellpy data directory
        raw_datadir: path to raw data directory
        filestatuschecker: filestatuschecker object
        force_step_table_creation: force step table creation
        ensure_step_table: ensure step table
        limit_loaded_cycles: limit loaded cycles
        profile: profile
        select_minimal: select minimal
        empty: empty
        forced_errors: forced errors
        capacity_modifiers: capacity modifiers
        sep: separator
        cycle_mode: cycle mode
        tester: tester
        cell_name: cell name
        cellpy_file_version: cellpy file version
    """

    def __repr__(self):
        txt = f"<CellpyCell> (id={hex(id(self))})"
        if self.cell_name:
            txt += f" [name={self.cell_name}]"
        return txt

    def _repr_html_(self):
        header = f"""
            <h2>CellpyCell-object</h2>
            <b>id</b>: {hex(id(self))} <br>
            <b>name</b>: {self.cell_name} <br>
            <b>tester</b>: {self.tester} <br>
            <b>cycle_mode</b>: {self.cycle_mode} <br>
            <b>sep</b>: {self.sep} <br>
            <b>cellpy_datadir</b>: {self.cellpy_datadir} <br>
            <b>raw_datadir</b>: {self.raw_datadir} <br>
        """
        all_vars = "<p>"
        all_vars += f"""
            <b>capacity_modifiers</b>: {self.capacity_modifiers} <br>
            <b>empty</b>: {self.empty} <br>
            <b>ensure_step_table</b>: {self.ensure_step_table} <br>
            <b>filestatuschecker</b>: {self.filestatuschecker} <br>
            <b>force_step_table_creation</b>: {self.force_step_table_creation} <br>
            <b>forced_errors</b>: {self.forced_errors} <br>
            <b>limit_loaded_cycles</b>: {self.limit_loaded_cycles} <br>
            <b>profile</b>: {self.profile} <br>
            <b>cellpy_units</b>: {self.cellpy_units} <br>
            <b>select_minimal</b>: {self.select_minimal} <br>
            <b>selected_scans</b>: {self.selected_scans} <br>
        """
        all_vars += "</p>"

        cell_txt = ""

        try:
            data_txt = self.data._repr_html_()
        except NoDataFound:
            cell_txt += f"<h3>No data</h3>"
        else:
            cell_txt += f"<h3>data</h3>"
            cell_txt += f"<blockquote>{data_txt}</blockquote>"
        return header + all_vars + cell_txt

    def __str__(self):
        txt = "CellpyCell\n"
        txt += "----------\n"
        if self.cell_name:
            txt += f"session name: {self.cell_name}\n"
        if self.tester:
            txt += f"tester: {self.tester}\n"
        if self.data:
            txt += "data:\n"
            for t in str(self.data).split("\n"):
                txt += "     "
                txt += t
                txt += "\n"
            txt += "\n"
        else:
            txt += "datasets: EMPTY"
        txt += "\n"
        return txt

    def __bool__(self):
        if self.data:
            return True
        else:
            return False

    def __len__(self):
        if self.data:
            return 1
        else:
            return 0

    def __init__(
        self,
        filenames=None,
        selected_scans=None,
        profile=False,
        filestatuschecker=None,  # "modified"
        tester=None,
        initialize=False,
        cellpy_units=None,
        output_units=None,
        debug=False,
    ):
        """CellpyCell object

        Args:
            filenames: list of files to load.
            selected_scans:
            profile: experimental feature.
            filestatuschecker: property to compare cellpy and raw-files;
               default read from prms-file.
            tester: instrument used (e.g. "arbin_res") (checks prms-file as
               default).
            initialize: create a dummy (empty) dataset; defaults to False.
            cellpy_units (dict): sent to cellpy.parameters.internal_settings.get_cellpy_units
            output_units (dict): sent to cellpy.parameters.internal_settings.get_default_output_units
            debug (bool): set to True if you want to see debug messages.
        """

        # TODO v 1.1: move to data (allow for multiple testers for same cell)
        if tester is None:
            self.tester = prms.Instruments.tester
            logging.debug(f"reading instrument from prms: {prms.Instruments}")
        else:
            self.tester = tester

        self.loader = None  # this will be set in the function set_instrument
        self.debug = debug
        logging.debug("created CellpyCell instance")

        self._cell_name = None
        self.profile = profile

        self.minimum_selection = {}
        self.filestatuschecker = filestatuschecker or prms.Reader.filestatuschecker
        self.forced_errors = 0

        self.file_names = filenames or []
        if not self._is_listtype(self.file_names):
            self.file_names = [self.file_names]

        self.selected_scans = selected_scans or []
        if not self._is_listtype(self.selected_scans):
            self.selected_scans = [self.selected_scans]

        self._data = None
        self.overwrite_able = True  # attribute that prevents saving to the same filename as loaded from if False

        self.capacity_modifiers = ["reset"]

        self.list_of_step_types = [
            "charge",
            "discharge",
            "cv_charge",
            "cv_discharge",
            "taper_charge",
            "taper_discharge",
            "charge_cv",
            "discharge_cv",
            "ocvrlx_up",
            "ocvrlx_down",
            "ir",
            "rest",
            "not_known",
        ]
        # - options
        self.force_step_table_creation = prms.Reader.force_step_table_creation
        self.force_all = prms.Reader.force_all
        self.sep = prms.Reader.sep
        self._cycle_mode = None
        self.select_minimal = prms.Reader.select_minimal
        self.limit_loaded_cycles = prms.Reader.limit_loaded_cycles
        self.limit_data_points = None
        self.ensure_step_table = prms.Reader.ensure_step_table
        self.ensure_summary_table = prms.Reader.ensure_summary_table
        self.raw_datadir = OtherPath(prms.Paths.rawdatadir)
        self.cellpy_datadir = OtherPath(prms.Paths.cellpydatadir)
        self.auto_dirs = prms.Reader.auto_dirs  # v2.0

        # - headers and instruments
        self.headers_normal = headers_normal
        self.headers_summary = headers_summary
        self.headers_step_table = headers_step_table
        self.instrument_factory = None
        self.register_instrument_readers()
        self.set_instrument()
        # - units used by cellpy
        self.cellpy_units = get_cellpy_units(cellpy_units)
        self.output_units = get_default_output_units(output_units)  # v2.0

        if initialize:
            self.initialize()

    def initialize(self):
        """Initialize the CellpyCell object with empty Data instance."""

        logging.debug("Initializing...")
        self._data = Data()

    # the batch utility might be using session name
    # the cycle and ica collector are using session name
    # improvement suggestion: use data.cell_name instead
    @property
    def cell_name(self):
        """returns the session name"""

        if not self._cell_name:
            try:
                return self.data.cell_name
            except NoDataFound:
                return None
        else:
            return self._cell_name

    @cell_name.setter
    def cell_name(self, n):
        """sets the session name"""

        self._cell_name = n
        if not self.data.cell_name:
            self.data.cell_name = n

    def _invent_a_cell_name(self, filename=None, override=False):
        if filename is None:
            self.cell_name = "nameless"
            return
        if self.cell_name and not override:
            return
        if isinstance(filename, (list, tuple)):
            names = [Path(n).with_suffix("").name for n in filename]
            names = [
                n.replace(" ", "_").replace("-", "_").replace(".", "_") for n in names
            ]
            names = list(set(names))
            if len(names) == 1:
                self.cell_name = names[0]
            else:
                self.cell_name = "-".join(names)
        else:
            self.cell_name = Path(filename).with_suffix("").name

    @property
    def mass(self):
        """returns the mass"""
        return self.data.mass

    @mass.setter
    def mass(self, m):
        self.data.mass = self._dump_cellpy_unit(m, "mass")

    @property
    def active_mass(self):
        """returns the active mass (same as mass)"""
        return self.data.mass

    @active_mass.setter
    def active_mass(self, m):
        self.data.mass = self._dump_cellpy_unit(m, "mass")

    @property
    def tot_mass(self):
        """returns the total mass"""
        return self.data.tot_mass

    @tot_mass.setter
    def tot_mass(self, m):
        self.data.tot_mass = self._dump_cellpy_unit(m, "mass")

    @property
    def active_electrode_area(self):
        """returns the area"""
        return self.data.active_electrode_area

    @active_electrode_area.setter
    def active_electrode_area(self, a):
        self.data.active_electrode_area = self._dump_cellpy_unit(a, "area")

    @property
    def nom_cap(self):
        """returns the nominal capacity"""
        return self.data.nom_cap

    @nom_cap.setter
    def nom_cap(self, c):
        self.data.nom_cap = self._dump_cellpy_unit(c, "nominal_capacity")

    @property
    def nominal_capacity(self):
        """returns the nominal capacity"""
        return self.data.nom_cap

    @nominal_capacity.setter
    def nominal_capacity(self, c):
        self.data.nom_cap = self._dump_cellpy_unit(c, "nominal_capacity")

    def _dump_cellpy_unit(self, value, parameter):
        """Parse for unit, update cellpy_units class, and return magnitude."""
        if isinstance(value, numbers.Number):
            return value
        logging.critical(f"Parsing {parameter} ({value})")

        try:
            c = Q(value)
            c_unit = c.units
            self.cellpy_units[parameter] = f"{c_unit}"
            logging.critical(f"Updated your cellpy_units['{parameter}'] to '{c_unit}'")
            c = c.magnitude
        except ValueError:
            logging.debug(f"Could not parse {value}")
            return
        return c

    @property
    def nom_cap_specifics(self):
        """returns the nominal capacity specific"""
        return self.data.meta_common.nom_cap_specifics

    @nom_cap_specifics.setter
    def nom_cap_specifics(self, c):
        if c.lower() == "areal":
            self.cellpy_units.nominal_capacity = (
                f"{self.cellpy_units.charge}/{self.cellpy_units.specific_areal}"
            )
        elif c.lower() == "gravimetric":
            self.cellpy_units.nominal_capacity = (
                f"{self.cellpy_units.charge}/{self.cellpy_units.specific_gravimetric}"
            )
        elif c.lower() == "volumetric":
            self.cellpy_units.nominal_capacity = (
                f"{self.cellpy_units.charge}/{self.cellpy_units.specific_volumetric}"
            )
        else:
            logging.warning(f"Unknown nominal capacity specific: {c}")
            return
        self.data.meta_common.nom_cap_specifics = c

    @property
    def raw_units(self):
        """returns the raw_units dictionary"""

        return self.data.raw_units

    @property
    def data(self):
        """returns the DataSet instance"""

        if not self._data:
            logging.debug(
                "NoDataFound - might consider defaulting to create one in the future"
            )
            raise NoDataFound
        else:
            return self._data

    @data.setter
    def data(self, new_cell):
        """sets the DataSet instance"""

        self._data = new_cell

    @property
    def empty(self):
        """Gives True if the CellpyCell object is empty (or non-functional)"""

        return not self._validate_cell()

    @classmethod
    def vacant(cls, cell=None):
        """Create a CellpyCell instance.

        Args:
            cell (CellpyCell instance): the attributes from the data will be
                copied to the new Cellpydata instance.

         Returns:
            CellpyCell instance.

        """

        new_cell = cls(initialize=True)
        if cell is not None:
            new_cell.data.meta_common = cell.data.meta_common
            new_cell.data.meta_test_dependent = cell.data.meta_test_dependent

            new_cell.data.raw_data_files = cell.data.raw_data_files
            new_cell.data.raw_data_files_length = cell.data.raw_data_files_length
            new_cell.data.raw_units = cell.data.raw_units
            new_cell.data.raw_limits = cell.data.raw_limits

            new_cell.data.loaded_from = cell.data.loaded_from
            new_cell.data._raw_id = cell.data.raw_id
        return new_cell

    def mod_raw_split_cycle(self, data_points: List) -> None:
        """Split cycle(s) into several cycles.

        Args:
            data_points: list of the first data point(s) for additional cycle(s).

        """

        logging.info(f"splitting cycles at {data_points}")
        for data_point in data_points:
            self._mod_raw_split_cycle(data_point)
        logging.warning(
            f"splitting cycles at {data_points} -re-run make_step_table and make_summary to propagate change!"
        )

    def _mod_raw_split_cycle(self, data_point: int) -> None:
        r = self.data.raw

        hdr_data_point = self.headers_normal.data_point_txt
        hdr_cycle = self.headers_normal.cycle_index_txt
        hdr_c_cap = self.headers_normal.charge_capacity_txt
        hdr_d_cap = self.headers_normal.discharge_capacity_txt
        hdr_c_energy = self.headers_normal.charge_energy_txt
        hdr_d_energy = self.headers_normal.discharge_energy_txt

        # modifying cycle numbers
        c_mask = r[hdr_data_point] >= data_point
        r.loc[c_mask, hdr_cycle] = r.loc[c_mask, hdr_cycle] + 1

        # resetting capacities
        initial_values = r.loc[r[hdr_data_point] == data_point - 1, :]
        cycle = r.loc[r[hdr_data_point] == data_point, hdr_cycle].values[0]

        c_cap, d_cap, c_energy, d_energy = initial_values[
            [hdr_c_cap, hdr_d_cap, hdr_c_energy, hdr_d_energy]
        ].values[0]
        cycle_mask = r[hdr_cycle] == cycle
        r.loc[cycle_mask, hdr_c_cap] = r.loc[cycle_mask, hdr_c_cap] - c_cap
        r.loc[cycle_mask, hdr_d_cap] = r.loc[cycle_mask, hdr_d_cap] - d_cap
        r.loc[cycle_mask, hdr_c_energy] = r.loc[cycle_mask, hdr_c_energy] - c_energy
        r.loc[cycle_mask, hdr_d_energy] = r.loc[cycle_mask, hdr_d_energy] - d_energy

    def split(self, cycle=None):
        """Split experiment (CellpyCell object) into two sub-experiments. if cycle
        is not give, it will split on the median cycle number"""

        if isinstance(cycle, int) or cycle is None:
            return self.split_many(base_cycles=cycle)

    def drop_from(self, cycle=None):
        """Select first part of experiment (CellpyCell object) up to cycle number
        'cycle'"""

        if isinstance(cycle, int):
            c1, c2 = self.split_many(base_cycles=cycle)
            return c1

    def drop_to(self, cycle=None):
        """Select last part of experiment (CellpyCell object) from cycle number
        'cycle'"""

        if isinstance(cycle, int):
            c1, c2 = self.split_many(base_cycles=cycle)
            return c2

    def drop_edges(self, start, end):
        """Select middle part of experiment (CellpyCell object) from cycle
        number 'start' to 'end'"""

        if end < start:
            raise ValueError("end cannot be larger than start")
        if end == start:
            raise ValueError("end cannot be the same as start")
        return self.split_many([start, end])[1]

    def split_many(self, base_cycles=None):
        """Split experiment (CellpyCell object) into several sub-experiments.

        Args:
            base_cycles (int or list of ints): cycle(s) to do the split on.

        Returns:
            List of CellpyCell objects

        """
        h_summary_index = HEADERS_SUMMARY.cycle_index
        h_raw_index = HEADERS_NORMAL.cycle_index_txt
        h_step_cycle = HEADERS_STEP_TABLE.cycle

        if base_cycles is None:
            all_cycles = self.get_cycle_numbers()
            base_cycles = int(np.median(all_cycles))

        cells = list()
        if not isinstance(base_cycles, (list, tuple)):
            base_cycles = [base_cycles]

        dataset = self.data
        steptable = dataset.steps
        data = dataset.raw
        summary = dataset.summary

        # In case Cycle_Index has been promoted to index [#index]
        if h_summary_index not in summary.columns:
            summary = summary.reset_index(drop=False)

        for b_cycle in base_cycles:
            steptable0, steptable = [
                steptable[steptable[h_step_cycle] < b_cycle],
                steptable[steptable[h_step_cycle] >= b_cycle],
            ]
            data0, data = [
                data[data[h_raw_index] < b_cycle],
                data[data[h_raw_index] >= b_cycle],
            ]
            summary0, summary = [
                summary[summary[h_summary_index] < b_cycle],
                summary[summary[h_summary_index] >= b_cycle],
            ]

            new_cell = CellpyCell.vacant(cell=self)
            old_cell = CellpyCell.vacant(cell=self)

            summary0 = summary0.set_index(h_summary_index)

            new_cell.data.steps = steptable0
            new_cell.data.raw = data0
            new_cell.data.summary = summary0
            new_cell.data = identify_last_data_point(new_cell.data)

            old_cell.data.steps = steptable
            old_cell.data.raw = data
            old_cell.data.summary = summary
            old_cell.data = identify_last_data_point(old_cell.data)

            cells.append(new_cell)

        cells.append(old_cell)
        return cells

    def __register_external_readers(self):
        logging.debug(
            "Not implemented yet. Should allow registering readers "
            "for example installed as plug-ins."
        )
        self.__external_readers = dict()
        return

    def register_instrument_readers(self):
        """Register instrument readers."""

        self.instrument_factory = generate_default_factory()
        # instruments = find_all_instruments()
        # for instrument_id, instrument in instruments.items():
        #     self.instrument_factory.register_builder(instrument_id, instrument)

    def _set_raw_units(self):
        raw_units = get_default_raw_units()
        new_raw_units = self.loader_class.get_raw_units()
        for key in new_raw_units:
            if key in raw_units:
                raw_units[key] = new_raw_units[key]
            else:
                logging.debug(f"Got unconventional raw-unit label: {key}")
        return raw_units

    def _set_instrument(self, instrument, **kwargs):
        logging.debug(f"Setting new instrument: {instrument}")
        self.loader_class = self.instrument_factory.create(instrument, **kwargs)
        self.raw_limits = self.loader_class.get_raw_limits()
        # ----- create the loader ------------------------
        self.loader = self.loader_class.loader_executor

    def set_instrument(
        self,
        instrument=None,
        model=None,
        instrument_file=None,
        **kwargs,
    ):
        """Set the instrument (i.e. tell cellpy the file-type you use).

        Three different modes of setting instruments are currently supported. You can
        provide the already supported instrument names (see the documentation, e.g. "arbin_res").
        You can use the "custom" loader by providing the path to a yaml-file
        describing the file format. This can be done either by setting instrument to
        "instrument_name::instrument_definition_file_name", or by setting instrument to "custom" and
        provide the definition file name through the instrument_file keyword argument. A last option
        exists where you provide the yaml-file name directly to the instrument parameter. Cellpy
        will then look into your local instrument folder and search for the yaml-file. Some
        instrument types also supports a model key-word.

        Args:
            instrument: (str) in ["arbin_res", "maccor_txt",...]. If
                instrument ends with ".yml" a local instrument file will be used. For example,
                if instrument is "my_instrument.yml", cellpy will look into the local
                instruments folders for a file called "my_instrument.yml" and then
                use LocalTxtLoader to load after registering the instrument. If the instrument
                name contains a '::' separator, the part after the separator will be interpreted
                as 'instrument_file'.
            model: (str) optionally specify if the instrument loader supports handling several models
                (some instruments allow for exporting data in slightly different formats depending on
                the choices made during the export or the model of the instrument, e.g. different number of
                header lines, different encoding).
            instrument_file: (path) instrument definition file,
            kwargs (dict): key-word arguments sent to the initializer of the
                loader class

        Notes:
            If you are using a local instrument loader, you will have to register it first to the loader factory.

            >>> c = CellpyCell()  # this will automatically register the already implemented loaders
            >>> c.instrument_factory.register_builder(instrument_id, (module_name, path_to_instrument_loader_file))

            It is highly recommended using the module_name as the instrument_id.

        """

        # constants:
        custom_instrument_splitter = "::"
        # consume keyword arguments:
        _override_local_instrument_path = kwargs.pop(
            "_override_local_instrument_path", False
        )

        # parse input (need instrument, instrument_file and model)
        if instrument is None and instrument_file is None:
            instrument = self.tester

        if not instrument_file:
            instrument, instrument_file = self._parse_instrument_str(
                instrument, custom_instrument_splitter
            )

        if instrument_file and not model:
            instrument, model = self._parse_instrument_str(
                instrument, custom_instrument_splitter
            )

        if instrument and instrument.endswith(".yml"):
            instrument_file = instrument
            instrument = "local_instrument"
            prms.Instruments.custom_instrument_definitions_file = instrument_file
            if _override_local_instrument_path:
                instrument_file = Path(instrument_file)
            else:
                instrument_file = Path(prms.Paths.instrumentdir) / instrument_file

            if not instrument_file.is_file():
                raise FileNotFoundError(f"Could not locate {instrument_file}")

        self._set_instrument(
            instrument, instrument_file=instrument_file, model=model, **kwargs
        )

    @staticmethod
    def _parse_instrument_str(instrument, custom_instrument_splitter="::"):
        if not instrument:
            return None, None

        _instrument = instrument.split(custom_instrument_splitter)
        if len(_instrument) < 2:
            return instrument, None

        return _instrument

    @property
    def cycle_mode(self):
        # TODO: v2.0 edit this from scalar to list
        try:
            data = self.data
            m = data.meta_test_dependent.cycle_mode
            # cellpy saves this as a list (ready for v2.0),
            # but we want to return a scalar for the moment
            # Temporary fix to make sure that cycle_mode is a scalar:
            if isinstance(m, (tuple, list)):
                return m[0]
            return m
        except NoDataFound:
            return self._cycle_mode

    @cycle_mode.setter
    def cycle_mode(self, cycle_mode):
        # TODO: v2.0 edit this from scalar to list
        logging.debug(f"-> cycle_mode: {cycle_mode}")
        try:
            data = self.data
            data.meta_test_dependent.cycle_mode = cycle_mode
            self._cycle_mode = cycle_mode
        except NoDataFound:
            self._cycle_mode = cycle_mode

    def set_raw_datadir(self, directory=None):
        """Set the directory containing .res-files.

        Used for setting directory for looking for res-files.@
        A valid directory name is required.

        Args:
            directory (str): path to res-directory

        Example:
            >>> d = CellpyCell()
            >>> directory = "MyData/Arbindata"
            >>> d.set_raw_datadir(directory)

        """

        if directory is None:
            logging.info("No directory name given")
            return
        if not os.path.isdir(directory):
            logging.info(directory)
            logging.info("Directory does not exist")
            return
        self.raw_datadir = directory

    def set_cellpy_datadir(self, directory=None):
        """Set the directory containing .hdf5-files.

        Used for setting directory for looking for hdf5-files.
        A valid directory name is required.

        Args:
            directory (str): path to hdf5-directory

        Example:
            >>> d = CellpyCell()
            >>> directory = "MyData/HDF5"
            >>> d.set_raw_datadir(directory)

        """

        if directory is None:
            logging.info("No directory name given")
            return
        if not os.path.isdir(directory):
            logging.info("Directory does not exist")
            return
        self.cellpy_datadir = directory

    def check_file_ids(self, rawfiles, cellpyfile, detailed=False):
        """Check the stats for the files (raw-data and cellpy hdf5).

        This function checks if the hdf5 file and the res-files have the same
        timestamps etc. to find out if we need to bother to load .res -files.

        Args:
            cellpyfile (str): filename of the cellpy hdf5-file.
            rawfiles (list of str): name(s) of raw-data file(s).
            detailed (bool): return a dict containing True or False for each individual raw-file.

        Returns:
            If detailed is False:
                False if the raw files are newer than the cellpy hdf5-file
                (update needed). True if update is not needed.
             If detailed is True it returns a dict containing True or False for each
                individual raw-file.
        """

        txt = f"Checking file ids - using '{self.filestatuschecker}'"
        logging.info(txt)

        ids_cellpy_file = self._check_cellpy_file(cellpyfile)

        logging.debug(f"cellpyfile ids: {ids_cellpy_file}")

        if not ids_cellpy_file:
            # logging.debug("hdf5 file does not exist - needs updating")
            return False

        ids_raw = self._check_raw(rawfiles)

        if detailed:
            similar = self._parse_ids(ids_raw, ids_cellpy_file)
            return similar

        else:
            similar = self._compare_ids(ids_raw, ids_cellpy_file)
            if not similar:
                # logging.debug("hdf5 file needs updating")
                return False
            else:
                # logging.debug("hdf5 file is updated")
                return True

    def _check_raw(self, file_names, abort_on_missing=False):
        """Get the file-ids for the res_files."""

        strip_file_names = True
        check_on = self.filestatuschecker
        if not self._is_listtype(file_names):
            file_names = [file_names]

        ids = dict()
        for f in file_names:
            logging.debug(f"checking raw file {f}")
            fid = FileID(f)
            # logging.debug(fid)
            if fid.name is None:
                warnings.warn(f"file does not exist: {f}")
                if abort_on_missing:
                    sys.exit(-1)
            else:
                if strip_file_names:
                    name = f.name
                else:
                    name = f
                if check_on == "size":
                    ids[name] = int(fid.size)
                elif check_on == "modified":
                    ids[name] = int(fid.last_modified)
                else:
                    ids[name] = int(fid.last_modified)
        return ids

    def _check_cellpy_file(self, filename: OtherPath):
        """Get the file-ids for the cellpy_file."""

        if not isinstance(filename, OtherPath):
            logging.debug("filename must be an OtherPath object")
            filename = OtherPath(filename)

        use_full_filename_path = False
        parent_level = prms._cellpyfile_root  # noqa
        fid_dir = prms._cellpyfile_fid  # noqa
        check_on = self.filestatuschecker
        logging.debug("checking cellpy-file")
        logging.debug(filename)
        if not filename.is_file():
            logging.debug("cellpy-file does not exist")
            return None
        try:
            # TODO: implement external handling of hdf5-files
            if filename.is_external:
                # I have not implemented any external handling of hdf5-files yet. So we need to
                # copy the file to temporary directory (this will take some time, and therefore it is
                # probably best not to put your cellpy files in a remote directory yet):
                filename = filename.copy()
            store = pd.HDFStore(filename)
        except Exception as e:
            logging.debug(f"could not open cellpy-file ({e})")
            return None
        fidtable = None
        try:
            fidtable = store.select(parent_level + fid_dir)
        except KeyError:
            logging.warning("no fidtable - you should update your hdf5-file")
        except NotImplementedError:
            logging.warning(
                "your system cannot read the fid-table (posix-windows confusion) "
                "hopefully this will be solved in a newer version of pytables."
            )
        finally:
            store.close()
        if fidtable is not None:
            raw_data_files, raw_data_files_length = self._convert2fid_list(fidtable)
            txt = "contains %i res-files" % (len(raw_data_files))
            logging.debug(txt)
            ids = dict()
            for fid in raw_data_files:
                full_name = fid.full_name
                name = fid.name
                size = fid.size
                mod = fid.last_modified
                logging.debug(f"fileID information for: {full_name}")
                logging.debug(f"   modified: {mod}")
                logging.debug(f"   size: {size}")

                if use_full_filename_path:
                    name = full_name

                if check_on == "size":
                    ids[name] = int(fid.size)
                elif check_on == "modified":
                    ids[name] = int(fid.last_modified)
                else:
                    ids[name] = int(fid.last_modified)
            return ids
        else:
            return None

    @staticmethod
    def _compare_ids(ids_raw, ids_cellpy_file):
        similar = True
        l_res = len(ids_raw)
        l_cellpy = len(ids_cellpy_file)
        if l_res == l_cellpy and l_cellpy > 0:
            for name, value in list(ids_raw.items()):
                try:
                    c_value = ids_cellpy_file[name]
                except KeyError:
                    logging.debug("KeyError when comparing raw and cellpy file.")
                    logging.debug(
                        "Could be due to upper case vs. lower case confusion."
                    )
                    similar = False
                else:
                    if c_value != value:
                        similar = False
        else:
            similar = False

        return similar

    @staticmethod
    def _parse_ids(ids_raw, ids_cellpy_file):
        similar = dict()
        for name in ids_raw:
            v_cellpy = ids_cellpy_file.get(name, None)
            v_raw = ids_raw[name]
            similar[name] = False
            if v_raw is not None:
                if v_raw == v_cellpy:
                    similar[name] = True
        return similar

    # TODO: v2.0 - remove this
    def loadcell(
        self,
        raw_files,
        cellpy_file=None,
        mass=None,
        summary_on_raw=True,
        summary_on_cellpy_file=True,
        find_ir=True,
        find_end_voltage=True,
        force_raw=False,
        use_cellpy_stat_file=None,
        cell_type=None,
        loading=None,
        area=None,
        estimate_area=True,
        selector=None,
        **kwargs,
    ):
        """Loads data for given cells.

        Args:
            raw_files (list): name of res-files
            cellpy_file (path): name of cellpy-file
            mass (float): mass of electrode or active material
            summary_on_raw (bool): calculate summary if loading from raw
            summary_on_cellpy_file (bool): calculate summary if loading from cellpy-file.
            find_ir (bool): summarize ir
            find_end_voltage (bool): summarize end voltage
            force_raw (bool): only use raw-files
            use_cellpy_stat_file (bool): use stat file if creating summary
                from raw
            cell_type (str): set the data type (e.g. "anode"). If not, the default from
               the config file is used.
            loading (float): loading in units [mass] / [area], used to calculate area if area not given
            area (float): area of active electrode
            estimate_area (bool): calculate area from loading if given (defaults to True).
            selector (dict): passed to load.
            **kwargs: passed to from_raw

        Example:

            >>> srnos = my_dbreader.select_batch("testing_new_solvent")
            >>> cell_datas = []
            >>> for srno in srnos:
            >>> ... my_run_name = my_dbreader.get_cell_name(srno)
            >>> ... mass = my_dbreader.get_mass(srno)
            >>> ... rawfiles, cellpyfiles = \
            >>> ...     filefinder.search_for_files(my_run_name)
            >>> ... cell_data = cellreader.CellpyCell()
            >>> ... cell_data.loadcell(raw_files=rawfiles,
            >>> ...                    cellpy_file=cellpyfiles)
            >>> ... cell_data.set_mass(mass)
            >>> ... cell_data.make_summary() # etc. etc.
            >>> ... cell_datas.append(cell_data)
            >>>
        """
        # This is a part of a dramatic API change. It will not be possible to
        # load more than one set of datasets (i.e. one single cellpy-file or
        # several raw-files that will be automatically merged)

        # TODO @jepe Make setting or prm so that it is possible to update only new data
        # TODO @jepe Allow passing handle to progress-bar or update a global progressbar

        warnings.warn(
            DeprecationWarning("loadcell is deprecated. Use cellpy.get instead.")
        )
        logging.debug("Started cellpy.cellreader.loadcell ")

        if cellpy_file is None:
            similar = False
        elif force_raw:
            similar = False
        else:
            similar = self.check_file_ids(raw_files, cellpy_file)
            logging.debug(f"checked if the files were similar")
        logging.debug(f"similar: {similar}")

        if similar:
            logging.debug(f"loading cellpy-file: {cellpy_file}")
            self.load(cellpy_file, selector=selector)

        else:
            logging.debug("cellpy file(s) needs updating - loading raw")
            logging.info("Loading raw-file")
            logging.debug(raw_files)

            self.from_raw(raw_files, **kwargs)

        logging.debug("loaded files")

        if not self._validate_cell():
            logging.warning("Empty run!")
            return self

        logging.debug("setting cell_type")
        if cell_type is not None:
            self.cycle_mode = cell_type
            logging.debug(f"setting cycle mode: {cell_type}")

        logging.debug("setting mass")
        if mass is not None:
            self.set_mass(mass)

        logging.debug("setting nom_cap")
        nom_cap = kwargs.pop("nom_cap", None)
        if nom_cap is not None:
            self.set_nom_cap(nom_cap)

        logging.debug("calculating area")
        if area is not None:
            logging.debug(f"got area: {area}")
            self.data.meta_common.active_electrode_area = area
        elif loading and estimate_area:
            logging.debug(f"got loading: {logging}")
            area = self.data.mass / loading
            logging.debug(
                f"calculating area from loading ({loading}) and mass ({self.data.mass}): {area}"
            )
            self.data.meta_common.active_electrode_area = area
        else:
            logging.debug("using default area")

        if similar:
            if summary_on_cellpy_file:
                self.make_summary(
                    find_ir=find_ir,
                    find_end_voltage=find_end_voltage,
                    use_cellpy_stat_file=use_cellpy_stat_file,
                )

        else:
            if summary_on_raw:
                self.make_summary(
                    find_ir=find_ir,
                    find_end_voltage=find_end_voltage,
                    use_cellpy_stat_file=use_cellpy_stat_file,
                )

        return self

    def from_raw(
        self,
        file_names=None,
        pre_processor_hook=None,
        post_processor_hook=None,
        is_a_file=True,
        refuse_copying=False,
        **kwargs,
    ):
        """Load a raw data-file.

        Args:
            file_names (list of raw-file names): uses CellpyCell.file_names if
                None. If the list contains more than one file name, then the
                runs will be merged together. Remark! the order of the files in
                the list is important.
            pre_processor_hook (callable): function that will be applied to the data within the loader.
            post_processor_hook (callable): function that will be applied to the
                cellpy.Dataset object after initial loading.
            is_a_file (bool): set this to False if it is a not a file-like object.
            refuse_copying (bool): if set to True, the raw-file will not be copied before loading.

        Keyword Args for merging:
            recalc (bool): set to false if you don't want cellpy to automatically shift cycle number
                and time (e.g. add last cycle number from previous file to the cycle numbers
                in the next file).

        Other keywords depending on loader:
            [ArbinLoader]:
                bad_steps (list of tuples): (c, s) tuples of steps s (in cycle c)
                    to skip loading.
                data_points (tuple of ints): load only data from data_point[0] to
                    data_point[1] (use None for infinite). NOT IMPLEMENTED YET.

        """
        if file_names:
            self.file_names = file_names

        if not isinstance(self.file_names, (list, tuple)):
            self.file_names = [file_names]

        # file_type = self.tester
        instrument = kwargs.pop("instrument", None)
        instrument_file = kwargs.pop("instrument_file", None)
        if instrument_file:
            logging.info("Setting custom instrument")
            logging.info(f"-> {instrument}")
            logging.info(f"-> instrument file: {instrument_file}")
            self.set_instrument(instrument="custom", instrument_file=instrument_file)
        elif instrument:
            logging.info("Setting custom instrument")
            logging.info(f"-> {instrument}")
            self.set_instrument(instrument)

        raw_file_loader = self.loader
        try:
            self.tester = self.loader_class.instrument_name
        except AttributeError:
            logging.debug(f"could not set instrument name")

        # TODO: include this into prms (and config-file):
        max_raw_files_to_merge = 20
        if len(self.file_names) > max_raw_files_to_merge:
            logging.debug("ERROR? Too many files to merge")
            raise ValueError(
                f"Too many files to merge (max allowed is {max_raw_files_to_merge})"
                f" - could be a bug in the code (please report if you know you have"
                f" less than {max_raw_files_to_merge} files)"
            )

        logging.debug("start iterating through file(s)")
        recalc = kwargs.pop("recalc", True)
        data = None
        for file_name in self.file_names:
            logging.debug("loading raw file:")
            logging.debug(f"{file_name}")
            if is_a_file:
                file_name = OtherPath(file_name)
                if not file_name.is_file():
                    raise NoDataFound(f"Could not find the file {file_name}")

            new_data = raw_file_loader(
                file_name,
                pre_processor_hook=pre_processor_hook,
                refuse_copying=refuse_copying,
                **kwargs,
            )  # list of tests

            if new_data is None:
                raise IOError(
                    f"Could not read {file_name}. Loader returned None. Aborting."
                )
            if not new_data.has_data:
                raise IOError(f"Could not read any data from {file_name}. Aborting.")

            if post_processor_hook is not None:
                # REMARK! this needs to be changed if we stop returning the datasets in a list
                # (i.e. if we chose to remove option for having more than one test pr instance)
                new_data = post_processor_hook(new_data)

            if data is None:
                # retrieving the first cell data (e.g. first file)
                logging.debug("getting data from first file")
                data = new_data
            else:
                # appending cell data file to existing
                logging.debug("continuing reading files...")
                data = self._append(data, new_data, recalc=recalc)

                # retrieving file info in a for-loop in case of multiple files
                # Remark!
                #    - the raw_data_files attribute is a list
                #    - the raw_data_files_length attribute is a list

                logging.debug("added the data set - merging file info")

                data.raw_data_files.extend(new_data.raw_data_files)
                data.raw_data_files_length.extend(new_data.raw_data_files_length)

        logging.debug("finished loading the raw-files")

        if not prms.Reader.sorted_data:
            logging.debug("sorting data")
            data = self._sort_data(data)
            data.raw_units = self._set_raw_units()

        self.data = data
        self._invent_a_cell_name(self.file_names)  # TODO (v1.0.0): fix me
        return self

    def _validate_cell(self, level=0):
        logging.debug("validating test")
        # simple validation for finding empty datasets - should be expanded to
        # find not-complete datasets, datasets with missing parameters etc
        v = True
        if level == 0:
            try:
                data = self.data
                return True
            except NoDataFound:
                return False
        return v

    def partial_load(self, **kwargs):
        """Load only a selected part of the cellpy file."""
        raise NotImplementedError

    def link(self, **kwargs):
        """Create a link to a cellpy file.

        If the file is very big, it is sometimes better to work with the data
        out of memory (i.e. on disk). A CellpyCell object with a linked file
        will in most cases work as a normal object. However, some methods
        might be disabled. And it will be slower.

        Notes:
            2020.02.08 - maybe this functionality is not needed and can be replaced
                by using dask or similar?
        """
        raise NotImplementedError

    def load(
        self,
        cellpy_file,
        parent_level=None,
        return_cls=True,
        accept_old=True,
        selector=None,
        **kwargs,
    ):
        """Loads a cellpy file.

        Args:
            cellpy_file (OtherPath, str): Full path to the cellpy file.
            parent_level (str, optional): Parent level. Warning! Deprecating this soon!
            return_cls (bool): Return the class.
            accept_old (bool): Accept loading old cellpy-file versions.
                Instead of raising WrongFileVersion it only issues a warning.
            selector (): under development

        Returns:
            cellpy.CellPyCellpy class if return_cls is True
        """

        # This is what happens:
        # 1) (this is not implemented yet, using only hdf5) chose what file format to load from
        # 2) in reader (currently only _load_hdf5): check version and select sub-reader.
        # 3) in sub-reader: read data
        # 4) in this method: add data to CellpyCell object (i.e. self)
        for kwarg in kwargs:
            logging.debug(f"received (still) un-supported keyword argument {kwarg=}")

        try:
            logging.debug("loading cellpy-file (hdf5):")
            logging.debug(cellpy_file)
            logging.debug(f"{type(cellpy_file)=}")
            cellpy_file = OtherPath(cellpy_file)
            with pickle_protocol(PICKLE_PROTOCOL):
                logging.debug(f"using pickle protocol {PICKLE_PROTOCOL}")
                data = self._load_hdf5(
                    cellpy_file, parent_level, accept_old, selector=selector
                )
            logging.debug("cellpy-file loaded")

        except AttributeError:
            data = None
            logging.warning(
                "This cellpy-file version is not supported by"
                "current reader (try to update cellpy)."
            )

        if data:
            self.data = data
        else:
            # raise LoadError
            logging.warning("Could not load")
            logging.warning(str(cellpy_file))

        self._invent_a_cell_name(cellpy_file)
        if return_cls:
            return self

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _get_cellpy_file_version(self, filename, meta_dir=None, parent_level=None):
        if meta_dir is None:
            meta_dir = prms._cellpyfile_common_meta

        if parent_level is None:
            parent_level = prms._cellpyfile_root

        with pd.HDFStore(filename) as store:
            try:
                meta_table = store.select(parent_level + meta_dir)
            except KeyError:
                raise WrongFileVersion(
                    "This file is VERY old - cannot read file version number"
                )
        try:
            # cellpy_file_version = self._extract_from_dict(
            #     meta_table, "cellpy_file_version"
            # )
            meta_dict = meta_table.to_dict(orient="list")
            cellpy_file_version = self._extract_from_meta_dictionary(
                meta_dict, "cellpy_file_version"
            )
        except Exception as e:
            warnings.warn(f"Unhandled exception raised: {e}")
            return 0

        return cellpy_file_version

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _load_hdf5(self, filename, parent_level=None, accept_old=False, selector=None):
        """Load a cellpy-file.

        Args:
            filename (str): Name of the cellpy file.
            parent_level (str) (optional): name of the parent level
                (defaults to "CellpyData").
            accept_old (bool): accept old file versions.
            selector (): select specific ranges (under development)

        Returns:
            loaded datasets (DataSet-object)
        """

        if parent_level is None:
            parent_level = prms._cellpyfile_root

        if parent_level != prms._cellpyfile_root:
            logging.debug(
                f"Using non-default parent label for the " f"hdf-store: {parent_level}"
            )

        if not os.path.isfile(filename):
            logging.info(f"File does not exist: {filename}")
            raise IOError(f"File does not exist: {filename}")

        cellpy_file_version = self._get_cellpy_file_version(filename)
        logging.debug(f"Cellpy file version {cellpy_file_version}; selector={selector}")

        if cellpy_file_version > CELLPY_FILE_VERSION:
            raise WrongFileVersion(
                f"File format too new: {filename} :: version: {cellpy_file_version}"
                f"Reload from raw or upgrade your cellpy!"
            )

        elif cellpy_file_version < MINIMUM_CELLPY_FILE_VERSION:
            raise WrongFileVersion(
                f"File format too old: {filename} :: version: {cellpy_file_version}"
                f"Reload from raw or downgrade your cellpy!"
            )

        elif cellpy_file_version < CELLPY_FILE_VERSION:
            if accept_old:
                logging.debug(f"old cellpy file version {cellpy_file_version}")
                logging.debug(f"filename: {filename}")
                logging.warning(
                    f"Loading old file-type. It is recommended that you remake the step table and the "
                    f"summary table."
                )
                new_data = self._load_old_hdf5(filename, cellpy_file_version)
            else:
                raise WrongFileVersion(
                    f"File format too old: {filename} :: version: {cellpy_file_version}"
                    f"Try loading setting accept_old=True"
                )

        else:
            logging.debug(f"Loading {filename} :: v{cellpy_file_version}")
            new_data = self._load_hdf5_current_version(filename, selector=selector)

        # self.__check_loaded_data(new_data)

        return new_data

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _load_hdf5_current_version(self, filename, parent_level=None, selector=None):
        if parent_level is None:
            parent_level = prms._cellpyfile_root

        raw_dir = prms._cellpyfile_raw
        step_dir = prms._cellpyfile_step
        summary_dir = prms._cellpyfile_summary
        fid_dir = prms._cellpyfile_fid
        common_meta_dir = prms._cellpyfile_common_meta
        test_dependent_meta_dir = prms._cellpyfile_test_dependent_meta

        logging.debug(f"filename: {filename}")
        logging.debug(f"selector: {selector}")
        with pd.HDFStore(filename) as store:
            (
                data,
                meta_table,
                test_dependent_meta_table,
            ) = self._create_initial_data_set_from_cellpy_file(
                common_meta_dir,
                parent_level,
                store,
                test_dependent_meta_dir=test_dependent_meta_dir,
            )
            self._check_keys_in_cellpy_file(
                common_meta_dir, parent_level, raw_dir, store, summary_dir
            )
            self._extract_summary_from_cellpy_file(
                data, parent_level, store, summary_dir, selector=selector
            )
            self._extract_raw_from_cellpy_file(
                data, parent_level, raw_dir, store, selector=selector
            )
            self._extract_steps_from_cellpy_file(
                data, parent_level, step_dir, store, selector=selector
            )
            fid_table, fid_table_selected = self._extract_fids_from_cellpy_file(
                fid_dir, parent_level, store
            )

        self._extract_meta_from_cellpy_file(
            data, meta_table, test_dependent_meta_table, filename
        )

        if fid_table_selected:
            (
                data.raw_data_files,
                data.raw_data_files_length,
            ) = self._convert2fid_list(fid_table)
        else:
            data.raw_data_files = []
            data.raw_data_files_length = []
        return data

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _load_hdf5_v7(self, filename, selector=None, **kwargs):
        logging.debug("--- loading v7")
        meta_dir = "/info"
        parent_level = kwargs.pop("parent_level", "CellpyData")
        raw_dir = kwargs.pop("raw_dir", "/raw")
        step_dir = kwargs.pop("step_dir", "/steps")
        summary_dir = kwargs.pop("summary_dir", "/summary")
        fid_dir = kwargs.pop("fid_dir", "/fid")

        logging.debug(f"filename: {filename}")
        logging.debug(f"selector: {selector}")

        with pd.HDFStore(filename) as store:
            data, meta_table = self._create_initial_data_set_from_cellpy_file(
                meta_dir, parent_level, store
            )
            self._check_keys_in_cellpy_file(
                meta_dir, parent_level, raw_dir, store, summary_dir
            )
            self._extract_summary_from_cellpy_file(
                data, parent_level, store, summary_dir, selector=selector
            )
            self._extract_raw_from_cellpy_file(
                data, parent_level, raw_dir, store, selector=selector
            )
            self._extract_steps_from_cellpy_file(
                data, parent_level, step_dir, store, selector=selector
            )
            fid_table, fid_table_selected = self._extract_fids_from_cellpy_file(
                fid_dir, parent_level, store
            )

        self._extract_meta_from_old_cellpy_file_max_v7(
            data, meta_table, filename, upgrade_from_to=(7, CELLPY_FILE_VERSION)
        )

        if fid_table_selected:
            (
                data.raw_data_files,
                data.raw_data_files_length,
            ) = self._convert2fid_list(fid_table)
        else:
            data.raw_data_files = []
            data.raw_data_files_length = []
        return data

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _load_hdf5_v6(self, filename, selector=None):
        logging.critical("--- loading v6")
        parent_level = "CellpyData"
        raw_dir = "/raw"
        step_dir = "/steps"
        summary_dir = "/summary"
        fid_dir = "/fid"
        meta_dir = "/info"

        with pd.HDFStore(filename) as store:
            data, meta_table = self._create_initial_data_set_from_cellpy_file(
                meta_dir,
                parent_level,
                store,
            )
            self._check_keys_in_cellpy_file(
                meta_dir, parent_level, raw_dir, store, summary_dir
            )
            self._extract_summary_from_cellpy_file(
                data,
                parent_level,
                store,
                summary_dir,
                selector=selector,
                upgrade_from_to=(6, CELLPY_FILE_VERSION),
            )
            self._extract_raw_from_cellpy_file(
                data,
                parent_level,
                raw_dir,
                store,
                selector=selector,
                upgrade_from_to=(6, CELLPY_FILE_VERSION),
            )
            self._extract_steps_from_cellpy_file(
                data,
                parent_level,
                step_dir,
                store,
                selector=selector,
            )
            fid_table, fid_table_selected = self._extract_fids_from_cellpy_file(
                fid_dir, parent_level, store
            )

        self._extract_meta_from_old_cellpy_file_max_v7(
            data, meta_table, filename, upgrade_from_to=(6, CELLPY_FILE_VERSION)
        )

        if fid_table_selected:
            (
                data.raw_data_files,
                data.raw_data_files_length,
            ) = self._convert2fid_list(fid_table)
        else:
            data.raw_data_files = []
            data.raw_data_files_length = []

        logging.debug("loaded new test")
        return data

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _load_hdf5_v5(self, filename, selector=None):
        logging.critical("--- loading v5")
        parent_level = "CellpyData"
        raw_dir = "/raw"
        step_dir = "/steps"
        summary_dir = "/summary"
        fid_dir = "/fid"
        meta_dir = "/info"

        with pd.HDFStore(filename) as store:
            data, meta_table = self._create_initial_data_set_from_cellpy_file(
                meta_dir, parent_level, store
            )
            self._check_keys_in_cellpy_file(
                meta_dir, parent_level, raw_dir, store, summary_dir
            )
            self._extract_summary_from_cellpy_file(
                data,
                parent_level,
                store,
                summary_dir,
                selector=selector,
                upgrade_from_to=(5, CELLPY_FILE_VERSION),
            )
            self._extract_raw_from_cellpy_file(
                data,
                parent_level,
                raw_dir,
                store,
                selector=selector,
                upgrade_from_to=(5, CELLPY_FILE_VERSION),
            )
            self._extract_steps_from_cellpy_file(
                data, parent_level, step_dir, store, selector=selector
            )
            fid_table, fid_table_selected = self._extract_fids_from_cellpy_file(
                fid_dir, parent_level, store
            )

        self._extract_meta_from_old_cellpy_file_max_v7(data, meta_table, filename)

        if fid_table_selected:
            (
                data.raw_data_files,
                data.raw_data_files_length,
            ) = self._convert2fid_list(fid_table)
        else:
            data.raw_data_files = []
            data.raw_data_files_length = []

        logging.debug("loaded new test")
        return data

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _load_old_hdf5(self, filename, cellpy_file_version):
        if cellpy_file_version < 5:
            data = self._load_old_hdf5_v3_to_v4(filename)
        elif cellpy_file_version == 5:
            data = self._load_hdf5_v5(filename)
        elif cellpy_file_version == 6:
            data = self._load_hdf5_v6(filename)
        elif cellpy_file_version == 7:
            data = self._load_hdf5_v7(filename)
        else:
            raise WrongFileVersion(f"version {cellpy_file_version} is not supported")

        # if cellpy_file_version < 6:
        #     logging.debug("legacy cellpy file version needs translation")
        #     # data.raw = cellpy_file_upgrade_settings()
        #     data.raw = rename_raw_columns(data.raw, old, new)
        #     # data = old_settings.translate_headers(data, cellpy_file_version)
        #     # self.__check_loaded_data(data)
        return data

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _load_old_hdf5_v3_to_v4(self, filename):
        logging.critical("--- loading v < 5")
        parent_level = "CellpyData"
        meta_dir = "/info"
        _raw_dir = "/dfdata"
        _step_dir = "/step_table"
        _summary_dir = "/dfsummary"
        _fid_dir = "/fidtable"

        with pd.HDFStore(filename) as store:
            data, meta_table = self._create_initial_data_set_from_cellpy_file(
                meta_dir, parent_level, store
            )

            self._check_keys_in_cellpy_file(
                meta_dir, parent_level, _raw_dir, store, _summary_dir
            )
            self._extract_summary_from_cellpy_file(
                data,
                parent_level,
                store,
                _summary_dir,
                upgrade_from_to=(4, CELLPY_FILE_VERSION),
            )
            self._extract_raw_from_cellpy_file(
                data,
                parent_level,
                _raw_dir,
                store,
                upgrade_from_to=(4, CELLPY_FILE_VERSION),
            )
            self._extract_steps_from_cellpy_file(
                data,
                parent_level,
                _step_dir,
                store,
                upgrade_from_to=(4, CELLPY_FILE_VERSION),
            )
            fid_table, fid_table_selected = self._extract_fids_from_cellpy_file(
                _fid_dir, parent_level, store
            )
        self._extract_meta_from_old_cellpy_file_max_v7(data, meta_table, filename)
        warnings.warn(
            "Loaded old cellpy-file version (<5). Please update and save again."
        )
        if fid_table_selected:
            (
                data.raw_data_files,
                data.raw_data_files_length,
            ) = self._convert2fid_list(fid_table)
        else:
            data.raw_data_files = []
            data.raw_data_files_length = []

        # new_tests = [data]
        # return new_tests
        return data

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _create_initial_data_set_from_cellpy_file(
        self, meta_dir, parent_level, store, test_dependent_meta_dir=None
    ):
        # Remark that this function is run before selecting loading method
        # based on version. If you change the common_meta_dir prm to something else than
        # "/info" it will most likely fail.
        # Remark! Used from versions 3
        if test_dependent_meta_dir is not None:
            common_meta_table = store.select(parent_level + meta_dir)
            test_dependent_meta = store.select(parent_level + test_dependent_meta_dir)
            data = Data()
            # data.cellpy_file_version = CELLPY_FILE_VERSION
            return data, common_meta_table, test_dependent_meta

        data = Data()
        meta_table = None

        try:
            meta_table = store.select(parent_level + meta_dir)
        except KeyError as e:
            logging.info("This file is VERY old - no info given here")
            logging.info("You should convert the files to a newer version!")
            logging.debug(e)
            return data, meta_table

        try:
            meta_dict = meta_table.to_dict(orient="list")
            # data.cellpy_file_version = self._extract_from_meta_dictionary(
            #     meta_dict, "cellpy_file_version"
            # )
        except Exception as e:
            # data.cellpy_file_version = 0
            warnings.warn(f"Unhandled exception raised: {e}")
            return data, meta_table

        # logging.debug(f"cellpy file version. {data.cellpy_file_version}")
        return data, meta_table

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    @staticmethod
    def _check_keys_in_cellpy_file(meta_dir, parent_level, raw_dir, store, summary_dir):
        required_keys = [raw_dir, summary_dir, meta_dir]
        required_keys = ["/" + parent_level + _ for _ in required_keys]

        for key in required_keys:
            if key not in store.keys():
                logging.info(
                    f"This cellpy-file is not good enough - "
                    f"at least one key is missing: {key}"
                )
                raise Exception(
                    f"OH MY GOD! At least one crucial key is missing {key}!"
                )
        logging.debug(f"Keys in current cellpy-file: {store.keys()}")

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _hdf5_locate_data_points_from_max_cycle_number(
        self, table_name, max_cycle, parent_level, store, child_level
    ):
        if table_name == prms._cellpyfile_step:
            _cycle_header = self.headers_step_table.cycle
            table_path = parent_level + child_level
        elif table_name == prms._cellpyfile_raw:
            _cycle_header = self.headers_normal.cycle_index_txt
            table_path = parent_level + child_level
        else:
            raise IOError(
                f"provided wrong table name: {table_name} "
                f"(valid options: ({prms._cellpyfile_step}, {prms._cellpyfile_raw}))"
            )

        cycles = store.select(table_path, where="columns=[_cycle_header]")
        return cycles[_cycle_header] <= max_cycle

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _hdf5_cycle_filter(self, table=None):
        # this is not the best way to do it
        if max_cycle := self.limit_loaded_cycles:
            if table == "summary":
                logging.debug(f"limited to cycle_number {max_cycle}")
                return f"index <= {int(max_cycle)}"
            elif table == "raw":
                # update this by finding the last data point
                #  by making a function setting self.limit_data_points
                logging.debug(f"limited to data_point {self.limit_data_points}")
                return f"index <= {int(self.limit_data_points)}"
            elif table == "steps":
                # update this by finding the last data point
                #  by making a function setting self.limit_data_points
                logging.debug(f"limited to data_point {self.limit_data_points}")
                return f"index <= {int(self.limit_data_points)}"

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _unpack_selector(self, selector):
        # not implemented yet
        # should be used for trimming the selector so that it is not necessary to parse it individually
        # for all the _extract_xxx_from_cellpy_file methods.
        return selector

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _extract_summary_from_cellpy_file(
        self,
        data: Data,
        parent_level: str,
        store: pd.HDFStore,
        summary_dir: str,
        selector: Union[None, str] = None,
        upgrade_from_to: tuple = None,
    ):
        if selector is not None:
            cycle_filter = []
            if max_cycle := selector.get("max_cycle", None):
                # self.overwrite_able = False

                cycle_filter.append(f"index <= {int(max_cycle)}")
                self.limit_loaded_cycles = max_cycle
        else:
            # getting cycle filter by setting attributes:
            self.limit_loaded_cycles = None
            cycle_filter = self._hdf5_cycle_filter("summary")

        data.summary = store.select(parent_level + summary_dir, where=cycle_filter)
        if upgrade_from_to is not None:
            old, new = upgrade_from_to
            logging.debug(f"upgrading from {old} to {new}")
            data.summary = rename_summary_columns(data.summary, old, new)

        # TODO: max data point should be an attribute
        try:
            max_data_point = data.summary[self.headers_summary.data_point].max()
        except KeyError as e:
            raise KeyError(
                f"You are most likely trying to open a too old cellpy file"
            ) from e

        self.limit_data_points = int(max_data_point)
        logging.debug(f"data-point max limit: {self.limit_data_points}")

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _extract_raw_from_cellpy_file(
        self,
        data,
        parent_level,
        raw_dir,
        store,
        selector: Union[None, str] = None,
        upgrade_from_to: tuple = None,
    ):
        # selector is not implemented yet for only raw data
        # however, selector for max_cycle will still work since
        # the attribute self.limit_data_points is set while reading the summary
        cycle_filter = self._hdf5_cycle_filter(table="raw")
        data.raw = store.select(parent_level + raw_dir, where=cycle_filter)
        if upgrade_from_to is not None:
            old, new = upgrade_from_to
            logging.debug(f"upgrading from {old} to {new}")
            data.raw = rename_raw_columns(data.raw, old, new)

    def _extract_steps_from_cellpy_file(
        self,
        data,
        parent_level,
        step_dir,
        store,
        selector: Union[None, str] = None,
        upgrade_from_to: tuple = None,
    ):
        try:
            data.steps = store.select(parent_level + step_dir)
            if self.limit_data_points:
                data.steps = data.steps.loc[
                    data.steps["point_last"] <= self.limit_data_points
                ]
                logging.debug(f"limited to data_point {self.limit_data_points}")
            if upgrade_from_to is not None:
                old, new = upgrade_from_to
                logging.debug(f"upgrading from {old} to {new}")
                data.steps = rename_step_columns(data.steps, old, new)
        except Exception as e:
            print(e)
            logging.debug("could not get steps from cellpy-file")
            data.steps = pd.DataFrame()
            warnings.warn(f"Unhandled exception raised: {e}")

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _extract_fids_from_cellpy_file(
        self, fid_dir, parent_level, store, upgrade_from_to: tuple = None
    ):
        logging.debug(f"Extracting fid table from {fid_dir} in hdf5 store")
        try:
            fid_table = store.select(
                parent_level + fid_dir
            )  # remark! changed spelling from
            # lower letter to camel-case!
            fid_table_selected = True
            if upgrade_from_to is not None:
                old, new = upgrade_from_to
                logging.debug(f"upgrading from {old} to {new}")
                fid_table = rename_fid_columns(fid_table, old, new)
        except Exception as e:
            logging.debug(e)
            logging.debug("could not get fid from cellpy-file")
            fid_table = []
            warnings.warn("no fid_table - you should update your cellpy-file")
            fid_table_selected = False
        return fid_table, fid_table_selected

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _extract_meta_from_cellpy_file(
        self,
        data: Data,
        meta_table: pd.DataFrame,
        test_dependent_meta_table: pd.DataFrame,
        filename: Union[Path, str],
        upgrade_from_to: tuple = None,
    ) -> None:
        if upgrade_from_to is not None:
            old, new = upgrade_from_to
            print(f"upgrading meta from {old} to {new}")
            logging.debug(f"upgrading meta from {old} to {new}")
            # fid_table = rename_fid_columns(fid_table, old, new)

        data.loaded_from = str(filename)
        meta_dict = meta_table.to_dict(orient="list")

        # unpacking the raw data limits
        # remark! stored as scalars (not test dependent)
        for key in data.raw_limits:
            h5_key = f"{prms._cellpyfile_raw_limit_pre_id}{key}"
            try:
                v = meta_dict.pop(h5_key)
                data.raw_units[key] = v[0]
            except KeyError:
                logging.debug(f"missing key in meta_table: {h5_key}")
                # warnings.warn("OLD-TYPE: Recommend to save in new format!")

        # unpacking the raw data units
        # remark! stored as scalars (not test dependent)
        for key in data.raw_units:
            h5_key = f"{prms._cellpyfile_raw_unit_pre_id}{key}"
            try:
                v = meta_dict.pop(h5_key)
                data.raw_units[key] = v[0]
            except KeyError:
                logging.critical(f"missing key in meta_table: {h5_key}")
                # warnings.warn("OLD-TYPE: Recommend to save in new format!")

        data.meta_common.update(as_list=False, **meta_dict)
        test_dependent_meta_dict = test_dependent_meta_table.to_dict(orient="list")
        data.meta_test_dependent.update(as_list=True, **test_dependent_meta_dict)

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    def _extract_meta_from_old_cellpy_file_max_v7(
        self,
        data: Data,
        meta_table: pd.DataFrame,
        filename: Union[Path, str],
        upgrade_from_to: tuple,
    ) -> None:
        # get attributes from meta table
        # remark! could also utilise the pandas to dictionary method directly
        # for example: meta_table.T.to_dict()
        # Maybe a good task for someone who would like to learn more about
        # how cellpy works.

        old, new = upgrade_from_to
        logging.debug(f"upgrading meta from {old} to {new}")
        if old > 7:
            raise IOError("using this method for processing v>7 is not allowed!")

        meta_dict = meta_table.to_dict(orient="list")

        # unpacking the raw data limits
        # remark! stored as scalars (not test dependent)
        for key in data.raw_limits:
            h5_key = f"{prms._cellpyfile_raw_limit_pre_id}{key}"
            try:
                v = meta_dict.pop(h5_key)
                data.raw_units[key] = v[0]
            except KeyError:
                logging.debug(f"missing key in meta_table: {h5_key}")
                # warnings.warn("OLD-TYPE: Recommend to save in new format!")

        # unpacking the raw data units
        # remark! stored as scalars (not test dependent)
        for key in data.raw_units:
            h5_key = f"{prms._cellpyfile_raw_unit_pre_id}{key}"
            try:
                v = meta_dict.pop(h5_key)
                v = v[0]
                if not isinstance(v, str):
                    logging.debug(f"{v} is not of type string")
                    v = convert_from_simple_unit_label_to_string_unit_label(key, v)
                data.raw_units[key] = v
            except KeyError:
                logging.critical(f"missing key in meta_table: {h5_key}")
                # warnings.warn("OLD-TYPE: Recommend to save in new format!")

        meta_dict = data.meta_common.digest(as_list=False, **meta_dict)
        data.meta_test_dependent.update(as_list=True, **meta_dict)

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    @staticmethod
    def _extract_from_meta_dictionary(
        meta_dict, attribute, default_value=None, hard=False
    ):
        try:
            value = meta_dict[attribute][0]
            if not value:
                value = None
        except KeyError as e:
            if hard:
                raise KeyError from e
            value = default_value
        return value

    # TODO @jepe: move this to its own module (e.g. as a cellpy-exporters?):
    def _create_infotable(self):
        # needed for saving class/DataSet to hdf5
        cell = self.data
        new_info_table = asdict(cell.meta_common)
        new_info_table_test_dependent = asdict(cell.meta_test_dependent)
        new_info_table["cellpy_file_version"] = CELLPY_FILE_VERSION

        limits = cell.raw_limits
        for key in limits:
            h5_key = f"{prms._cellpyfile_raw_limit_pre_id}{key}"
            new_info_table[key] = limits[h5_key]

        units = cell.raw_units
        for key in units:
            h5_key = f"{prms._cellpyfile_raw_unit_pre_id}{key}"
            value = units[key]
            if not isinstance(value, str):
                raise IOError(
                    f"raw unit for {key} ({value}) must be of type string, not {type(value)}"
                )
            new_info_table[h5_key] = value

        new_info_table = pd.DataFrame.from_records([new_info_table])
        new_info_table_test_dependent = pd.DataFrame.from_records(
            [new_info_table_test_dependent]
        )

        fidtable = self._convert2fid_table(cell)
        fidtable = pd.DataFrame(fidtable)
        # TODO: test_dependent with several tests (and possibly merge with FID)
        # TODO: save
        # TODO: load old
        # TODO: find out if it is possible to initiate dataclasses with **kwargs (for loading)
        # TODO: update getters and setters (cell_name etc)
        return new_info_table, new_info_table_test_dependent, fidtable

    # TODO @jepe: move this to its own module (e.g. as a cellpy-exporters?):
    @staticmethod
    def _convert2fid_table(cell):
        # used when saving cellpy-file
        logging.debug("converting FileID object to fid-table that can be saved")
        fidtable = collections.OrderedDict()
        fidtable["raw_data_name"] = []
        fidtable["raw_data_full_name"] = []
        fidtable["raw_data_size"] = []
        fidtable["raw_data_last_modified"] = []
        fidtable["raw_data_last_accessed"] = []
        fidtable["raw_data_last_info_changed"] = []
        fidtable["raw_data_location"] = []
        # TODO: consider deprecating this as we now have implemented last_data_point:
        fidtable["raw_data_files_length"] = []

        fidtable["last_data_point"] = []
        fids = cell.raw_data_files
        if fids:
            for fid, length in zip(fids, cell.raw_data_files_length):
                try:
                    fidtable["raw_data_name"].append(fid.name)
                    fidtable["raw_data_full_name"].append(fid.full_name)
                    fidtable["raw_data_size"].append(fid.size)
                    fidtable["raw_data_last_modified"].append(fid.last_modified)
                    fidtable["raw_data_last_accessed"].append(fid.last_accessed)
                    fidtable["raw_data_last_info_changed"].append(fid.last_info_changed)
                except AttributeError:  # TODO: this is probably not needed anymore
                    logging.debug("this is probably not from a file")
                    fidtable["raw_data_name"].append("db")
                    fidtable["raw_data_full_name"].append("db")
                    fidtable["raw_data_size"].append(fid.size)
                    fidtable["raw_data_last_modified"].append("db")
                    fidtable["raw_data_last_accessed"].append("db")
                    fidtable["raw_data_last_info_changed"].append("db")

                fidtable["raw_data_location"].append(fid.location)
                fidtable["raw_data_files_length"].append(length)
                fidtable["last_data_point"].append(
                    fid.last_data_point
                )  # will most likely be the same as length
        else:
            warnings.warn("seems you lost info about your raw-data (missing fids)")
        return fidtable

    # TODO @jepe: move this to its own module (e.g. as a cellpy-loader in instruments?):
    @staticmethod
    def _convert2fid_list(tbl):
        # used when reading cellpy-file
        logging.debug("converting loaded fid-table to FileID object")
        fids = []
        lengths = []
        min_amount = 0
        for counter, item in enumerate(tbl["raw_data_name"]):
            fid = FileID()
            try:
                fid.name = OtherPath(item).name
            except NotImplementedError:
                fid.name = item
            fid.full_name = tbl["raw_data_full_name"][counter]
            fid.size = tbl["raw_data_size"][counter]
            fid.last_modified = tbl["raw_data_last_modified"][counter]
            fid.last_accessed = tbl["raw_data_last_accessed"][counter]
            fid.last_info_changed = tbl["raw_data_last_info_changed"][counter]
            fid.location = tbl["raw_data_location"][counter]
            length = tbl["raw_data_files_length"][counter]
            if "last_data_point" in tbl.columns:
                fid.last_data_point = tbl["last_data_point"][counter]
            else:
                fid.last_data_point = 0
            if "is_db" in tbl.columns:
                fid.is_db = tbl["is_db"][counter]
            fids.append(fid)
            lengths.append(length)
            min_amount = 1
        if min_amount < 1:
            logging.debug("info about raw files missing")
        return fids, lengths

    def merge(self, datasets: list, **kwargs):
        """This function merges datasets into one set."""
        logging.info("Merging")
        self.data = datasets.pop(0)
        for data in datasets:
            self.data = self._append(self.data, data, **kwargs)
            for raw_data_file, file_size in zip(
                data.raw_data_files,
                data.raw_data_files_length,
            ):
                self.data.raw_data_files.append(raw_data_file)
                self.data.raw_data_files_length.append(file_size)
        return self

    def _append(self, t1, t2, merge_summary=False, merge_step_table=False, recalc=True):
        logging.debug(
            f"merging two datasets\n(merge summary = {merge_summary})\n"
            f"(merge step table = {merge_step_table})"
        )
        if t1.raw.empty:
            logging.debug("OBS! the first dataset is empty")
            logging.debug(" -> merged contains only second")
            return t2

        if t2.raw.empty:
            logging.debug("OBS! the second dataset was empty")
            logging.debug(" -> merged contains only first")
            return t1

        if not isinstance(t1.loaded_from, (list, tuple)):
            t1.loaded_from = [t1.loaded_from]

        cycle_index_header = self.headers_summary.cycle_index
        data = t1
        if recalc:
            # finding diff of time
            start_time_1 = t1.meta_common.start_datetime
            start_time_2 = t2.meta_common.start_datetime

            if self.tester in ["arbin_res"]:
                diff_time = xldate_as_datetime(start_time_2) - xldate_as_datetime(
                    start_time_1
                )
            else:
                diff_time = start_time_2 - start_time_1
            diff_time = diff_time.total_seconds()

            if diff_time < 0:
                logging.warning("Wow! your new dataset is older than the old!")
            logging.debug(f"diff time: {diff_time}")

            sort_key = self.headers_normal.datetime_txt  # DateTime
            # mod data points for set 2
            data_point_header = self.headers_normal.data_point_txt
            try:
                last_data_point = max(t1.raw[data_point_header])
            except ValueError:
                logging.debug("ValueError when getting last data point for r1")
                last_data_point = 0

            t2.raw[data_point_header] = t2.raw[data_point_header] + last_data_point
            logging.debug("No error getting last data point for r2")
            # mod cycle index for set 2

            try:
                last_cycle = max(t1.raw[cycle_index_header])
            except ValueError:
                logging.debug("ValueError when getting last cycle index for r1")
                last_cycle = 0
            t2.raw[cycle_index_header] = t2.raw[cycle_index_header] + last_cycle
            # mod test time for set 2
            test_time_header = self.headers_normal.test_time_txt
            t2.raw[test_time_header] = t2.raw[test_time_header] + diff_time
        else:
            logging.debug("not doing recalc")
        # merging
        logging.debug("performing concat")
        raw = pd.concat([t1.raw, t2.raw], ignore_index=True)
        data.raw = raw
        data.loaded_from.append(t2.loaded_from)
        step_table_made = False

        if merge_summary:
            # checking if we already have made a summary file of these datasets
            # (to be used if merging summaries (but not properly implemented yet))
            if t1.summary.empty or t2.summary.empty:
                summary_made = False
            else:
                summary_made = True

            try:
                _ = t1.summary[
                    cycle_index_header
                ]  # during loading arbin res files, a stats-frame is loaded into
                _ = t2.summary[
                    cycle_index_header
                ]  # the summary. This prevents merging those.
            except KeyError:
                summary_made = False
                logging.info("The summary is not complete - run make_summary()")

            # checking if we already have made step tables for these datasets
            if t1.has_steps and t2.has_steps:
                step_table_made = True
            else:
                step_table_made = False

            if summary_made:
                # check if (self-made) summary exists.
                logging.debug("merge summaries")
                if recalc:
                    # This part of the code is seldom ran. Careful!
                    # mod cycle index for set 2
                    last_cycle = max(t1.summary[cycle_index_header])
                    t2.summary[cycle_index_header] = (
                        t2.summary[cycle_index_header] + last_cycle
                    )
                    # mod test time for set 2
                    t2.summary[test_time_header] = (
                        t2.summary[test_time_header] + diff_time
                    )
                    # to-do: mod all the cumsum stuff in the summary (best to make
                    # summary after merging) merging

                    t2.summary[data_point_header] = (
                        t2.summary[data_point_header] + last_data_point
                    )

                summary2 = pd.concat([t1.summary, t2.summary], ignore_index=True)

                data.summary = summary2
            else:
                logging.debug(
                    "could not merge summary tables "
                    "(non-existing) -"
                    "create them first!"
                )

        if merge_step_table:
            if step_table_made:
                cycle_index_header = self.headers_normal.cycle_index_txt
                t2.steps[self.headers_step_table.cycle] = (
                    t2.raw[self.headers_step_table.cycle] + last_cycle
                )

                steps2 = pd.concat([t1.steps, t2.steps], ignore_index=True)
                data.steps = steps2
            else:
                logging.debug(
                    "could not merge step tables "
                    "(non-existing) -"
                    "create them first!"
                )

        logging.debug(" -> merged with new dataset")
        # TODO: @jepe -  update merging for more variables
        return data

    # TODO: check if this can be moved to helpers
    def _validate_step_table(self, simple=False):
        step_index_header = self.headers_normal.step_index_txt
        logging.debug("-validating step table")
        d = self.data.raw
        s = self.data.steps

        if not self.data.has_steps:
            return False

        no_cycles_raw = np.amax(d[self.headers_normal.cycle_index_txt])
        headers_step_table = self.headers_step_table
        no_cycles_step_table = np.amax(s[headers_step_table.cycle])

        if simple:
            logging.debug("  (simple)")
            if no_cycles_raw == no_cycles_step_table:
                return True
            else:
                return False

        else:
            validated = True
            if no_cycles_raw != no_cycles_step_table:
                logging.debug("  differ in no. of cycles")
                validated = False
            else:
                for j in range(1, no_cycles_raw + 1):
                    cycle_number = j
                    no_steps_raw = len(
                        np.unique(
                            d.loc[
                                d[self.headers_normal.cycle_index_txt] == cycle_number,
                                self.headers_normal.step_index_txt,
                            ]
                        )
                    )
                    no_steps_step_table = len(
                        s.loc[
                            s[headers_step_table.cycle] == cycle_number,
                            headers_step_table.step,
                        ]
                    )
                    if no_steps_raw != no_steps_step_table:
                        validated = False
            return validated

    def print_steps(self):
        """Print the step table."""
        st = self.data.steps
        print(st)

    def get_step_numbers(
        self,
        steptype="charge",
        allctypes=True,
        pdtype=False,
        cycle_number=None,
        trim_taper_steps=None,
        steps_to_skip=None,
        steptable=None,
    ):
        # TODO: @jepe - include sub_steps here
        # TODO: @jepe - include option for not selecting taper steps here
        # TODO: @jepe - refactor this method!
        """Get the step numbers of selected type.

        Returns the selected step_numbers for the selected type of step(s).

        Args:
            steptype (string): string identifying type of step.
            allctypes (bool): get all types of charge (or discharge).
            pdtype (bool): return results as pandas.DataFrame
            cycle_number (int): selected cycle, selects all if not set.
            trim_taper_steps (integer): number of taper steps to skip (counted
                from the end, i.e. 1 means skip last step in each cycle).
            steps_to_skip (list): step numbers that should not be included.
            steptable (pandas.DataFrame): optional steptable

        Returns:
            A dictionary containing a list of step numbers corresponding
                to the selected steptype for the cycle(s).
            Returns a pandas.DataFrame instead of a dict of lists if pdtype is
                set to True. The frame is a sub-set of the step-table frame
                (i.e. all the same columns, only filtered by rows).

        Example:
            >>> my_charge_steps = CellpyCell.get_step_numbers(
            >>>    "charge",
            >>>    cycle_number = 3
            >>> )
            >>> print my_charge_steps
            {3: [5,8]}

        """
        t0 = time.time()
        # logging.debug("Trying to get step-types")
        if steps_to_skip is None:
            steps_to_skip = []

        if steptable is None:
            if not self.data.has_steps:
                logging.debug("steps is not made")

                if self.force_step_table_creation or self.force_all:
                    logging.debug("creating step_table for")
                    logging.debug(self.data.loaded_from)
                    self.make_step_table()

                else:
                    logging.info("ERROR! Cannot use get_steps: create step_table first")
                    logging.info("You could use find_step_numbers method instead")
                    logging.info("(but I don't recommend it)")
                    return None

        # check if steptype is valid
        steptype = steptype.lower()
        steptypes = []
        helper_step_types = ["ocv", "charge_discharge"]
        valid_step_type = True
        # logging.debug(f"dt 2: {time.time() - t0}")
        if steptype in self.list_of_step_types:
            steptypes.append(steptype)
        else:
            txt = f"{steptype} is not a valid core steptype"
            if steptype in helper_step_types:
                txt = "but a helper steptype"
                if steptype == "ocv":
                    steptypes.append("ocvrlx_up")
                    steptypes.append("ocvrlx_down")
                elif steptype == "charge_discharge":
                    steptypes.append("charge")
                    steptypes.append("discharge")
            else:
                valid_step_type = False
            # logging.debug(txt)
        if not valid_step_type:
            return None

        # in case of selection `allctypes`, then modify charge, discharge
        if allctypes:
            add_these = []
            for st in steptypes:
                if st in ["charge", "discharge"]:
                    st1 = st + "_cv"
                    add_these.append(st1)
                    st1 = "cv_" + st
                    add_these.append(st1)
            for st in add_these:
                steptypes.append(st)

        # logging.debug("Your steptypes:")
        # logging.debug(steptypes)

        if steptable is None:
            st = self.data.steps
        else:
            st = steptable
        shdr = self.headers_step_table

        # retrieving cycle numbers
        # logging.debug(f"dt 3: {time.time() - t0}")
        if cycle_number is None:
            cycle_numbers = self.get_cycle_numbers(steptable=steptable)
        else:
            if isinstance(cycle_number, collections.abc.Iterable):
                cycle_numbers = cycle_number
            else:
                cycle_numbers = [cycle_number]

        if trim_taper_steps is not None:
            trim_taper_steps = -trim_taper_steps
            # logging.debug("taper steps to trim given")

        if pdtype:
            # logging.debug("Return pandas dataframe.")
            if trim_taper_steps:
                logging.info(
                    "Trimming taper steps is currently not"
                    "possible when returning pd.DataFrame. "
                    "Do it manually instead."
                )
            out = st[st[shdr.type].isin(steptypes) & st[shdr.cycle].isin(cycle_numbers)]
            return out

        # if not pdtype, return a dict instead
        # logging.debug("out as dict; out[cycle] = [s1,s2,...]")
        # logging.debug("(same behaviour as find_step_numbers)")
        # logging.debug("return dict of lists")
        # logging.warning(
        #     "returning dict will be deprecated",
        # )
        out = dict()
        # logging.debug(f"return a dict")
        # logging.debug(f"dt 4: {time.time() - t0}")
        for cycle in cycle_numbers:
            steplist = []
            for s in steptypes:
                mask_type_and_cycle = (st[shdr.type] == s) & (st[shdr.cycle] == cycle)
                if not any(mask_type_and_cycle):
                    logging.debug(f"found nothing for cycle {cycle}")
                else:
                    step = st[mask_type_and_cycle][shdr.step].tolist()
                    for newstep in step[:trim_taper_steps]:
                        if newstep in steps_to_skip:
                            logging.debug(f"skipping step {newstep}")
                        else:
                            steplist.append(int(newstep))

            if not steplist:
                steplist = [0]
            out[cycle] = steplist
        # logging.debug(f"dt tot: {time.time() - t0}")
        return out

    def load_step_specifications(self, file_name, short=False):
        """Load a table that contains step-type definitions.

        This function loads a file containing a specification for each step or
        for each (cycle_number, step_number) combinations if short==False. The
        step_cycle specifications that are allowed are stored in the variable
        cellreader.list_of_step_types.
        """

        # if short:
        #     # the table only consists of steps (not cycle,step pairs) assuming
        #     # that the step numbers uniquely defines step type (this is true
        #     # for arbin at least).
        #     raise NotImplementedError

        step_specs = pd.read_csv(file_name, sep=prms.Reader.sep)
        if "step" not in step_specs.columns:
            logging.info("Missing column: step")
            raise IOError

        if "type" not in step_specs.columns:
            logging.info("Missing column: type")
            raise IOError

        if not short and "cycle" not in step_specs.columns:
            logging.info("Missing column: cycle")
            raise IOError

        self.make_step_table(step_specifications=step_specs, short=short)

    def _sort_data(self, dataset):
        # TODO: [# index]
        if self.headers_normal.data_point_txt in dataset.raw.columns:
            dataset.raw = dataset.raw.sort_values(
                self.headers_normal.data_point_txt
            ).reset_index()
            return dataset

        logging.debug("_sort_data: no datapoint header to sort by")

    def _ustep(self, n):
        un = []
        c = 0
        dn = n.diff()
        for i in dn:
            if i != 0:
                c += 1
            un.append(c)
        logging.debug("created u-steps")
        return un

    def make_step_table(
        self,
        step_specifications=None,
        short=False,
        override_step_types=None,
        override_raw_limits=None,
        profiling=False,
        all_steps=False,
        add_c_rate=True,
        skip_steps=None,
        sort_rows=True,
        from_data_point=None,
        nom_cap_specifics=None,
    ):
        """Create a table (v.4) that contains summary information for each step.

        This function creates a table containing information about the
        different steps for each cycle and, based on that, decides what type of
        step it is (e.g. charge) for each cycle.

        The format of the steps is:

            index: cycleno - stepno - sub-step-no - ustep
            Time info: average, stdev, max, min, start, end, delta
            Logging info: average, stdev, max, min, start, end, delta
            Current info: average, stdev, max, min, start, end, delta
            Voltage info: average,  stdev, max, min, start, end, delta
            Type: (from pre-defined list) - SubType
            Info: not used.

        Args:
            step_specifications (pandas.DataFrame): step specifications
            short (bool): step specifications in short format
            override_step_types (dict): override the provided step types, for example set all
                steps with step number 5 to "charge" by providing {5: "charge"}.
            override_raw_limits (dict): override the instrument limits (resolution), for example set
                'current_hard' to 0.1 by providing {'current_hard': 0.1}.
            profiling (bool): turn on profiling
            all_steps (bool): investigate all steps including same steps within
                one cycle (this is useful for e.g. GITT).
            add_c_rate (bool): include a C-rate estimate in the steps
            skip_steps (list of integers): list of step numbers that should not
                be processed (future feature - not used yet).
            sort_rows (bool): sort the rows after processing.
            from_data_point (int): first data point to use.
            nom_cap_specifics (str): "gravimetric", "areal", or "absolute".

        Returns:
            None

        """
        # TODO: @jepe - include option for omitting steps
        # TODO: @jepe  - make it is possible to update only new data

        time_00 = time.time()

        if nom_cap_specifics is None:
            nom_cap_specifics = self.nom_cap_specifics

        if profiling:
            print("PROFILING MAKE_STEP_TABLE".center(80, "="))

        def first(x):
            return x.iloc[0]

        def last(x):
            return x.iloc[-1]

        def delta(x):
            # Remark! this will not work if x is a TimeDelta object
            if x.iloc[0] == 0.0:
                # starts from a zero value
                difference = 100.0 * x.iloc[-1]
            else:
                difference = (x.iloc[-1] - x.iloc[0]) * 100 / abs(x.iloc[0])

            return difference

        nhdr = self.headers_normal
        shdr = self.headers_step_table

        if from_data_point is not None:
            df = self.data.raw.loc[
                self.data.raw[nhdr.data_point_txt] >= from_data_point
            ]
        else:
            df = self.data.raw
        # df[shdr.internal_resistance_change] = \
        #     df[nhdr.internal_resistance_txt].pct_change()

        # selecting only the most important columns from raw:
        keep = [
            nhdr.data_point_txt,
            nhdr.test_time_txt,
            nhdr.step_time_txt,
            nhdr.step_index_txt,
            nhdr.cycle_index_txt,
            nhdr.current_txt,
            nhdr.voltage_txt,
            nhdr.ref_voltage_txt,
            nhdr.charge_capacity_txt,
            nhdr.discharge_capacity_txt,
            nhdr.internal_resistance_txt,
            # "ir_pct_change"
        ]

        # only use col-names that exist:
        keep = [col for col in keep if col in df.columns]
        df = df[keep]
        # preparing for implementation of sub_steps (will come in the future):
        df[nhdr.sub_step_index_txt] = 1

        # using headers as defined in the internal_settings.py file
        rename_dict = {
            nhdr.cycle_index_txt: shdr.cycle,
            nhdr.step_index_txt: shdr.step,
            nhdr.sub_step_index_txt: shdr.sub_step,
            nhdr.data_point_txt: shdr.point,
            nhdr.test_time_txt: shdr.test_time,
            nhdr.step_time_txt: shdr.step_time,
            nhdr.current_txt: shdr.current,
            nhdr.voltage_txt: shdr.voltage,
            nhdr.charge_capacity_txt: shdr.charge,
            nhdr.discharge_capacity_txt: shdr.discharge,
            nhdr.internal_resistance_txt: shdr.internal_resistance,
        }

        df = df.rename(columns=rename_dict)
        by = [shdr.cycle, shdr.step, shdr.sub_step]

        if skip_steps is not None:
            logging.debug(f"omitting steps {skip_steps}")
            df = df.loc[~df[shdr.step].isin(skip_steps)]

        if all_steps:
            by.append(shdr.ustep)
            df[shdr.ustep] = self._ustep(df[shdr.step])

        logging.debug(f"groupby: {by}")

        if profiling:
            time_01 = time.time()

        # TODO: make sure that all columns are numeric

        gf = df.groupby(by=by)

        # TODO: FutureWarning: The provided callable <function mean at 0x000002BD4D332840>
        #  is currently using SeriesGroupBy.mean. In a future version of pandas, the provided
        #  callable will be used directly. To keep current behavior pass the string "mean" instead.
        df_steps = gf.agg(["mean", "std", "min", "max", "first", "last", delta]).rename(
            columns={"amin": "min", "amax": "max", "mean": "avr"}
        )

        df_steps = df_steps.reset_index()

        if profiling:
            print(f"*** groupby-agg: {time.time() - time_01} s")
            time_01 = time.time()

        # column with C-rates:
        if add_c_rate:
            logging.debug("adding c-rates")
            nom_cap = self.data.nom_cap
            if nom_cap_specifics == "gravimetric":
                mass = self.data.mass
                nom_cap = self.nominal_capacity_as_absolute(
                    nom_cap, mass, nom_cap_specifics
                )

            elif nom_cap_specifics == "areal":
                area = self.data.active_electrode_area
                nom_cap = self.nominal_capacity_as_absolute(
                    nom_cap, area, nom_cap_specifics
                )
            df_steps[shdr.rate_avr] = abs(
                round(
                    df_steps.loc[:, (shdr.current, "avr")] / nom_cap,
                    DIGITS_C_RATE,
                )
            )
        df_steps[shdr.type] = ""
        df_steps[shdr.sub_type] = ""
        df_steps[shdr.info] = ""

        if step_specifications is None:
            # TODO: refactor this:
            if override_raw_limits is None:
                override_raw_limits = {}
            current_limit_value_hard = (
                override_raw_limits.get("current_hard", None)
                or self.raw_limits["current_hard"]
            )
            current_limit_value_soft = (
                override_raw_limits.get("current_soft", None)
                or self.raw_limits["current_soft"]
            )
            stable_current_limit_hard = (
                override_raw_limits.get("stable_current_hard", None)
                or self.raw_limits["stable_current_hard"]
            )
            stable_current_limit_soft = (
                override_raw_limits.get("stable_current_soft", None)
                or self.raw_limits["stable_current_soft"]
            )
            stable_voltage_limit_hard = (
                override_raw_limits.get("stable_voltage_hard", None)
                or self.raw_limits["stable_voltage_hard"]
            )
            stable_voltage_limit_soft = (
                override_raw_limits.get("stable_voltage_soft", None)
                or self.raw_limits["stable_voltage_soft"]
            )
            stable_charge_limit_hard = (
                override_raw_limits.get("stable_charge_hard", None)
                or self.raw_limits["stable_charge_hard"]
            )
            stable_charge_limit_soft = (
                override_raw_limits.get("stable_charge_soft", None)
                or self.raw_limits["stable_charge_soft"]
            )
            ir_change_limit = (
                override_raw_limits.get("ir_change", None)
                or self.raw_limits["ir_change"]
            )

            mask_no_current_hard = (
                df_steps.loc[:, (shdr.current, "max")].abs()
                + df_steps.loc[:, (shdr.current, "min")].abs()
            ) < current_limit_value_hard / 2

            mask_voltage_down = (
                df_steps.loc[:, (shdr.voltage, "delta")] < -stable_voltage_limit_hard
            )

            mask_voltage_up = (
                df_steps.loc[:, (shdr.voltage, "delta")] > stable_voltage_limit_hard
            )

            mask_voltage_stable = (
                df_steps.loc[:, (shdr.voltage, "delta")].abs()
                < stable_voltage_limit_hard
            )

            mask_current_down = (
                df_steps.loc[:, (shdr.current, "delta")] < -stable_current_limit_soft
            )

            mask_current_up = (
                df_steps.loc[:, (shdr.current, "delta")] > stable_current_limit_soft
            )

            mask_current_negative = (
                df_steps.loc[:, (shdr.current, "avr")] < -current_limit_value_hard
            )

            mask_current_positive = (
                df_steps.loc[:, (shdr.current, "avr")] > current_limit_value_hard
            )

            mask_galvanostatic = (
                df_steps.loc[:, (shdr.current, "delta")].abs()
                < stable_current_limit_soft
            )

            mask_charge_changed = (
                df_steps.loc[:, (shdr.charge, "delta")].abs() > stable_charge_limit_hard
            )

            mask_discharge_changed = (
                df_steps.loc[:, (shdr.discharge, "delta")].abs()
                > stable_charge_limit_hard
            )

            mask_no_change = (
                (df_steps.loc[:, (shdr.voltage, "delta")] == 0)
                & (df_steps.loc[:, (shdr.current, "delta")] == 0)
                & (df_steps.loc[:, (shdr.charge, "delta")] == 0)
                & (df_steps.loc[:, (shdr.discharge, "delta")] == 0)
            )

            # TODO: make an option for only checking unique steps
            #     e.g.
            #     df_x = df_steps.where.steps.are.unique

            # TODO: FutureWarning: Setting an item of incompatible dtype is deprecated and will raise in a future error
            #  of pandas. Value 'rest' has dtype incompatible with float64, please explicitly cast to a
            #  compatible dtype first.

            df_steps.loc[
                mask_no_current_hard & mask_voltage_stable, (shdr.type, slice(None))
            ] = "rest"

            df_steps.loc[
                mask_no_current_hard & mask_voltage_up, (shdr.type, slice(None))
            ] = "ocvrlx_up"

            df_steps.loc[
                mask_no_current_hard & mask_voltage_down, (shdr.type, slice(None))
            ] = "ocvrlx_down"

            df_steps.loc[
                mask_discharge_changed & mask_current_negative, (shdr.type, slice(None))
            ] = "discharge"

            df_steps.loc[
                mask_charge_changed & mask_current_positive, (shdr.type, slice(None))
            ] = "charge"

            df_steps.loc[
                mask_voltage_stable & mask_current_negative & mask_current_down,
                (shdr.type, slice(None)),
            ] = "cv_discharge"

            df_steps.loc[
                mask_voltage_stable & mask_current_positive & mask_current_down,
                (shdr.type, slice(None)),
            ] = "cv_charge"

            # --- internal resistance ----
            df_steps.loc[mask_no_change, (shdr.type, slice(None))] = "ir"
            # assumes that IR is stored in just one row

            # --- sub-step-txt -----------
            df_steps[shdr.sub_type] = None

            # --- CV steps ----

            # "voltametry_charge"
            # mask_charge_changed
            # mask_voltage_up
            # (could also include abs-delta-cumsum current)

            # "voltametry_discharge"
            # mask_discharge_changed
            # mask_voltage_down

            if override_step_types is not None:
                for step, step_type in override_step_types.items():
                    df_steps.loc[
                        df_steps[shdr.step] == step, (shdr.type, slice(None))
                    ] = step_type

            if profiling:
                print(f"*** masking: {time.time() - time_01} s")
                time_01 = time.time()

        else:
            logging.debug("parsing custom step definition")
            if not short:
                logging.debug("using long format (cycle,step)")
                for row in step_specifications.itertuples():
                    df_steps.loc[
                        (df_steps[shdr.step] == row.step)
                        & (df_steps[shdr.cycle] == row.cycle),
                        (shdr.type, slice(None)),
                    ] = row.type
                    df_steps.loc[
                        (df_steps[shdr.step] == row.step)
                        & (df_steps[shdr.cycle] == row.cycle),
                        (shdr.info, slice(None)),
                    ] = row.info
            else:
                logging.debug("using short format (step)")
                for row in step_specifications.itertuples():
                    df_steps.loc[
                        df_steps[shdr.step] == row.step, (shdr.type, slice(None))
                    ] = row.type
                    df_steps.loc[
                        df_steps[shdr.step] == row.step, (shdr.info, slice(None))
                    ] = row.info

        if profiling:
            print(f"*** introspect: {time.time() - time_01} s")

        # check if all the steps got categorizes
        logging.debug("looking for un-categorized steps")
        empty_rows = df_steps.loc[df_steps[shdr.type].isnull()]
        if not empty_rows.empty:
            logging.warning(
                f"found {len(empty_rows)}"
                f":{len(df_steps)} non-categorized steps "
                f"(please, check your raw-limits)"
            )
            # logging.debug(empty_rows)

        # flatten (possible remove in the future),

        logging.debug(f"flatten columns")
        if profiling:
            time_01 = time.time()
        flat_cols = []
        for col in df_steps.columns:
            if isinstance(col, tuple):
                if col[-1]:
                    col = "_".join(col)
                else:
                    col = col[0]
            flat_cols.append(col)

        df_steps.columns = flat_cols
        if sort_rows:
            logging.debug("sorting the step rows")
            # TODO: [#index]
            # if this throws a KeyError: 'test_time_first' it probably
            # means that the df contains a non-nummeric 'test_time' column.
            df_steps = df_steps.sort_values(by=shdr.test_time + "_first").reset_index()

        if profiling:
            print(f"*** flattening: {time.time() - time_01} s")

        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")

        if from_data_point is not None:
            return df_steps
        else:
            self.data.steps = df_steps
            return self

    def select_steps(self, step_dict, append_df=False):
        """Select steps (not documented yet)."""
        raise DeprecatedFeature

    def _select_step(self, cycle, step):
        # TODO: @jepe - insert sub_step here
        test = self.data

        # check if columns exist
        c_txt = self.headers_normal.cycle_index_txt
        s_txt = self.headers_normal.step_index_txt
        y_txt = self.headers_normal.voltage_txt
        x_txt = self.headers_normal.discharge_capacity_txt  # jepe fix

        # no_cycles=np.amax(test.raw[c_txt])
        # print d.columns

        if not any(test.raw.columns == c_txt):
            logging.info("ERROR - cannot find %s" % c_txt)
            sys.exit(-1)
        if not any(test.raw.columns == s_txt):
            logging.info("ERROR - cannot find %s" % s_txt)
            sys.exit(-1)

        # logging.debug(f"selecting cycle {cycle} step {step}")
        v = test.raw[(test.raw[c_txt] == cycle) & (test.raw[s_txt] == step)]

        if self._is_empty_array(v):
            logging.debug("empty dataframe")
            return None
        else:
            return v

    def populate_step_dict(self, step):
        """Returns a dict with cycle numbers as keys
        and corresponding steps (list) as values."""
        raise DeprecatedFeature

    def _export_cycles(
        self,
        setname=None,
        sep=None,
        outname=None,
        shifted=False,
        method=None,
        shift=0.0,
        last_cycle=None,
    ):
        # export voltage - capacity curves to .csv file

        logging.debug("START exporing cycles")
        time_00 = time.time()
        lastname = "_cycles.csv"
        if sep is None:
            sep = self.sep
        if outname is None:
            outname = setname + lastname

        logging.debug(f"outname: {outname}")

        list_of_cycles = self.get_cycle_numbers()
        if last_cycle is not None:
            list_of_cycles = [c for c in list_of_cycles if c <= int(last_cycle)]
            logging.debug(f"only processing up to cycle {last_cycle}")
            logging.debug(f"you have {len(list_of_cycles)}" f"cycles to process")
        out_data = []
        c = None
        if not method:
            method = "back-and-forth"
        if shifted:
            method = "back-and-forth"
            shift = 0.0
            _last = 0.0
        logging.debug(f"number of cycles: {len(list_of_cycles)}")
        for cycle in list_of_cycles:
            try:
                if shifted and c is not None:
                    shift = _last
                    # print(f"shifted = {shift}, first={_first}")
                df = self.get_cap(cycle, method=method, shift=shift)
                if df.empty:
                    logging.debug("NoneType from get_cap")
                else:
                    c = df["capacity"]
                    v = df["voltage"]

                    _last = c.iat[-1]
                    _first = c.iat[0]

                    c = c.tolist()
                    v = v.tolist()
                    header_x = "cap cycle_no %i" % cycle
                    header_y = "voltage cycle_no %i" % cycle
                    c.insert(0, header_x)
                    v.insert(0, header_y)
                    out_data.append(c)
                    out_data.append(v)
                    # txt = "extracted cycle %i" % cycle
                    # logging.debug(txt)
            except IndexError as e:
                txt = "Could not extract cycle %i" % cycle
                logging.info(txt)
                logging.debug(e)

        # Saving cycles in one .csv file (x,y,x,y,x,y...)
        # print "saving the file with delimiter '%s' " % (sep)
        logging.debug("writing cycles to file")
        with open(outname, "w", newline="") as f:
            writer = csv.writer(f, delimiter=sep)
            writer.writerows(itertools.zip_longest(*out_data))
            # star (or asterix) means transpose (writing cols instead of rows)

        logging.info(f"The file {outname} was created")
        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")
        logging.debug("END exporting cycles")

    # TODO: remove this
    def _export_cycles_old(
        self,
        setname=None,
        sep=None,
        outname=None,
        shifted=False,
        method=None,
        shift=0.0,
        last_cycle=None,
    ):
        # export voltage - capacity curves to .csv file

        logging.debug("*** OLD EXPORT-CYCLES METHOD***")
        lastname = "_cycles.csv"
        if sep is None:
            sep = self.sep
        if outname is None:
            outname = setname + lastname

        list_of_cycles = self.get_cycle_numbers()
        logging.debug(f"you have {len(list_of_cycles)} cycles")
        if last_cycle is not None:
            list_of_cycles = [c for c in list_of_cycles if c <= int(last_cycle)]
            logging.debug(f"only processing up to cycle {last_cycle}")
            logging.debug(f"you have {len(list_of_cycles)}" f"cycles to process")
        out_data = []
        c = None
        if not method:
            method = "back-and-forth"
        if shifted:
            method = "back-and-forth"
            shift = 0.0
            _last = 0.0

        for cycle in list_of_cycles:
            try:
                if shifted and c is not None:
                    shift = _last
                    # print(f"shifted = {shift}, first={_first}")
                c, v = self.get_cap(cycle, method=method, shift=shift)
                if c is None:
                    logging.debug("NoneType from get_cap")
                else:
                    _last = c.iat[-1]
                    _first = c.iat[0]

                    c = c.tolist()
                    v = v.tolist()
                    header_x = "cap cycle_no %i" % cycle
                    header_y = "voltage cycle_no %i" % cycle
                    c.insert(0, header_x)
                    v.insert(0, header_y)
                    out_data.append(c)
                    out_data.append(v)
                    # txt = "extracted cycle %i" % cycle
                    # logging.debug(txt)
            except IndexError as e:
                txt = "Could not extract cycle %i" % cycle
                logging.info(txt)
                logging.debug(e)

        # Saving cycles in one .csv file (x,y,x,y,x,y...)
        # print "saving the file with delimiter '%s' " % (sep)
        logging.debug("writing cycles to file")
        with open(outname, "w", newline="") as f:
            writer = csv.writer(f, delimiter=sep)
            writer.writerows(itertools.zip_longest(*out_data))
            # star (or asterix) means transpose (writing cols instead of rows)
        logging.info(f"The file {outname} was created")

    def _export_normal(self, data, setname=None, sep=None, outname=None):
        time_00 = time.time()
        lastname = "_normal.csv"
        if sep is None:
            sep = self.sep
        if outname is None:
            outname = setname + lastname
        txt = outname
        try:
            data.raw.to_csv(outname, sep=sep)
            txt += " OK"
        except Exception as e:
            txt += " Could not save it!"
            logging.debug(e)
            warnings.warn(f"Unhandled exception raised: {e}")
        logging.info(txt)
        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")

    def _export_stats(self, data, setname=None, sep=None, outname=None):
        time_00 = time.time()
        lastname = "_stats.csv"
        if sep is None:
            sep = self.sep
        if outname is None:
            outname = setname + lastname
        txt = outname
        try:
            data.summary.to_csv(outname, sep=sep)
            txt += " OK"
        except Exception as e:
            txt += " Could not save it!"
            logging.debug(e)
            warnings.warn(f"Unhandled exception raised: {e}")
        logging.info(txt)
        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")

    def _export_steptable(self, data, setname=None, sep=None, outname=None):
        # TODO 259: rename to _export_steps_csv
        time_00 = time.time()
        lastname = "_steps.csv"
        if sep is None:
            sep = self.sep
        if outname is None:
            outname = setname + lastname
        txt = outname
        try:
            data.steps.to_csv(outname, sep=sep)
            txt += " OK"
        except Exception as e:
            txt += " Could not save it!"
            logging.debug(e)
            warnings.warn(f"Unhandled exception raised: {e}")
        logging.info(txt)
        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")

    def to_excel(
        self,
        filename=None,
        cycles=None,
        raw=False,
        steps=True,
        nice=True,
        get_cap_kwargs=None,
        to_excel_kwargs=None,
    ):
        """Saves the data as .xlsx file(s).

        Args:
            filename: name of the Excel file.
            cycles: (None, bool, or list of ints) export voltage-capacity curves if given.
            raw: (bool) export raw-data if True.
            steps: (bool) export steps if True.
            nice: (bool) use nice formatting if True.
            get_cap_kwargs: (dict) kwargs for CellpyCell.get_cap method.
            to_excel_kwargs: (dict) kwargs for pandas.DataFrame.to_excel method.
        """
        to_excel_method_kwargs = {"index": True, "header": True}
        get_cap_method_kwargs = {
            "method": "forth-and-forth",
            "label_cycle_number": True,
            "categorical_column": True,
            "interpolated": True,
            "number_of_points": 1000,
            "capacity_then_voltage": True,
        }
        if to_excel_kwargs is not None:
            to_excel_method_kwargs.update(to_excel_kwargs)
        if get_cap_kwargs is not None:
            get_cap_method_kwargs.update(get_cap_kwargs)

        border = openpyxl.styles.Border()
        face_color = "00EEEEEE"
        meta_alignment_left = openpyxl.styles.Alignment(
            horizontal="left", vertical="bottom"
        )
        meta_width = 34
        meta_alignment_right = openpyxl.styles.Alignment(
            horizontal="right", vertical="bottom"
        )
        fill = openpyxl.styles.PatternFill(
            start_color=face_color, end_color=face_color, fill_type="solid"
        )

        if filename is None:
            pre = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{pre}_cellpy.xlsx"
            filename = Path(filename).resolve()
            logging.critical(f"generating filename: {filename}")

        summary_frame = self.data.summary
        meta_common_frame = self.data.meta_common.to_frame()
        meta_test_dependent_frame = self.data.meta_test_dependent.to_frame()
        cellpy_units = self.cellpy_units.to_frame()
        cellpy_units.index = "cellpy_units_" + cellpy_units.index
        raw_units = self.raw_units.to_frame()
        raw_units.index = "raw_units_" + raw_units.index

        meta_common_frame = pd.concat([meta_common_frame, cellpy_units, raw_units])

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            meta_common_frame.to_excel(
                writer, sheet_name="meta_common", **to_excel_method_kwargs
            )
            meta_test_dependent_frame.to_excel(
                writer, sheet_name="meta_test_dependent", **to_excel_method_kwargs
            )
            summary_frame.to_excel(
                writer, sheet_name="summary", **to_excel_method_kwargs
            )

            if raw:
                # TODO: raw-table has two columns called "data_point" at the moment,
                #  so this should be fixed (probably the .set_index("data_point") should be checked)
                logging.debug("exporting raw data")
                raw = self.data.raw
                max_len = 1_048_576
                if len(raw) < max_len:
                    raw.to_excel(writer, sheet_name="raw", **to_excel_method_kwargs)
                else:
                    logging.warning(
                        "Raw data is too large to fit in one sheet. "
                        "Splitting raw data into chunks. This is not tested yet"
                    )
                    n_chunks = len(raw) // max_len + 1
                    for i in range(n_chunks):
                        raw.iloc[i * max_len : (i + 1) * max_len].to_excel(
                            writer, sheet_name=f"raw_{i:02}", **to_excel_method_kwargs
                        )

            if steps:
                logging.debug("exporting steps")
                # TODO: step-table has a columns called "index" at the moment,
                #  so setting index=False for dataframe.to_excel
                #  Maybe best to make sure that step table does not have a column called "index" in the future?
                self.data.steps.to_excel(
                    writer, sheet_name="steps", index=False, header=True
                )
            if cycles:
                logging.debug("exporting cycles")
                if cycles is True:
                    cycles = self.get_cycle_numbers()
                for cycle in cycles:
                    _curves = self.get_cap(cycle=cycle, **get_cap_method_kwargs)
                    _curves.to_excel(
                        writer,
                        sheet_name=f"cycle_{cycle:03}",
                        index=False,
                        header=True,
                    )
            if nice:
                for sheet in writer.sheets.values():
                    if sheet.title.startswith("meta"):
                        sheet.column_dimensions["A"].width = meta_width
                        for xl_cell in sheet["A"]:
                            xl_cell.alignment = meta_alignment_left
                            xl_cell.border = border
                        for xl_cell in sheet["B"]:
                            xl_cell.alignment = meta_alignment_right
                            xl_cell.border = border
                    else:
                        for xl_cell in sheet["A"]:
                            xl_cell.border = border

                    for xl_cell in sheet["1"]:
                        xl_cell.border = border
                        xl_cell.fill = fill

    def to_csv(
        self,
        datadir=None,
        sep=None,
        cycles=False,
        raw=True,
        summary=True,
        shifted=False,
        method=None,
        shift=0.0,
        last_cycle=None,
    ):
        """Saves the data as .csv file(s).

        Args:
            datadir: folder where to save the data (uses current folder if not
                given).
            sep: the separator to use in the csv file
                (defaults to CellpyCell.sep).
            cycles: (bool) export voltage-capacity curves if True.
            raw: (bool) export raw-data if True.
            summary: (bool) export summary if True.
            shifted (bool): export with cumulated shift.
            method (string): how the curves are given::

                "back-and-forth" - standard back and forth; discharge (or charge)
                    reversed from where charge (or discharge) ends.

                "forth" - discharge (or charge) continues along x-axis.

                "forth-and-forth" - discharge (or charge) also starts at 0
                    (or shift if not shift=0.0)

            shift: start-value for charge (or discharge)
            last_cycle: process only up to this cycle (if not None).

        Returns:
            Nothing

        """

        if sep is None:
            sep = self.sep

        logging.debug("saving to csv")

        try:
            data = self.data
        except NoDataFound:
            logging.info("to_csv -")
            logging.info("NoDataFound: not saved!")
            return

        if isinstance(data.loaded_from, (list, tuple)):
            txt = "merged file"
            txt += "using first file as basename"
            logging.debug(txt)
            no_merged_sets = len(data.loaded_from)
            no_merged_sets = "_merged_" + str(no_merged_sets).zfill(3)
            filename = data.loaded_from[0]
        else:
            filename = data.loaded_from
            no_merged_sets = ""

        firstname, extension = os.path.splitext(filename)
        firstname += no_merged_sets
        if datadir:
            firstname = os.path.join(datadir, os.path.basename(firstname))

        if raw:
            outname_normal = firstname + "_normal.csv"
            self._export_normal(data, outname=outname_normal, sep=sep)
            if data.has_steps is True:
                outname_steps = firstname + "_steps.csv"
                self._export_steptable(data, outname=outname_steps, sep=sep)
            else:
                logging.debug("steps_made is not True")

        if summary:
            outname_stats = firstname + "_stats.csv"
            self._export_stats(data, outname=outname_stats, sep=sep)

        if cycles:
            outname_cycles = firstname + "_cycles.csv"
            self._export_cycles(
                outname=outname_cycles,
                sep=sep,
                shifted=shifted,
                method=method,
                shift=shift,
                last_cycle=last_cycle,
            )

    def save(
        self,
        filename,
        force=False,
        overwrite=None,
        extension="h5",
        ensure_step_table=None,
        ensure_summary_table=None,
    ):
        """Save the data structure to cellpy-format.

        Args:
            filename: (str or pathlib.Path) the name you want to give the file
            force: (bool) save a file even if the summary is not made yet
                (not recommended)
            overwrite: (bool) save the new version of the file even if old one
                exists.
            extension: (str) filename extension.
            ensure_step_table: (bool) make step-table if missing.
            ensure_summary_table: (bool) make summary-table if missing.

        Returns: Nothing at all.
        """
        logging.debug(f"Trying to save cellpy-file to {filename}")
        logging.info(f" -> {filename}")

        cellpy_file_format = "hdf5"

        # some checks to find out what you want
        if overwrite is None:
            overwrite = self.overwrite_able

        if ensure_step_table is None:
            ensure_step_table = self.ensure_step_table

        if ensure_summary_table is None:
            ensure_summary_table = self.ensure_summary_table

        my_data = self.data
        summary_made = my_data.has_summary
        if not summary_made and not force and not ensure_summary_table:
            logging.info("File not saved!")
            logging.info("You should not save datasets without making a summary first!")
            logging.info("If you really want to do it, use save with force=True")
            return

        step_table_made = my_data.has_steps
        if not step_table_made and not force and not ensure_step_table:
            logging.info(
                "File not saved!"
                "You should not save datasets without making a step-table first!"
            )
            logging.info("If you really want to do it, use save with force=True")
            return

        outfile_all = OtherPath(filename)
        if not outfile_all.suffix:
            logging.debug("No suffix given - adding one")
            outfile_all = outfile_all.with_suffix(f".{extension}")

        if outfile_all.is_file():
            logging.debug("Outfile exists")
            if overwrite:
                logging.debug("overwrite = True")
                try:
                    os.remove(outfile_all)
                except PermissionError as e:
                    logging.critical("Could not over write old file")
                    logging.info(e)
                    return
            else:
                logging.critical("File exists - did not save")
                logging.info(outfile_all)
                return

        if ensure_step_table:
            logging.debug("ensure_step_table is on")
            if not my_data.has_steps:
                logging.debug("save: creating step table")
                self.make_step_table()

        if ensure_summary_table:
            logging.debug("ensure_summary_table is on")
            if not my_data.has_summary:
                logging.debug("save: creating summary table")
                self.make_summary_table()

        logging.debug("trying to make infotable")
        (
            common_meta_table,
            test_dependent_meta_table,
            fid_table,
        ) = self._create_infotable()

        logging.debug(f"trying to save to file: {outfile_all}")
        if cellpy_file_format == "hdf5":
            # --- saving to hdf5 -----------------------------------
            root = prms._cellpyfile_root  # noqa
            raw_dir = prms._cellpyfile_raw  # noqa
            step_dir = prms._cellpyfile_step  # noqa
            summary_dir = prms._cellpyfile_summary  # noqa
            common_meta_dir = prms._cellpyfile_common_meta  # noqa
            fid_dir = prms._cellpyfile_fid  # noqa
            test_dependent_meta_dir = prms._cellpyfile_test_dependent_meta  # noqa
            warnings.simplefilter("ignore", PerformanceWarning)
            try:
                with pickle_protocol(PICKLE_PROTOCOL):
                    store = self._save_to_hdf5(
                        fid_dir,
                        fid_table,
                        common_meta_table,
                        common_meta_dir,
                        test_dependent_meta_table,
                        test_dependent_meta_dir,
                        my_data,
                        outfile_all,
                        raw_dir,
                        root,
                        step_dir,
                        summary_dir,
                    )
            finally:
                store.close()
            logging.debug(" all -> hdf5 OK")
            warnings.simplefilter("default", PerformanceWarning)
            # del store
        # --- finished saving to hdf5 -------------------------------

    def _save_to_hdf5(
        self,
        fid_dir,
        fid_table,
        infotbl,
        meta_dir,
        test_dependent_meta_table,
        test_dependent_meta_dir,
        my_data,
        outfile_all,
        raw_dir,
        root,
        step_dir,
        summary_dir,
    ):
        store = pd.HDFStore(
            outfile_all,
            complib=prms._cellpyfile_complib,
            complevel=prms._cellpyfile_complevel,
        )
        logging.debug("trying to put raw data")
        logging.debug(" - lets set Data_Point as index")
        hdr_data_point = self.headers_normal.data_point_txt
        if my_data.raw.index.name != hdr_data_point:
            my_data.raw = my_data.raw.set_index(hdr_data_point, drop=False)
        store.put(root + raw_dir, my_data.raw, format=prms._cellpyfile_raw_format)
        logging.debug(" raw -> hdf5 OK")
        logging.debug("trying to put summary")
        store.put(
            root + summary_dir,
            my_data.summary,
            format=prms._cellpyfile_summary_format,
        )
        logging.debug(" summary -> hdf5 OK")

        logging.debug("trying to put meta data")
        store.put(root + meta_dir, infotbl, format=prms._cellpyfile_infotable_format)
        logging.debug(" common meta -> hdf5 OK")
        store.put(
            root + test_dependent_meta_dir,
            test_dependent_meta_table,
            format=prms._cellpyfile_infotable_format,
        )
        logging.debug(" test dependent meta -> hdf5 OK")

        logging.debug("trying to put fidtable")
        store.put(root + fid_dir, fid_table, format=prms._cellpyfile_fidtable_format)
        logging.debug(" fid -> hdf5 OK")
        logging.debug("trying to put step")
        try:
            store.put(
                root + step_dir,
                my_data.steps,
                format=prms._cellpyfile_stepdata_format,
            )
            logging.debug(" step -> hdf5 OK")
        except TypeError:
            my_data = self._fix_dtype_step_table(my_data)
            store.put(
                root + step_dir,
                my_data.steps,
                format=prms._cellpyfile_stepdata_format,
            )
            logging.debug(" fixed step -> hdf5 OK")
        # creating indexes
        # hdr_data_point = self.headers_normal.data_point_txt
        # hdr_cycle_steptable = self.headers_step_table.cycle
        # hdr_cycle_normal = self.headers_normal.cycle_index_txt
        # store.create_table_index(root + "/raw", columns=[hdr_data_point],
        #                          optlevel=9, kind='full')
        return store

    # --------------helper-functions--------------------------------------------
    def _fix_dtype_step_table(self, dataset):
        # used when saving to cellpy format
        hst = get_headers_step_table()
        try:
            cols = dataset.steps.columns
        except AttributeError:
            logging.info("Could not extract columns from steps")
            return
        for col in cols:
            if col not in [hst.cycle, hst.sub_step, hst.info]:
                dataset.steps[col] = dataset.steps[col].apply(pd.to_numeric)
            else:
                dataset.steps[col] = dataset.steps[col].astype("str")
        return dataset

    # TODO: check if this is useful and if it is rename, if not delete
    def _cap_mod_summary(self, summary, capacity_modifier="reset"):
        # Why did I make this method?
        # OBS! modifies the summary table
        time_00 = time.time()
        discharge_title = self.headers_normal.discharge_capacity_txt
        charge_title = self.headers_normal.charge_capacity_txt
        chargecap = 0.0
        dischargecap = 0.0

        # TODO: @jepe - use pd.loc[row,column]

        if capacity_modifier == "reset":
            for index, row in summary.iterrows():
                dischargecap_2 = row[discharge_title]
                summary.loc[index, discharge_title] = dischargecap_2 - dischargecap
                dischargecap = dischargecap_2
                chargecap_2 = row[charge_title]
                summary.loc[index, charge_title] = chargecap_2 - chargecap
                chargecap = chargecap_2
        else:
            raise NotImplementedError

        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")
        return summary

    # TODO: check if this is useful and if it is rename, if not delete
    def _cap_mod_normal(self, capacity_modifier="reset", allctypes=True):
        # Why did I make this method?
        # OBS! modifies the normal table
        time_00 = time.time()
        logging.debug("Not properly checked yet! Use with caution!")

        cycle_index_header = self.headers_normal.cycle_index_txt
        step_index_header = self.headers_normal.step_index_txt
        discharge_index_header = self.headers_normal.discharge_capacity_txt
        discharge_energy_index_header = self.headers_normal.discharge_energy_txt
        charge_index_header = self.headers_normal.charge_capacity_txt
        charge_energy_index_header = self.headers_normal.charge_energy_txt

        raw = self.data.raw

        chargecap = 0.0
        dischargecap = 0.0

        if capacity_modifier == "reset":
            # discharge cycles
            no_cycles = np.amax(raw[cycle_index_header])
            for j in range(1, no_cycles + 1):
                cap_type = "discharge"
                e_header = discharge_energy_index_header
                cap_header = discharge_index_header
                discharge_cycles = self.get_step_numbers(
                    steptype=cap_type, allctypes=allctypes, cycle_number=j
                )

                steps = discharge_cycles[j]
                txt = "Cycle  %i (discharge):  " % j
                logging.debug(txt)
                # TODO: @jepe - use pd.loc[row,column] e.g. pd.loc[:,"charge_cap"]
                # for col or pd.loc[(pd.["step"]==1),"x"]
                selection = (raw[cycle_index_header] == j) & (
                    raw[step_index_header].isin(steps)
                )
                c0 = raw[selection].iloc[0][cap_header]
                e0 = raw[selection].iloc[0][e_header]
                raw.loc[selection, cap_header] = raw.loc[selection, cap_header] - c0
                raw.loc[selection, e_header] = raw.loc[selection, e_header] - e0

                cap_type = "charge"
                e_header = charge_energy_index_header
                cap_header = charge_index_header
                charge_cycles = self.get_step_numbers(
                    steptype=cap_type, allctypes=allctypes, cycle_number=j
                )
                steps = charge_cycles[j]
                txt = "Cycle  %i (charge):  " % j
                logging.debug(txt)

                selection = (raw[cycle_index_header] == j) & (
                    raw[step_index_header].isin(steps)
                )

                if any(selection):
                    c0 = raw[selection].iloc[0][cap_header]
                    e0 = raw[selection].iloc[0][e_header]
                    raw.loc[selection, cap_header] = raw.loc[selection, cap_header] - c0
                    raw.loc[selection, e_header] = raw.loc[selection, e_header] - e0
        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")

    def get_mass(self):
        return self.data.meta_common.mass

    def sget_voltage(self, cycle, step):
        """Returns voltage for cycle, step.

        Convenience function; same as issuing
           raw[(raw[cycle_index_header] == cycle) &
                 (raw[step_index_header] == step)][voltage_header]

        Args:
            cycle: cycle number
            step: step number

        Returns:
            pandas.Series or None if empty
        """
        header = self.headers_normal.voltage_txt
        return self._sget(cycle, step, header, usteps=False)

    def sget_current(self, cycle, step):
        """Returns current for cycle, step.

        Convenience function; same as issuing
           raw[(raw[cycle_index_header] == cycle) &
                 (raw[step_index_header] == step)][current_header]

        Args:
            cycle: cycle number
            step: step number

        Returns:
            pandas.Series or None if empty
        """
        header = self.headers_normal.current_txt
        return self._sget(cycle, step, header, usteps=False)

    def get_raw(
        self,
        header,
        cycle: Optional[Union[Iterable, int]] = None,
        with_index: bool = True,
        with_step: bool = False,
        with_time: bool = False,
        additional_headers: Optional[list] = None,
        as_frame: bool = True,
        scaler: Optional[float] = None,
    ) -> Union[pd.DataFrame, List[np.array]]:
        """Returns the values for column with given header (in raw units).

        Args:
            header: header name.
            cycle: cycle number (all cycles if None).
            with_index: if True, includes the cycle index as a column in the returned pandas.DataFrame.
            with_step: if True, includes the step index as a column in the returned pandas.DataFrame.
            with_time: if True, includes the time as a column in the returned pandas.DataFrame.
            additional_headers (list): additional headers to include in the returned pandas.DataFrame.
            as_frame: if not True, returns a list of current values as numpy arrays (one for each cycle).
                Remark that with_time and with_index will be False if as_frame is set to False.
            scaler: if not None, the returned values are scaled by this value.

        Returns:
            pandas.DataFrame (or list of numpy arrays if as_frame=False)
        """
        y_header = header  # Consider including some lookup handling here
        cycle_index_header = self.headers_normal.cycle_index_txt
        time_header = self.headers_normal.test_time_txt
        step_index_header = self.headers_normal.step_index_txt

        if not as_frame:
            with_time = False
            with_index = True
            with_step = False
            additional_headers = None

        y_headers = [y_header]
        if with_time:
            y_headers.append(time_header)
        if with_step:
            y_headers.append(step_index_header)
        if with_index:
            y_headers.append(cycle_index_header)

        y_headers = reversed(y_headers)
        if additional_headers is not None:
            y_headers.extend(additional_headers)

        data = self.data.raw

        if cycle is None:
            cycle = self.get_cycle_numbers()
        else:
            if not isinstance(cycle, collections.abc.Iterable):
                cycle = [cycle]

        logging.debug(f"getting current for cycles {cycle}")
        c = data.loc[(data[cycle_index_header].isin(cycle)), y_headers]

        if scaler is not None:
            c[y_header] = c[y_header] * scaler

        if not as_frame:
            gb = c.groupby(cycle_index_header)
            c = [gb.get_group(x) for x in gb.groups]
            c = [x[y_header].values for x in c]
        return c

    def get_voltage(self, cycle=None, with_index=True, with_time=False, as_frame=True):
        """Returns voltage (in raw units).

        Args:
            cycle: cycle number (all cycles if None).
            with_index: if True, includes the cycle index as a column in the returned pandas.DataFrame.
            with_time: if True, includes the time as a column in the returned pandas.DataFrame.
            as_frame: if not True, returns a list of current values as numpy arrays (one for each cycle).
                Remark that with_time and with_index will be False if as_frame is set to False.

        Returns:
            pandas.DataFrame (or list of pandas.Series if cycle=None and as_frame=False)
        """

        y_header = self.headers_normal.voltage_txt
        return self.get_raw(
            y_header,
            cycle=cycle,
            with_index=with_index,
            with_time=with_time,
            as_frame=as_frame,
            with_step=False,
            additional_headers=None,
            scaler=None,
        )

    def get_current(self, cycle=None, with_index=True, with_time=False, as_frame=True):
        """Returns current (in raw units).

        Args:
            cycle: cycle number (all cycles if None).
            with_index: if True, includes the cycle index as a column in the returned pandas.DataFrame.
            with_time: if True, includes the time as a column in the returned pandas.DataFrame.
            as_frame: if not True, returns a list of current values as numpy arrays (one for each cycle).
                Remark that with_time and with_index will be False if as_frame is set to False.

        Returns:
            pandas.DataFrame (or list of pandas.Series if cycle=None and as_frame=False)
        """

        y_header = self.headers_normal.current_txt
        return self.get_raw(
            y_header,
            cycle=cycle,
            with_index=with_index,
            with_time=with_time,
            as_frame=as_frame,
            with_step=False,
            additional_headers=None,
            scaler=None,
        )

    def get_datetime(self, cycle=None, with_index=True, with_time=False, as_frame=True):
        """Returns datetime (in raw units).

        Args:
            cycle: cycle number (all cycles if None).
            with_index: if True, includes the cycle index as a column in the returned pandas.DataFrame.
            with_time: if True, includes the time as a column in the returned pandas.DataFrame.
            as_frame: if not True, returns a list of current values as numpy arrays (one for each cycle).
                Remark that with_time and with_index will be False if as_frame is set to False.

        Returns:
            pandas.DataFrame (or list of pandas.Series if cycle=None and as_frame=False)
        """

        y_header = self.headers_normal.datetime_txt
        return self.get_raw(
            y_header,
            cycle=cycle,
            with_index=with_index,
            with_time=with_time,
            as_frame=as_frame,
            with_step=False,
            additional_headers=None,
            scaler=None,
        )

    def get_timestamp(
        self, cycle=None, with_index=True, as_frame=True, in_minutes=False, units="raw"
    ):
        """Returns timestamp.

        Args:
            cycle: cycle number (all cycles if None).
            with_index: if True, includes the cycle index as a column in the returned pandas.DataFrame.
            as_frame: if not True, returns a list of current values as numpy arrays (one for each cycle).
                Remark that with_time and with_index will be False if as_frame is set to False.
            in_minutes: (deprecated, use units="minutes" instead) return values in minutes
                instead of seconds if True.
            units: return values in given time unit ("raw", "seconds", "minutes", "hours").

        Returns:
            pandas.DataFrame (or list of pandas.Series if cycle=None and as_frame=False)
        """

        y_header = self.headers_normal.test_time_txt

        if in_minutes:
            units = "minutes"

        if units == "raw":
            scaler = None
        else:
            scaler = self.unit_scaler_from_raw(units, "time")

        return self.get_raw(
            y_header,
            cycle=cycle,
            with_index=with_index,
            with_time=False,
            as_frame=as_frame,
            with_step=False,
            additional_headers=None,
            scaler=scaler,
        )

    def sget_steptime(self, cycle, step):
        """Returns step time for cycle, step.

        Convenience function; same as issuing
           raw[(raw[cycle_index_header] == cycle) &
                 (raw[step_index_header] == step)][step_time_header]

        Args:
            cycle: cycle number
            step: step number

        Returns:
            pandas.Series or None if empty
        """

        header = self.headers_normal.step_time_txt
        return self._sget(cycle, step, header, usteps=False)

    def _sget(self, cycle, step, header, usteps=False):
        logging.debug(f"searching for {header}")

        cycle_index_header = self.headers_normal.cycle_index_txt
        step_index_header = self.headers_normal.step_index_txt

        if usteps:
            print("Using sget for usteps is not supported yet.")
            print("I encourage you to work with the DataFrames directly instead.")
            print(" - look up the 'ustep' in the steps DataFrame")
            print(" - get the start and end 'data_point'")
            print(" - look up the start and end 'data_point' in the raw DataFrame")
            print("")
            print(
                "(Just remember to run make_step_table with the all_steps set to True"
                " before you do it. And re-run it with all_steps set to False afterwards"
                " to ensure that it behaves normally)"
            )
            return

        test = self.data.raw

        if not isinstance(step, (list, tuple)):
            step = [step]

        return test.loc[
            (test[cycle_index_header] == cycle) & (test[step_index_header].isin(step)),
            header,
        ].reset_index(drop=True)

    def sget_timestamp(self, cycle, step):
        """Returns timestamp for cycle, step.

        Convenience function; same as issuing
           raw[(raw[cycle_index_header] == cycle) &
                 (raw[step_index_header] == step)][timestamp_header]

        Args:
            cycle: cycle number
            step: step number (can be a list of several step numbers)

        Returns:
            pandas.Series
        """

        header = self.headers_normal.test_time_txt
        return self._sget(cycle, step, header, usteps=False)

    def sget_step_numbers(self, cycle, step):
        """Returns step number for cycle, step.

        Convenience function; same as issuing
           raw[(raw[cycle_index_header] == cycle) &
                 (raw[step_index_header] == step)][step_index_header]

        Args:
            cycle: cycle number
            step: step number (can be a list of several step numbers)

        Returns:
            pandas.Series
        """

        header = self.headers_normal.step_index_txt
        return self._sget(cycle, step, header, usteps=False)

    def get_dcap(
        self,
        cycle=None,
        converter=None,
        mode="gravimetric",
        as_frame=True,
        **kwargs,
    ):
        """Returns discharge-capacity and voltage for the selected cycle
        Args:
            cycle (int): cycle number.
            converter (float): a multiplication factor that converts the values to specific values (i.e.
                from Ah to mAh/g). If not provided (or None), the factor is obtained from the
                self.get_converter_to_specific() method.
            mode (string): 'gravimetric', 'areal' or 'absolute'. Defaults to 'gravimetric'.
                Used if converter is not provided (or None).
            as_frame (bool): if True: returns pd.DataFrame instead of capacity, voltage series.
            **kwargs:
        Returns:
            discharge_capacity, voltage (pd.Series or pd.DataFrame if return_dataframe is True).

        """

        if converter is None:
            converter = self.get_converter_to_specific(mode=mode)

        dc, v = self._get_cap(cycle, "discharge", converter=converter, **kwargs)
        if as_frame:
            cycle_df = pd.concat([v, dc], axis=1)
            return cycle_df
        else:
            return dc, v

    def get_ccap(
        self,
        cycle=None,
        converter=None,
        mode="gravimetric",
        as_frame=True,
        **kwargs,
    ):
        """Returns charge-capacity and voltage for the selected cycle.
        Args:
            cycle (int): cycle number.
            converter (float): a multiplication factor that converts the values to specific values (i.e.
                from Ah to mAh/g). If not provided (or None), the factor is obtained from the
                self.get_converter_to_specific() method.
            mode (string): 'gravimetric', 'areal' or 'absolute'. Defaults to 'gravimetric'.
                Used if converter is not provided (or None).
            as_frame (bool): if True: returns pd.DataFrame instead of capacity, voltage series.
        Returns:
            charge_capacity, voltage (pandas.Series or pandas.DataFrame if return_dataframe is True).

        """

        if converter is None:
            converter = self.get_converter_to_specific(mode=mode)
        cc, v = self._get_cap(cycle, "charge", converter=converter, **kwargs)

        if as_frame:
            cycle_df = pd.concat([v, cc], axis=1)
            return cycle_df
        else:
            return cc, v

    def get_cap(
        self,
        cycle=None,
        cycles=None,
        method="back-and-forth",
        insert_nan=None,
        shift=0.0,
        categorical_column=False,
        label_cycle_number=False,
        split=False,
        interpolated=False,
        dx=0.1,
        number_of_points=None,
        ignore_errors=True,
        dynamic=False,
        inter_cycle_shift=True,
        interpolate_along_cap=False,
        capacity_then_voltage=False,
        mode="gravimetric",
        mass=None,
        area=None,
        volume=None,
        cycle_mode=None,
        **kwargs,
    ):
        """Gets the capacity for the run.

        Args:
            cycle (int, list): cycle number (s).
            cycles (list): list of cycle numbers.
            method (string): how the curves are given
                "back-and-forth" - standard back and forth; discharge
                (or charge) reversed from where charge (or discharge) ends.
                "forth" - discharge (or charge) continues along x-axis.
                "forth-and-forth" - discharge (or charge) also starts at 0
                (or shift if not shift=0.0)
            insert_nan (bool): insert a np.nan between the charge and discharge curves.
                Defaults to True for "forth-and-forth", else False
            shift: start-value for charge (or discharge) (typically used when
                plotting shifted-capacity).
            categorical_column: add a categorical column showing if it is
                charge or discharge.
            label_cycle_number (bool): add column for cycle number
                (tidy format).
            split (bool): return a list of c and v instead of the default
                that is to return them combined in a DataFrame. This is only
                possible for some specific combinations of options (neither
                categorical_column=True or label_cycle_number=True are
                allowed).
            interpolated (bool): set to True if you would like to get
                interpolated data (typically if you want to save disk space
                or memory). Defaults to False.
            dx (float): the step used when interpolating.
            number_of_points (int): number of points to use (over-rides dx)
                for interpolation (i.e. the length of the interpolated data).
            ignore_errors (bool): don't break out of loop if an error occurs.
            dynamic: for dynamic retrieving data from cellpy-file.
                [NOT IMPLEMENTED YET]
            inter_cycle_shift (bool): cumulative shifts between consecutive
                cycles. Defaults to True.
            interpolate_along_cap (bool): interpolate along capacity axis instead
                of along the voltage axis. Defaults to False.
            capacity_then_voltage (bool): return capacity and voltage instead of
                voltage and capacity. Defaults to False.
            mode (string): 'gravimetric', 'areal', 'volumetric' or 'absolute'. Defaults
                to 'gravimetric'.
            mass (float): mass of active material (in set cellpy unit, typically mg).
            area (float): area of electrode (in set cellpy units, typically cm2).
            volume (float): volume of electrode (in set cellpy units, typically cm3).
            cycle_mode (string): if 'anode' the first step is assumed to be the discharge,
                else charge (defaults to CellpyCell.cycle_mode).
            **kwargs: sent to get_ccap and get_dcap.

        Returns:
            pandas.DataFrame ((cycle) voltage, capacity, (direction (-1, 1)))
            unless split is explicitly set to True. Then it returns a tuple
            with capacity and voltage.
        """

        # TODO: allow for fixing the interpolation range (so that it is possible
        #   to run the function on several cells and have a common x-axis)

        # if cycle is not given, then this function should
        # iterate through cycles

        def last(df):
            try:
                return df.iat[-1]
            except TypeError:
                return 0.0

        def first(df):
            try:
                return df.iat[0]
            except TypeError:
                return 0.0

        experimental = True

        cycle_mode = cycle_mode or self.cycle_mode

        if cycles is not None:
            cycle = cycles

        if cycle is None:
            cycle = self.get_cycle_numbers()

        if not isinstance(cycle, collections.abc.Iterable):
            cycle = [cycle]

        if split and not (categorical_column or label_cycle_number):
            return_dataframe = False
        else:
            return_dataframe = True

        method = method.lower()
        if method not in ["back-and-forth", "forth", "forth-and-forth"]:
            warnings.warn(
                f"method '{method}' is not a valid option "
                f"- setting to 'back-and-forth'"
            )
            method = "back-and-forth"

        if insert_nan is None:
            if method == "forth-and-forth":
                insert_nan = True
            else:
                insert_nan = False

        if insert_nan:
            _nan = pd.DataFrame({"capacity": [np.nan], "voltage": [np.nan]})

        converter_kwargs = dict()
        if mass is not None:
            logging.info(
                f"mass of {mass} {self.cellpy_units['mass']} given - using gravimetric mode"
            )
            converter_kwargs["value"] = mass
            mode = "gravimetric"

        if area is not None:
            logging.info(
                f"area of {area} {self.cellpy_units['area']} given - using areal mode"
            )
            converter_kwargs["value"] = area
            mode = "areal"

        if volume is not None:
            logging.info(
                f"volume of {volume} {self.cellpy_units['volume']} given - using volumetric mode"
            )
            converter_kwargs["value"] = volume
            mode = "volumetric"

        if mode == "absolute":
            logging.info(f"absolute mode - no conversion")

        capacity = None
        voltage = None
        specific_converter = self.get_converter_to_specific(
            mode=mode, **converter_kwargs
        )
        cycle_df = pd.DataFrame()

        initial = True
        for current_cycle in cycle:
            cc = pd.DataFrame()
            cv = pd.DataFrame()
            dc = pd.DataFrame()
            dv = pd.DataFrame()
            error = False
            try:
                cc, cv = self.get_ccap(
                    current_cycle,
                    converter=specific_converter,
                    as_frame=False,
                    **kwargs,
                )

            except NullData as e:
                error = True
                logging.info(e)
                if not ignore_errors:
                    logging.debug("breaking out of loop")
                    break

            try:
                dc, dv = self.get_dcap(
                    current_cycle,
                    converter=specific_converter,
                    as_frame=False,
                    **kwargs,
                )

            except NullData as e:
                error = True
                logging.info(e)
                if not ignore_errors:
                    logging.debug("breaking out of loop")
                    break

            if not error or experimental:
                if cc.empty:
                    logging.debug("get_ccap returns empty cc Series")

                if dc.empty:
                    logging.debug("get_ccap returns empty dc Series")

                if initial:
                    prev_end = shift
                    initial = False

                if cycle_mode == "anode":
                    first_interpolation_direction = -1
                    _first_step_c = dc
                    _first_step_v = dv
                    last_interpolation_direction = 1
                    _last_step_c = cc
                    _last_step_v = cv
                else:
                    first_interpolation_direction = 1
                    _first_step_c = cc
                    _first_step_v = cv
                    last_interpolation_direction = -1
                    _last_step_c = dc
                    _last_step_v = dv

                if method == "back-and-forth":
                    _last = last(_first_step_c)
                    _first = None
                    _new_first = None

                    if not inter_cycle_shift:
                        prev_end = 0.0

                    if _last_step_c is not None:
                        _last_step_c = _last - _last_step_c + prev_end

                    else:
                        logging.info("no last charge step found")

                    if _first_step_c is not None:
                        _first = first(_first_step_c)
                        _first_step_c += prev_end
                        _new_first = first(_first_step_c)

                    else:
                        logging.info("probably empty (_first_step_c is None)")
                    prev_end = last(_last_step_c)

                elif method == "forth":
                    _last = last(_first_step_c)
                    if _last_step_c is not None:
                        _last_step_c += _last + prev_end
                    else:
                        logging.debug("no last charge step found")
                    if _first_step_c is not None:
                        _first_step_c += prev_end
                    else:
                        logging.debug("no first charge step found")

                    if inter_cycle_shift:
                        prev_end = last(_last_step_c)
                    else:
                        prev_end = 0.0

                elif method == "forth-and-forth":
                    if _last_step_c is not None:
                        _last_step_c += shift
                    else:
                        logging.debug("no last charge step found")
                    if _first_step_c is not None:
                        _first_step_c += shift
                    else:
                        logging.debug("no first charge step found")

                if return_dataframe:
                    x_col = "voltage"
                    y_col = "capacity"
                    if interpolate_along_cap:
                        x_col, y_col = y_col, x_col

                    try:
                        # processing first
                        if not _first_step_c.empty:
                            _first_df = pd.DataFrame(
                                {
                                    "voltage": _first_step_v,
                                    "capacity": _first_step_c,
                                }
                            )
                            if interpolated:
                                _first_df = interpolate_y_on_x(
                                    _first_df,
                                    y=y_col,
                                    x=x_col,
                                    dx=dx,
                                    number_of_points=number_of_points,
                                    direction=first_interpolation_direction,
                                )
                            if insert_nan:
                                _first_df = pd.concat([_first_df, _nan])
                            if categorical_column:
                                _first_df["direction"] = -1
                        else:
                            _first_df = pd.DataFrame()

                        # processing last
                        if not _last_step_c.empty:
                            _last_df = pd.DataFrame(
                                {
                                    "voltage": _last_step_v.values,
                                    "capacity": _last_step_c.values,
                                }
                            )
                            if interpolated:
                                _last_df = interpolate_y_on_x(
                                    _last_df,
                                    y=y_col,
                                    x=x_col,
                                    dx=dx,
                                    number_of_points=number_of_points,
                                    direction=last_interpolation_direction,
                                )
                            if insert_nan:
                                _last_df = pd.concat([_last_df, _nan])
                            if categorical_column:
                                _last_df["direction"] = 1
                        else:
                            _last_df = pd.DataFrame()

                        if interpolate_along_cap:
                            if method == "forth":
                                _first_df = _first_df.loc[::-1].reset_index(drop=True)
                            elif method == "back-and-forth":
                                _first_df = _first_df.loc[::-1].reset_index(drop=True)
                                _last_df = _last_df.loc[::-1].reset_index(drop=True)

                    except AttributeError:
                        logging.info(f"Could not extract cycle {current_cycle}")
                    else:
                        c = pd.concat([_first_df, _last_df], axis=0)
                        if label_cycle_number:
                            c.insert(0, "cycle", current_cycle)
                            # c["cycle"] = current_cycle
                            # c = c[["cycle", "voltage", "capacity", "direction"]]
                        if cycle_df.empty:
                            cycle_df = c
                        else:
                            cycle_df = pd.concat([cycle_df, c], axis=0)
                    if capacity_then_voltage:
                        cols = cycle_df.columns.to_list()
                        new_cols = [
                            cols.pop(cols.index("capacity")),
                            cols.pop(cols.index("voltage")),
                        ]
                        new_cols.extend(cols)
                        cycle_df = cycle_df[new_cols]
                else:
                    logging.warning("returning non-dataframe")
                    _non_empty_c = []
                    _non_empty_v = []
                    if not _first_step_c.empty:
                        _non_empty_c.append(_first_step_c)
                        _non_empty_v.append(_first_step_v)
                    if not _last_step_c.empty:
                        _non_empty_c.append(_last_step_c)
                        _non_empty_v.append(_last_step_v)

                    c = pd.concat(_non_empty_c, axis=0)
                    v = pd.concat(_non_empty_v, axis=0)

                    if not c.empty:
                        capacity = pd.concat([capacity, c], axis=0)
                        voltage = pd.concat([voltage, v], axis=0)

        if return_dataframe:
            return cycle_df
        else:
            return capacity, voltage

    def _get_cap(
        self,
        cycle=None,
        cap_type="charge",
        trim_taper_steps=None,
        steps_to_skip=None,
        steptable=None,
        converter=None,
        usteps=False,
    ):
        if usteps:
            print(
                "Unfortunately, the ustep functionality is not implemented in this version of cellpy"
            )
            raise NotImplementedError("ustep == True not allowed!")
        # used when extracting capacities (get_ccap, get_dcap)
        # TODO: @jepe - does not allow for constant voltage yet?

        test = self.data

        if cap_type == "charge_capacity":
            cap_type = "charge"
        elif cap_type == "discharge_capacity":
            cap_type = "discharge"

        cycles = self.get_step_numbers(
            steptype=cap_type,
            allctypes=False,
            cycle_number=cycle,
            trim_taper_steps=trim_taper_steps,
            steps_to_skip=steps_to_skip,
            steptable=steptable,
        )
        if cap_type == "charge":
            column_txt = self.headers_normal.charge_capacity_txt
        else:
            column_txt = self.headers_normal.discharge_capacity_txt
        if cycle:
            steps = cycles[cycle]
            _v = []
            _c = []
            if len(set(steps)) < len(steps) and not usteps:
                raise ValueError(f"You have duplicate step numbers!")
            for step in sorted(steps):
                selected_step = self._select_step(cycle, step)
                if not self._is_empty_array(selected_step):
                    _v.append(selected_step[self.headers_normal.voltage_txt])
                    _c.append(selected_step[column_txt] * converter)
            try:
                voltage = pd.concat(_v, axis=0)
                cap = pd.concat(_c, axis=0)
            except Exception:
                logging.debug("could not find any steps for this cycle")
                raise NullData(f"no steps found (c:{cycle} s:{step} type:{cap_type})")
        else:
            # get all the discharge cycles
            # this is a dataframe filtered on step and cycle
            # This functionality is not crucial since get_cap (that uses this method) has it
            # (but it might be nice to improve performance)
            raise NotImplementedError(
                "Not yet possible to extract without giving cycle numbers (use get_cap instead)"
            )
        return cap, voltage

    def get_ocv(
        self,
        cycles=None,
        direction="up",
        remove_first=False,
        interpolated=False,
        dx=None,
        number_of_points=None,
    ) -> pd.DataFrame:
        """get the open circuit voltage relaxation curves.

        Args:
            cycles (list of ints or None): the cycles to extract from
                (selects all if not given).
            direction ("up", "down", or "both"): extract only relaxations that
                is performed during discharge for "up" (because then the
                voltage relaxes upwards) etc.
            remove_first: remove the first relaxation curve (typically,
                the first curve is from the initial rest period between
                assembling the data to the actual testing/cycling starts)
            interpolated (bool): set to True if you want the data to be
                interpolated (e.g. for creating smaller files)
            dx (float): the step used when interpolating.
            number_of_points (int): number of points to use (over-rides dx)
                for interpolation (i.e. the length of the interpolated data).

        Returns:
            A pandas.DataFrame with cycle-number, step-number, step-time, and
                voltage columns.
        """
        # TODO: use proper column header pickers
        if cycles is None:
            cycles = self.get_cycle_numbers()
        else:
            if not isinstance(cycles, (list, tuple, np.ndarray)):
                cycles = [cycles]
            else:
                remove_first = False

        ocv_rlx_id = "ocvrlx"
        if direction == "up":
            ocv_rlx_id += "_up"
        elif direction == "down":
            ocv_rlx_id += "_down"

        steps = self.data.steps
        raw = self.data.raw

        ocv_steps = steps.loc[steps["cycle"].isin(cycles), :]

        ocv_steps = ocv_steps.loc[
            ocv_steps.type.str.startswith(ocv_rlx_id, na=False), :
        ]

        if remove_first:
            ocv_steps = ocv_steps.iloc[1:, :]

        step_time_label = self.headers_normal.step_time_txt
        voltage_label = self.headers_normal.voltage_txt
        cycle_label = self.headers_normal.cycle_index_txt
        step_label = self.headers_normal.step_index_txt

        selected_df = raw.loc[
            (
                raw[cycle_label].isin(ocv_steps.cycle)
                & raw[step_label].isin(ocv_steps.step)
            ),
            [cycle_label, step_label, step_time_label, voltage_label],
        ]

        if interpolated:
            if dx is None and number_of_points is None:
                dx = prms.Reader.time_interpolation_step
            new_dfs = list()
            groupby_list = [cycle_label, step_label]

            for name, group in selected_df.groupby(groupby_list):
                new_group = interpolate_y_on_x(
                    group,
                    x=step_time_label,
                    y=voltage_label,
                    dx=dx,
                    number_of_points=number_of_points,
                )

                for i, j in zip(groupby_list, name):
                    new_group[i] = j
                new_dfs.append(new_group)

            selected_df = pd.concat(new_dfs)

        return selected_df

    def get_number_of_cycles(self, steptable=None):
        """Get the number of cycles in the test."""
        if steptable is None:
            d = self.data.raw
            number_of_cycles = np.amax(d[self.headers_normal.cycle_index_txt])
        else:
            number_of_cycles = np.amax(steptable[self.headers_step_table.cycle])
        return number_of_cycles

    def get_rates(self, steptable=None, agg="first", direction=None):
        """
        Get the rates in the test (only valid for constant current).
        Args:
            steptable: provide custom steptable (if None, the steptable from the cellpydata object will be used).
            agg (str): perform an aggregation if more than one step of charge or discharge is found
                (e.g. "mean", "first", "max"). For example, if agg='mean', the average rate for each cycle
                will be returned. Set to None if you want to keep all the rates.
            direction (str or list of str): only select rates for this direction (e.g. "charge" or "discharge").

        Returns:
            pandas.DataFrame with cycle, type, and rate_avr (i.e. C-rate) columns.
        """

        if steptable is None:
            steptable = self.data.steps
        rates = steptable[
            [
                self.headers_step_table.cycle,
                self.headers_step_table.type,
                self.headers_step_table.rate_avr,
            ]
        ].dropna()

        if agg:
            rates = (
                rates.groupby(
                    [self.headers_step_table.cycle, self.headers_step_table.type]
                )
                .agg(agg)
                .reset_index()
            )

        if direction is not None:
            if not isinstance(direction, (list, tuple)):
                direction = [direction]
            rates = rates.loc[rates[self.headers_step_table.type].isin(direction), :]

        return rates

    def get_cycle_numbers(
        self,
        steptable=None,
        rate=None,
        rate_on=None,
        rate_std=None,
        rate_agg="first",
        inverse=False,
    ):
        """Get a array containing the cycle numbers in the test.

        Parameters:
            steptable (pandas.DataFrame): the step-table to use (if None, the step-table
                from the cellpydata object will be used).
            rate (float): the rate to filter on. Remark that it should be given
                as a float, i.e. you will have to convert from C-rate to
                the actual numeric value. For example, use rate=0.05 if you want
                to filter on cycles that has a C/20 rate.
            rate_on (str): only select cycles if based on the rate of this step-type (e.g. on="discharge").
            rate_std (float): allow for this inaccuracy in C-rate when selecting cycles
            rate_agg (str): perform an aggregation on rate if more than one step of charge or discharge is found
                (e.g. "mean", "first", "max"). For example, if agg='mean', the average rate for each cycle
                will be returned. Set to None if you want to keep all the rates.
            inverse (bool): select steps that does not have the given C-rate.

        Returns:
            numpy.ndarray of cycle numbers.
        """

        # TODO: add support for selecting cycles based on other criteria (for example, based on the
        #   existence of particular step-types, or max, min values of current, voltage, etc)

        logging.debug("getting cycle numbers")

        if steptable is None:
            d = self.data.raw
            cycles = d[self.headers_normal.cycle_index_txt].dropna().unique()
            steptable = self.data.steps
        else:
            logging.debug("steptable is given as input parameter")
            cycles = steptable[self.headers_step_table.cycle].dropna().unique()

        if rate is None:
            return cycles

        logging.debug("filtering on rate")

        if rate is None:
            rate = 0.05

        if rate_std is None:
            rate_std = 0.1 * rate

        if rate_on is None:
            rate_on = ["charge", "discharge"]
        rates = self.get_rates(steptable=steptable, agg=rate_agg, direction=rate_on)
        rate_column = self.headers_step_table.rate_avr
        cycles_mask = (rates[rate_column] < (rate + rate_std)) & (
            rates[rate_column] > (rate - rate_std)
        )

        if inverse:
            cycles_mask = ~cycles_mask

        filtered_rates = rates[cycles_mask]
        filtered_cycles = filtered_rates[self.headers_step_table["cycle"]].unique()

        return filtered_cycles

    def get_ir(self):
        """Get the IR data (Deprecated)."""
        raise DeprecatedFeature

    def has_data_point_as_index(self):
        """Check if the raw data has data_point as index."""
        return self.data.raw.index.name == self.headers_normal.data_point_txt

    def has_data_point_as_column(self):
        """Check if the raw data has data_point as column."""
        return self.headers_normal.data_point_txt in self.data.raw.columns

    def has_no_full_duplicates(self):
        """Check if the raw data has no full duplicates."""
        return not self.data.raw.duplicated().any()

    def has_no_partial_duplicates(self, subset="data_point"):
        """Check if the raw data has no partial duplicates."""
        if subset == "data_point":
            subset = self.headers_normal.data_point_txt
        return not self.data.raw.duplicated(subset=subset).any()

    def total_time_at_voltage_level(
        self,
        cycles=None,
        voltage_limit=0.5,
        sampling_unit="S",
        at="low",
    ):
        """Experimental method for getting the total time spent at low / high voltage.

        Args:
            cycles: cycle number (all cycles if None).
            voltage_limit: voltage limit (default 0.5 V). Can be a tuple (low, high) if at="between".
            sampling_unit: sampling unit (default "S")
                    H: hourly frequency
                    T, min: minutely frequency
                    S: secondly frequency
                    L, ms:  milliseconds
                    U, us: microseconds
                    N: nanoseconds
            at (str): "low", "high", or "between" (default "low")
        """

        from pandas.api.types import is_datetime64_any_dtype as is_datetime

        if at not in ["low", "high", "between"]:
            raise ValueError("at must be either 'low', 'high', or 'between'")

        if sampling_unit not in ["S"]:
            logging.critical("Only 'S' (seconds) has been tested so far.")
            logging.critical(
                f"It might work with sampling_unit='{sampling_unit}'"
                f"however, the result you get is probably in {sampling_unit} and not"
                f"seconds."
            )

        date_time_hdr = "date_time"
        cycle_index_hdr = "cycle_index"
        voltage_hdr = "voltage"
        date_time_format = prms._date_time_format

        if cycles is not None:
            if not isinstance(cycles, (list, tuple)):
                cycles = [cycles]
            v = self.data.raw.loc[
                self.data.raw[cycle_index_hdr].isin(cycles),
                [date_time_hdr, cycle_index_hdr, voltage_hdr],
            ].copy()
        else:
            v = self.data.raw[[date_time_hdr, cycle_index_hdr, voltage_hdr]].copy()

        # make sure data_time is datetime64[ns] (not sure if this works for all tester formats/loaders):
        col_has_date_time_dtype = is_datetime(v[date_time_hdr])
        duplicated = v[date_time_hdr].duplicated().any()

        if not col_has_date_time_dtype:
            logging.debug("converting date_time to datetime64[ns]")
            v[date_time_hdr] = pd.to_datetime(v[date_time_hdr], format=date_time_format)

        if duplicated:
            logging.debug("removing duplicated date_time values")
            v = v.loc[~v[date_time_hdr].duplicated(), :]

        v = v.set_index(date_time_hdr, drop=True)
        v = v.resample(sampling_unit).ffill().bfill()
        v["is_at_target"] = 0

        if at == "low":
            v.loc[v[voltage_hdr] < voltage_limit, "is_at_target"] = 1
        elif at == "high":
            v.loc[v[voltage_hdr] > voltage_limit, "is_at_target"] = 1
        elif at == "between":
            v.loc[
                (v[voltage_hdr] > voltage_limit[0])
                & (v[voltage_hdr] < voltage_limit[1]),
                "is_at_target",
            ] = 1
        else:
            # This will never occur, but keeping it here for completeness
            raise ValueError("at must be either 'low' or 'high'")

        # missing option - convert to seconds
        return v.is_at_target.sum()

    def nominal_capacity_as_absolute(
        self,
        value=None,
        specific=None,
        nom_cap_specifics=None,
        convert_charge_units=False,
    ):
        """Get the nominal capacity as absolute value."""

        # TODO: implement handling of edge-cases (i.e. the raw capacity is not in absolute values)
        if self.debug:
            print("nominal_capacity_as_absolute".center(80, "="))
            print(f"{value=}")
            print(f"{specific=}")
            print(f"{nom_cap_specifics=}")
            print(f"{convert_charge_units=}")
            print(80 * "-")
        if nom_cap_specifics is None:
            nom_cap_specifics = "gravimetric"

        if specific is None:
            if nom_cap_specifics == "gravimetric":
                specific = self.data.mass
            elif nom_cap_specifics == "areal":
                specific = self.data.active_electrode_area

            # TODO: implement volumetric
            elif nom_cap_specifics == "volumetric":
                raise NotImplementedError("volumetric not implemented yet")

        if value is None:
            value = self.data.nom_cap

        value = Q(value, self.cellpy_units["nominal_capacity"])

        if nom_cap_specifics == "gravimetric":
            specific = Q(specific, self.cellpy_units["mass"])
        elif nom_cap_specifics == "areal":
            specific = Q(specific, self.cellpy_units["area"])

        # TODO: implement volumetric
        elif nom_cap_specifics == "volumetric":
            raise NotImplementedError("volumetric not implemented yet")

        if convert_charge_units:
            conversion_factor_charge = Q(1, self.cellpy_units["charge"]) / Q(
                1, self.data.raw_units["charge"]
            )
        else:
            conversion_factor_charge = 1.0

        try:
            absolute_value = (
                (value * conversion_factor_charge * specific)
                .to_reduced_units()
                .to("Ah")
            )
        except DimensionalityError as e:
            print(" DimensionalityError ".center(80, "="))
            print("Could not convert nominal capacity to absolute value!")
            print(
                "This is probably because the nominal capacity is given in "
                "different unit than the given specifics."
            )
            print(
                " - Maybe you have given nominal capacity in mAh/cm**2 and your "
                "specifics is set to 'gravimetric'?"
            )
            print(
                " - Maybe you have given nominal capacity in mAh/g and your "
                "specifics is set to 'areal'?"
            )
            print("Please check your input parameters!")
            print(
                "\n[hint 1] try to set the parameter 'nom_cap_specifics' in the get function:\n"
            )
            print(
                "    c = cellpy.get(filename, area=1.55, nom_cap='1.2 mAh/cm**2', nom_cap_specifics='areal')"
            )
            print(
                "\n[hint 2] try to set it on the cellpy object directly after loading, "
                "\n  but before processing (making the step-table etc):\n"
            )
            print("    c = cellpy.get(filename, auto_summary=False)")
            print("    c.nom_cap_specifics = 'areal'")
            print("    ... # set other stuff if needed")
            print("    c.make_step_table()")
            print("    c.make_summary()")
            print("\nRe-raising the exception.")
            print(80 * "=")
            raise e

        if self.debug:
            print(f"{self.mass=}")
            print(f"{self.active_electrode_area=}")
            print(f"{self.nom_cap=}")
            print(f"{self.cellpy_units=}")
            print(
                f"nominal capacity: {value} [{self.cellpy_units.nominal_capacity}] -> {absolute_value:.3f} [Ah]"
            )
            print(80 * "=")
        return absolute_value.m

    def with_cellpy_unit(self, parameter, as_str=False):
        """Return quantity as `pint.Quantity` object."""
        _look_up = {
            "nom_cap": "nominal_capacity",
            "active_electrode_area": "area",
        }
        _parameter = parameter
        if parameter in _look_up.keys():
            _parameter = _look_up[parameter]

        try:
            _unit = self.cellpy_units[_parameter]
        except KeyError:
            print(f"Did not find any units registered for {parameter}")
            return

        try:
            _value = getattr(self.data, parameter)
        except AttributeError:
            print(
                f"{parameter} is not a valid cellpy data attribute (but the unit is {_unit})"
            )
            return

        if as_str:
            return f"{_value} {_unit}"

        return Q(_value, _unit)

    def to_cellpy_unit(self, value, physical_property):
        """Convert value to cellpy units.

        Args:
            value (numeric, pint.Quantity or str): what you want to convert from
            physical_property (str): What this value is a measure of
                (must correspond to one of the keys in the CellpyUnits class).

        Returns (numeric):
            the value in cellpy units
        """
        logging.debug(f"value {value} is numeric? {isinstance(value, numbers.Number)}")
        logging.debug(
            f"value {value} is a pint quantity? {isinstance(value, Quantity)}"
        )

        if not isinstance(value, Quantity):
            if isinstance(value, numbers.Number):
                try:
                    value = Q(value, self.data.raw_units[physical_property])
                    logging.debug(f"With unit from raw-units: {value}")
                except NoDataFound:
                    raise NoDataFound(
                        "If you dont have any cells you cannot convert"
                        " values to cellpy units without providing what"
                        " unit to convert from!"
                    )
                except KeyError as e:
                    raise KeyError(
                        "You have to provide a valid physical_property"
                    ) from e
            elif isinstance(value, tuple):
                value = Q(*value)
            else:
                value = Q(value)

        value = value.to(self.cellpy_units[physical_property])

        return value.m

    def unit_scaler_from_raw(self, unit, physical_property):
        """Get the conversion factor going from raw to given unit.

        Args:
            unit (str): what you want to convert to
            physical_property (str): what this value is a measure of
                (must correspond to one of the keys in the CellpyUnits class).

        Returns (numeric):
            conversion factor (scaler)
        """
        logging.debug(f"value {unit} is a pint quantity? {isinstance(unit, Quantity)}")

        old_unit = self.data.raw_units[physical_property]
        value = Q(1, old_unit)
        value = value.to(unit)
        return value.m

    def get_converter_to_specific(
        self,
        dataset: Data = None,
        value: float = None,
        from_units: CellpyUnits = None,
        to_units: CellpyUnits = None,
        mode: str = "gravimetric",
    ) -> float:
        """Convert from absolute units to specific (areal or gravimetric).

        Args:
            dataset: data instance
            value: value used to scale on.
            from_units: defaults to data.raw_units.
            to_units: defaults to cellpy_units.
            mode (str): gravimetric, areal or absolute

        Returns:
            conversion factor (multiply with this to get your values into specific values).
        """
        # TODO @jepe: implement handling of edge-cases
        # TODO @jepe: fix all the instrument readers (replace floats in raw_units with strings)
        if dataset is None:
            dataset = self.data

        new_units = to_units or self.cellpy_units
        old_units = from_units or dataset.raw_units

        if mode == "gravimetric":
            value = value or dataset.mass
            value = Q(value, new_units["mass"])
            to_unit_specific = Q(1.0, new_units["specific_gravimetric"])

        elif mode == "areal":
            value = value or dataset.active_electrode_area
            value = Q(value, new_units["area"])
            to_unit_specific = Q(1.0, new_units["specific_areal"])

        elif mode == "volumetric":
            value = value or dataset.volume
            value = Q(value, new_units["volume"])
            to_unit_specific = Q(1.0, new_units["specific_volumetric"])

        elif mode == "absolute":
            value = Q(1.0, None)
            to_unit_specific = Q(1.0, None)

        elif mode is None:
            return 1.0

        else:
            logging.debug(f"mode={mode} not supported!")
            return 1.0

        from_unit_cap = Q(1.0, old_units["charge"])
        to_unit_cap = Q(1.0, new_units["charge"])

        # from unit is always in absolute values:
        from_unit = from_unit_cap

        to_unit = to_unit_cap / to_unit_specific

        conversion_factor = (from_unit / to_unit / value).to_reduced_units()
        logging.debug(f"conversion factor: {conversion_factor}")
        return conversion_factor.m

    def get_diagnostics_plots(self, scaled=False):
        raise DeprecatedFeature(
            "This feature is deprecated. "
            "Extract diagnostics from the summary instead."
        )

    def _set_mass(self, value):
        # TODO: replace with setter
        try:
            self.data.meta_common.mass = value
        except AttributeError as e:
            logging.info("This test is empty")
            logging.info(e)

    def _set_tot_mass(self, value):
        # TODO: replace with setter
        try:
            self.data.meta_common.tot_mass = value
        except AttributeError as e:
            logging.info("This test is empty")
            logging.info(e)

    def _set_nom_cap(self, value):
        # TODO: replace with setter
        try:
            self.data.meta_common.nom_cap = value
        except AttributeError as e:
            logging.info("This test is empty")
            logging.info(e)

    def _set_run_attribute(self, attr, val, validated=None):
        # Sets the val (vals) for the test (datasets).
        # Remark! This is left-over code from old ages when we thought we needed
        #   to have data-sets with multiple cells. And before we learned about
        #   setters and getters in Python. Feel free to refactor it.

        # TODO: deprecate it

        if attr == "mass":
            setter = self._set_mass
        elif attr == "tot_mass":
            setter = self._set_tot_mass
        elif attr == "nom_cap":
            setter = self._set_nom_cap

        if not self.data:
            logging.info("No datasets have been loaded yet")
            logging.info(f"Cannot set {attr} before loading datasets")
            sys.exit(-1)

        if validated is None:
            setter(val)
        else:
            if validated:
                setter(val)
            else:
                logging.debug("_set_run_attribute: this set is empty")

    def set_mass(self, mass, validated=None):
        """Sets the mass (masses) for the test (datasets)."""

        warnings.warn(
            "This function is deprecated. Use the setter instead (mass = value).",
            DeprecationWarning,
            stacklevel=2,
        )
        self._set_run_attribute("mass", mass, validated=validated)

    def set_tot_mass(self, mass, validated=None):
        warnings.warn(
            "This function is deprecated. Use the setter instead (tot_mass = value).",
            DeprecationWarning,
            stacklevel=2,
        )

        self._set_run_attribute("tot_mass", mass, validated=validated)

    def set_nom_cap(self, nom_cap, validated=None):
        warnings.warn(
            "This function is deprecated. Use the setter instead (nom_cap = value).",
            DeprecationWarning,
            stacklevel=2,
        )

        self._set_run_attribute("nom_cap", nom_cap, validated=validated)

    @staticmethod
    def set_col_first(df, col_names):
        """set selected columns first in a pandas.DataFrame.

        This function sets cols with names given in  col_names (a list) first in
        the DataFrame. The last col in col_name will come first (processed last)
        """

        column_headings = df.columns
        column_headings = column_headings.tolist()
        try:
            for col_name in col_names:
                i = column_headings.index(col_name)
                column_headings.pop(column_headings.index(col_name))
                column_headings.insert(0, col_name)

        finally:
            df = df.reindex(columns=column_headings)
            return df

    def get_summary(self, use_summary_made=False):
        """Retrieve summary returned as a pandas DataFrame."""
        cell = self.data

        # This is a bit convoluted; in the old days, we used an attribute
        # called summary_made,
        # that was set to True when the summary was made successfully.
        # It is most likely never
        # used anymore. And will most probably be deleted.

        warnings.warn(
            "get_summary is deprecated. Use the CellpyCell.data.summary property instead.",
            DeprecationWarning,
        )

        if use_summary_made:
            summary_made = cell.has_summary
        else:
            summary_made = True

        if not summary_made:
            warnings.warn("Summary is not made yet")
            return None
        else:
            logging.info("Returning datasets[test_no].summary")
            return cell.summary

    # -----------internal-helpers-----------------------------------------------

    @staticmethod
    def _is_empty_array(v):
        try:
            if not v:
                return True
            else:
                return False
        except Exception:
            try:
                if v.empty:
                    return True
                else:
                    return False
            except Exception:
                if v.isnull:
                    return False
                else:
                    return True

    @staticmethod
    def _is_listtype(x):
        if isinstance(x, (list, tuple)):
            return True
        else:
            return False

    @staticmethod
    def _bounds(x):
        return np.amin(x), np.amax(x)

    @staticmethod
    def _roundup(x):
        n = 1000.0
        x = np.ceil(x * n)
        x /= n
        return x

    def _rounddown(self, x):
        x = self._roundup(-x)
        x = -x
        return x

    @staticmethod
    def _reverse(x):
        x = x[::-1]
        # x = x.sort_index(ascending=True)
        return x

    def _select_y(self, x, y, points):
        # uses interpolation to select y = f(x)
        min_x, max_x = self._bounds(x)
        if x[0] > x[-1]:
            # need to reverse
            x = self._reverse(x)
            y = self._reverse(y)
        f = interpolate.interp1d(y, x)
        y_new = f(points)
        return y_new

    def _select_last(self, raw):
        # this legacy method gives a set of indexes pointing to the last
        # datapoints for each cycle in the dataset (only used in the old
        # summary method and for the new summary method if use_cellpy_stat_file is True)

        c_txt = self.headers_normal.cycle_index_txt
        d_txt = self.headers_normal.data_point_txt
        steps = []
        unique_steps = raw[c_txt].unique()
        max_step = max(raw[c_txt])
        for j in range(int(max_step)):
            if j + 1 not in unique_steps:
                logging.debug(f"Warning: Cycle {j + 1} is missing!")
            else:
                last_item = max(raw.loc[raw[c_txt] == j + 1, d_txt])
                steps.append(last_item)

        last_items = raw[d_txt].isin(steps)
        return last_items

    # TODO: @jepe - this method might be valuable for users and could be made
    #  public when it is fixed:
    def _select_without(self, exclude_types=None, exclude_steps=None, replace_nan=True):
        steps = self.data.steps
        raw = self.data.raw.copy()

        # unravel the headers:
        d_n_txt = self.headers_normal.data_point_txt
        v_n_txt = self.headers_normal.voltage_txt
        c_n_txt = self.headers_normal.cycle_index_txt
        s_n_txt = self.headers_normal.step_index_txt
        i_n_txt = self.headers_normal.current_txt
        ch_n_txt = self.headers_normal.charge_capacity_txt
        dch_n_txt = self.headers_normal.discharge_capacity_txt

        d_st_txt = self.headers_step_table.point
        v_st_txt = self.headers_step_table.voltage
        c_st_txt = self.headers_step_table.cycle
        i_st_txt = self.headers_step_table.current
        ch_st_txt = self.headers_step_table.charge
        dch_st_txt = self.headers_step_table.discharge
        t_st_txt = self.headers_step_table.type
        s_st_txt = self.headers_step_table.step

        _first = "_first"
        _last = "_last"
        _delta_label = "_diff"

        # TODO: implement also for energy and power (and probably others as well) - this will
        #  require changing step-table to also include energy and power etc. If implementing
        #  this, you should also include diff in the step-table. You should preferably also use this
        #  opportunity to also make both the headers in the tables as well as the names used for
        #  the headers more aligned (e.g. for header_normal.data_point_txt -> header_normal.point;
        #  "cycle_index" -> "cycle")

        # TODO: @jepe - this method might be a bit slow for large datasets - consider using
        #  more "native" pandas methods and get rid of all looping (need some timing to check first)

        last_data_points = (
            steps.loc[:, [c_st_txt, d_st_txt + _last]]
            .groupby(c_st_txt)
            .last()
            .values.ravel()
        )
        last_items = raw[d_n_txt].isin(last_data_points)
        selected = raw[last_items]

        if exclude_types is None and exclude_steps is None:
            return selected

        if not isinstance(exclude_types, (list, tuple)):
            exclude_types = [exclude_types]

        if not isinstance(exclude_steps, (list, tuple)):
            exclude_steps = [exclude_steps]

        q = None
        for exclude_type in exclude_types:
            _q = ~steps[t_st_txt].str.startswith(exclude_type)
            q = _q if q is None else q & _q

        if exclude_steps:
            _q = ~steps[t_st_txt].isin(exclude_steps)
            q = _q if q is None else q & _q

        _delta_columns = [
            i_st_txt,
            v_st_txt,
            ch_st_txt,
            dch_st_txt,
        ]
        _raw_columns = [
            i_n_txt,
            v_n_txt,
            ch_n_txt,
            dch_n_txt,
        ]
        _diff_columns = [f"{col}{_delta_label}" for col in _delta_columns]

        delta_first = [f"{col}{_first}" for col in _delta_columns]
        delta_last = [f"{col}{_last}" for col in _delta_columns]
        delta_columns = delta_first + delta_last

        delta = steps.loc[~q, [c_st_txt, d_st_txt + _last, *delta_columns]].copy()

        for col in _delta_columns:
            delta[col + _delta_label] = delta[col + _last] - delta[col + _first]
        delta = delta.drop(columns=delta_columns)
        delta = delta.groupby(c_st_txt).sum()
        delta = delta.reset_index()

        selected = selected.merge(delta, how="left", left_on=c_n_txt, right_on=c_st_txt)
        if replace_nan:
            selected = selected.fillna(0.0)

        for col_n, col_diff in zip(_raw_columns, _diff_columns):
            selected[col_n] -= selected[col_diff]
        selected = selected.drop(columns=_diff_columns)

        return selected

    # ----------making-summary------------------------------------------------------
    def make_summary(
        self,
        find_ir=False,
        find_end_voltage=True,
        use_cellpy_stat_file=None,
        ensure_step_table=True,
        remove_duplicates=True,
        normalization_cycles=None,
        nom_cap=None,
        nom_cap_specifics=None,
        old=False,
        create_copy=False,
        exclude_types=None,
        exclude_steps=None,
        selector_type=None,
        selector=None,
        **kwargs,
    ):
        """Convenience function that makes a summary of the cycling data.

        find_ir (bool): if True, the internal resistance will be calculated.
        find_end_voltage (bool): if True, the end voltage will be calculated.
        use_cellpy_stat_file (bool): if True, the summary will be made from
            the cellpy_stat file (soon to be deprecated).
        ensure_step_table (bool): if True, the step-table will be made if it does not exist.
        remove_duplicates (bool): if True, duplicates will be removed from the summary.
        normalization_cycles (int or list of int): cycles to use for normalization.
        nom_cap (float or str): nominal capacity (if None, the nominal capacity from the data will be used).
        nom_cap_specifics (str): gravimetric, areal, or volumetric.
        old (bool): if True, the old summary method will be used.
        create_copy (bool): if True, a copy of the cellpy object will be returned.
        exclude_types (list of str): exclude these types from the summary.
        exclude_steps (list of int): exclude these steps from the summary.
        selector_type (str): select based on type (e.g. "non-cv", "non-rest", "non-ocv", "only-cv").
        selector (callable): custom selector function.


        """
        # TODO: @jepe  - make it is possible to update only new data by implementing
        #  from_cycle (only calculate summary from a given cycle number).
        #  Probably best to keep the old summary and make
        #  a new one for the rest, then use pandas.concat to merge them.
        #  Might have to create the cumulative cols etc after merging?

        # first - check if we need some "instrument-specific" prms
        if ensure_step_table is None:
            ensure_step_table = self.ensure_step_table

        if use_cellpy_stat_file is None:
            use_cellpy_stat_file = prms.Reader.use_cellpy_stat_file
            logging.debug("using use_cellpy_stat_file from prms")
            logging.debug(f"use_cellpy_stat_file: {use_cellpy_stat_file}")

        txt = "creating summary for file "
        try:
            test = self.data
        except NoDataFound:
            logging.info(f"Empty test (no data found)")
            return

        if isinstance(test.loaded_from, (list, tuple)):
            for f in test.loaded_from:
                txt += f"{f}\n"
        else:
            txt += str(test.loaded_from)

        logging.debug(txt)

        if old:
            self._make_summar_legacy(
                # find_ocv=find_ocv,
                find_ir=find_ir,
                find_end_voltage=find_end_voltage,
                use_cellpy_stat_file=use_cellpy_stat_file,
                ensure_step_table=ensure_step_table,
                # add_c_rate=add_c_rate,
                normalization_cycles=normalization_cycles,
                nom_cap=nom_cap,
                nom_cap_specifics=nom_cap_specifics,
            )
            return self

        data = self._make_summary(
            find_ir=find_ir,
            find_end_voltage=find_end_voltage,
            use_cellpy_stat_file=use_cellpy_stat_file,
            ensure_step_table=ensure_step_table,
            remove_duplicates=remove_duplicates,
            normalization_cycles=normalization_cycles,
            nom_cap=nom_cap,
            nom_cap_specifics=nom_cap_specifics,
            create_copy=create_copy,
            exclude_types=exclude_types,
            exclude_steps=exclude_steps,
            selector_type=selector_type,
            selector=selector,
            **kwargs,
        )
        if create_copy:
            other = copy.deepcopy(self)
            other.data = data
            return other
        else:
            # TODO: check if anything is using this feature (returning self), if not, remove it.
            return self

    def _make_summary(
        self,
        mass=None,
        nom_cap=None,
        nom_cap_specifics="gravimetric",
        update_mass=False,
        select_columns=True,
        find_ir=True,
        find_end_voltage=False,
        ensure_step_table=True,
        remove_duplicates=True,
        sort_my_columns=True,
        use_cellpy_stat_file=False,
        normalization_cycles=None,
        create_copy=True,
        exclude_types=None,
        exclude_steps=None,
        selector_type=None,
        selector=None,
    ):
        # ---------------- discharge loss --------------------------------------
        # Assume that both charge and discharge is defined as positive.
        # The gain for cycle n (compared to cycle n-1)
        # is then cap[n] - cap[n-1]. The loss is the negative of gain.
        # discharge loss = discharge_cap[n-1] - discharge_cap[n]

        # ---------------- charge loss -----------------------------------------
        # charge loss = charge_cap[n-1] - charge_cap[n]

        # --------- shifted capacities ------------------------------------------
        #  as defined by J. Dahn et al.
        # Note! Should double-check this (including checking
        # if it is valid in cathode mode).

        # --------- relative irreversible capacities -----------------------------
        #  as defined by Gauthier et al.
        # RIC = discharge_cap[n-1] - charge_cap[n] /  charge_cap[n-1]
        # RIC_SEI = discharge_cap[n] - charge_cap[n-1] / charge_cap[n-1]
        # RIC_disconnect = charge_cap[n-1] - charge_cap[n] / charge_cap[n-1]

        # --------- notes --------------------------------------------------------
        # @jepe 2022.09.11: trying to use .assign from now on
        #   as it is recommended (but this will likely increase memory usage)

        if selector is None:
            if selector_type == "non-cv":
                exclude_types = ["cv_"]
            elif selector_type == "non-rest":
                exclude_types = ["rest_"]
            elif selector_type == "non-ocv":
                exclude_types = ["ocv_"]
            elif selector_type == "only-cv":
                exclude_types = ["charge", "discharge"]
            selector = functools.partial(
                self._select_without,
                exclude_types=exclude_types,
                exclude_steps=exclude_steps,
            )

        # TODO: add this to arguments and possible prms:
        if nom_cap_specifics is None:
            nom_cap_specifics = self.nom_cap_specifics
        specifics = ["gravimetric", "areal"]
        cycle_index_as_index = True
        time_00 = time.time()
        logging.debug("start making summary")

        if create_copy:
            data = copy.deepcopy(self.data)
        else:
            data = self.data

        if not mass:
            mass = data.mass or 1.0
        else:
            if update_mass:
                data.mass = mass

        if use_cellpy_stat_file:
            warnings.warn(
                "using cellpy 'statfile' - this feature is not properly supported anymore"
            )

        if nom_cap is None:
            nom_cap = data.nom_cap

        logging.info(f"Using the following nominal capacity: {nom_cap}")

        # cellpy has historically assumed that the nominal capacity (nom_cap) is specific gravimetric
        # (i.e. in units of for example mAh/g), but now we need it in absolute units (e.g. Ah). The plan
        # is to set stuff like this during initiation of the cell (but not yet)

        # generating absolute nominal capacity (this should be refactored):
        if nom_cap_specifics == "gravimetric":
            nom_cap_abs = self.nominal_capacity_as_absolute(
                nom_cap, mass, nom_cap_specifics
            )
        elif nom_cap_specifics == "areal":
            nom_cap_abs = self.nominal_capacity_as_absolute(
                nom_cap, data.active_electrode_area, nom_cap_specifics
            )

        # TODO: this will break because cell.volume (data.volume) is not set yet
        elif nom_cap_specifics == "volumetric":
            nom_cap_abs = self.nominal_capacity_as_absolute(
                nom_cap, data.volume, nom_cap_specifics
            )

        else:
            nom_cap_abs = self.nominal_capacity_as_absolute(
                nom_cap, mass, nom_cap_specifics
            )

        # ensuring that a step table exists:
        if ensure_step_table:
            logging.debug("ensuring existence of step-table")
            if not data.has_steps:
                logging.debug("dataset.step_table_made is not True")
                logging.info("running make_step_table")

                # update nom_cap in case it is given as argument to make_summary:
                data.nom_cap = nom_cap
                self.make_step_table()

        if not self.data.raw.index.is_unique:
            warnings.warn(f"{self.cell_name}: index is not unique for raw data")
            if remove_duplicates:
                logging.debug("removing duplicates before making summary")
                self.data.raw = self.data.raw[
                    ~self.data.raw.index.duplicated(keep="first")
                ]
            else:
                warnings.warn(
                    "You should remove the duplicates before making summary. For example using"
                    "c.data.raw = c.data.raw[~raw.index.duplicated(keep='first')]"
                )

        if use_cellpy_stat_file:
            summary_df = data.summary
            try:
                summary = self.data.raw[self.headers_normal.data_point_txt].isin(
                    summary_df[self.headers_normal.data_point_txt]
                )
            except KeyError:
                # TODO: remove this "escape" and instead raise Error asking
                #  the user to not use the stat-file if it is not working properly:
                logging.info("Error in stat_file (?) - using _select_last")
                summary = selector()
        else:
            summary = selector()

        if not summary.index.is_unique:
            warnings.warn(f"{self.cell_name}: index is not unique for summary data")

        column_names = summary.columns
        # TODO @jepe: use pandas.DataFrame properties instead (.len, .reset_index), but maybe first
        #  figure out if this is really needed and why it was implemented in the first place.
        summary_length = len(summary[column_names[0]])
        summary.index = list(range(summary_length))

        if select_columns:
            logging.debug("keeping only selected set of columns")
            columns_to_keep = [
                self.headers_normal.charge_capacity_txt,
                self.headers_normal.cycle_index_txt,
                self.headers_normal.data_point_txt,
                self.headers_normal.datetime_txt,
                self.headers_normal.discharge_capacity_txt,
                self.headers_normal.test_time_txt,
            ]
            for cn in column_names:
                if not columns_to_keep.count(cn):
                    summary.pop(cn)

        data.summary = summary

        # ----------------- calculated values -----------------------

        if self.cycle_mode == "anode":
            logging.info(
                "Assuming cycling in anode half-data (discharge before charge) mode"
            )
            _first_step_txt = self.headers_summary.discharge_capacity
            _second_step_txt = self.headers_summary.charge_capacity
        else:
            logging.info("Assuming cycling in full-data / cathode mode")
            _first_step_txt = self.headers_summary.charge_capacity
            _second_step_txt = self.headers_summary.discharge_capacity

        # ---------------- absolute -------------------------------

        data = self._generate_absolute_summary_columns(
            data, _first_step_txt, _second_step_txt
        )
        data = self._equivalent_cycles_to_summary(
            data, _first_step_txt, _second_step_txt, nom_cap_abs, normalization_cycles
        )

        # getting the C-rates, using values from step-table (so it will not be changed
        # even though you provide make_summary with a new nom_cap unfortunately):
        data = self._c_rates_to_summary(data)

        # ----------------- specifics ----------------------------------------
        specific_columns = self.headers_summary.specific_columns
        for mode in specifics:
            data = self._generate_specific_summary_columns(data, mode, specific_columns)

        # TODO @jepe: refactor this to method:
        if find_end_voltage:
            data = self._end_voltage_to_summary(data)

        if find_ir and (
            self.headers_normal.internal_resistance_txt in data.raw.columns
        ):
            data = self._ir_to_summary(data)

        if sort_my_columns:
            logging.debug("sorting columns")
            new_first_col_list = [
                self.headers_normal.datetime_txt,
                self.headers_normal.test_time_txt,
                self.headers_normal.data_point_txt,
                self.headers_normal.cycle_index_txt,
            ]
            data.summary = self.set_col_first(data.summary, new_first_col_list)

        if cycle_index_as_index:
            index_col = self.headers_summary.cycle_index
            try:
                data.summary.set_index(index_col, inplace=True)
            except KeyError:
                logging.debug("Setting cycle_index as index failed")

        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")
        return data

    def _make_summar_legacy(
        self,
        mass=None,
        update_it=False,
        select_columns=True,
        # find_ocv=False,  # deprecated
        find_ir=True,
        find_end_voltage=False,
        ensure_step_table=True,
        sort_my_columns=True,
        use_cellpy_stat_file=False,
        normalization_cycles=None,
        nom_cap=None,
        nom_cap_specifics="gravimetric",
        add_daniel_columns=False,  # deprecated
        **kwargs,
    ):
        # ---------------- discharge loss --------------------------------------
        # Assume that both charge and discharge is defined as positive.
        # The gain for cycle n (compared to cycle n-1)
        # is then cap[n] - cap[n-1]. The loss is the negative of gain.
        # discharge loss = discharge_cap[n-1] - discharge_cap[n]

        # ---------------- charge loss -----------------------------------------
        # charge loss = charge_cap[n-1] - charge_cap[n]

        # --------------- D.L. -------------------------------------------------
        # NH_n: high level at cycle n. The slope NHn=f(n) is linked to SEI loss
        # NB_n: low level (summation of irreversible capacities) at cycle n
        # Ref_n: sum[i=1 to ref](Q_charge_i - Q_discharge_i) + Q_charge_ref
        # Typically, ref should be a number where the electrode has become
        # stable (i.e. 5).
        # NBn/100 = sum[i=1 to n](Q_charge_i - Q_discharge_i) / Ref_n
        # NHn/100 = Q_charge_n + sum[i=1 to n-1](Q_charge_i - Q_discharge_i)
        #  / Ref_n
        # NH = 100%  ok if NH<120 at n=200
        # NB = 20% stable (or less)

        # --------- shifted capacities ------------------------------------------
        #  as defined by J. Dahn et al.
        # Note! Should double-check this (including checking
        # if it is valid in cathode mode).

        # --------- relative irreversible capacities -----------------------------
        #  as defined by Gauthier et al.
        # RIC = discharge_cap[n-1] - charge_cap[n] /  charge_cap[n-1]
        # RIC_SEI = discharge_cap[n] - charge_cap[n-1] / charge_cap[n-1]
        # RIC_disconnect = charge_cap[n-1] - charge_cap[n] / charge_cap[n-1]

        # --------- notes --------------------------------------------------------
        # @jepe 2022.09.11: trying to use .assign from now on
        #   as it is recommended (but this will likely increase memory usage)

        # TODO: add this to arguments and possible prms:
        if nom_cap_specifics is None:
            nom_cap_specifics = self.nom_cap_specifics
        specifics = ["gravimetric", "areal"]
        cycle_index_as_index = True
        time_00 = time.time()
        logging.debug("start making summary")

        cell = self.data
        if not mass:
            mass = cell.mass or 1.0
        else:
            if update_it:
                cell.mass = mass

        if not use_cellpy_stat_file:
            logging.debug("not using cellpy statfile")

        if nom_cap is None:
            nom_cap = cell.nom_cap

        logging.info(f"Using the following nominal capacity: {nom_cap}")

        # cellpy has historically assumed that the nominal capacity (nom_cap) is specific gravimetric
        # (i.e. in units of for example mAh/g), but now we need it in absolute units (e.g. Ah). The plan
        # is to set stuff like this during initiation of the cell (but not yet):

        # generating absolute nominal capacity:
        if nom_cap_specifics == "gravimetric":
            nom_cap_abs = self.nominal_capacity_as_absolute(
                nom_cap, mass, nom_cap_specifics
            )
        elif nom_cap_specifics == "areal":
            nom_cap_abs = self.nominal_capacity_as_absolute(
                nom_cap, cell.active_electrode_area, nom_cap_specifics
            )

        # TODO: this will break because cell.volume (data.volume) is not set yet
        elif nom_cap_specifics == "volumetric":
            nom_cap_abs = self.nominal_capacity_as_absolute(
                nom_cap, cell.volume, nom_cap_specifics
            )

        # ensuring that a step table exists:
        if ensure_step_table:
            logging.debug("ensuring existence of step-table")
            if not cell.has_steps:
                logging.debug("dataset.step_table_made is not True")
                logging.info("running make_step_table")

                # update nom_cap in case it is given as argument to make_summary:
                cell.nom_cap = nom_cap
                self.make_step_table()

        summary_df = cell.summary

        raw = cell.raw
        if use_cellpy_stat_file:
            try:
                summary_requirement = raw[self.headers_normal.data_point_txt].isin(
                    summary_df[self.headers_normal.data_point_txt]
                )
            except KeyError:
                logging.info("Error in stat_file (?) - using _select_last")
                summary_requirement = self._select_last(raw)
        else:
            summary_requirement = self._select_last(raw)
        summary = raw[summary_requirement].copy()
        column_names = summary.columns
        # TODO @jepe: use pandas.DataFrame properties instead (.len, .reset_index), but maybe first
        #  figure out if this is really needed and why it was implemented in the first place.
        summary_length = len(summary[column_names[0]])
        summary.index = list(range(summary_length))

        if select_columns:
            logging.debug("keeping only selected set of columns")
            columns_to_keep = [
                self.headers_normal.charge_capacity_txt,
                self.headers_normal.cycle_index_txt,
                self.headers_normal.data_point_txt,
                self.headers_normal.datetime_txt,
                self.headers_normal.discharge_capacity_txt,
                self.headers_normal.test_time_txt,
            ]
            for cn in column_names:
                if not columns_to_keep.count(cn):
                    summary.pop(cn)

        cell.summary = summary

        if self.cycle_mode == "anode":
            logging.info(
                "Assuming cycling in anode half-data (discharge before charge) mode"
            )
            _first_step_txt = self.headers_summary.discharge_capacity
            _second_step_txt = self.headers_summary.charge_capacity
        else:
            logging.info("Assuming cycling in full-data / cathode mode")
            _first_step_txt = self.headers_summary.charge_capacity
            _second_step_txt = self.headers_summary.discharge_capacity

        # ---------------- absolute -------------------------------

        cell = self._generate_absolute_summary_columns(
            cell, _first_step_txt, _second_step_txt
        )
        cell = self._equivalent_cycles_to_summary(
            cell, _first_step_txt, _second_step_txt, nom_cap_abs, normalization_cycles
        )

        # getting the C-rates, using values from step-table (so it will not be changed
        # even though you provide make_summary with a new nom_cap unfortunately):
        cell = self._c_rates_to_summary(cell)

        # ----------------- specifics ----------------------------------------
        specific_columns = self.headers_summary.specific_columns
        for mode in specifics:
            cell = self._generate_specific_summary_columns(cell, mode, specific_columns)

        if add_daniel_columns:
            warnings.warn(
                "Adding daniel columns is deprecated.", DeprecationWarning, stacklevel=2
            )

        # TODO @jepe: refactor this to method:
        if find_end_voltage:
            cell = self._end_voltage_to_summary(cell)

        if find_ir and (
            self.headers_normal.internal_resistance_txt in cell.raw.columns
        ):
            cell = self._ir_to_summary(cell)

        if sort_my_columns:
            logging.debug("sorting columns")
            new_first_col_list = [
                self.headers_normal.datetime_txt,
                self.headers_normal.test_time_txt,
                self.headers_normal.data_point_txt,
                self.headers_normal.cycle_index_txt,
            ]
            cell.summary = self.set_col_first(cell.summary, new_first_col_list)

        if cycle_index_as_index:
            index_col = self.headers_summary.cycle_index
            try:
                cell.summary.set_index(index_col, inplace=True)
            except KeyError:
                logging.debug("Setting cycle_index as index failed")

        logging.debug(f"(dt: {(time.time() - time_00):4.2f}s)")

    def _generate_absolute_summary_columns(
        self, data, _first_step_txt, _second_step_txt
    ) -> Data:
        summary = data.summary
        summary[self.headers_summary.coulombic_efficiency] = (
            100 * summary[_second_step_txt] / summary[_first_step_txt]
        )
        summary[self.headers_summary.cumulated_coulombic_efficiency] = summary[
            self.headers_summary.coulombic_efficiency
        ].cumsum()

        capacity_columns = {
            self.headers_summary.charge_capacity: summary[
                self.headers_normal.charge_capacity_txt
            ],
            self.headers_summary.discharge_capacity: summary[
                self.headers_normal.discharge_capacity_txt
            ],
        }
        summary = summary.assign(**capacity_columns)

        calculated_from_capacity_columns = {
            self.headers_summary.cumulated_charge_capacity: summary[
                self.headers_summary.charge_capacity
            ].cumsum(),
            self.headers_summary.cumulated_discharge_capacity: summary[
                self.headers_summary.discharge_capacity
            ].cumsum(),
            self.headers_summary.discharge_capacity_loss: (
                summary[self.headers_summary.discharge_capacity].shift(1)
                - summary[self.headers_summary.discharge_capacity]
            ),
            self.headers_summary.charge_capacity_loss: (
                summary[self.headers_summary.charge_capacity].shift(1)
                - summary[self.headers_summary.charge_capacity]
            ),
            self.headers_summary.coulombic_difference: (
                summary[_first_step_txt] - summary[_second_step_txt]
            ),
        }

        summary = summary.assign(**calculated_from_capacity_columns)

        calculated_from_coulombic_efficiency_columns = {
            self.headers_summary.cumulated_coulombic_difference: summary[
                self.headers_summary.coulombic_difference
            ].cumsum(),
        }

        summary = summary.assign(**calculated_from_coulombic_efficiency_columns)
        calculated_from_capacity_loss_columns = {
            self.headers_summary.cumulated_discharge_capacity_loss: summary[
                self.headers_summary.discharge_capacity_loss
            ].cumsum(),
            self.headers_summary.cumulated_charge_capacity_loss: summary[
                self.headers_summary.charge_capacity_loss
            ].cumsum(),
        }
        summary = summary.assign(**calculated_from_capacity_loss_columns)

        individual_edge_movement = summary[_first_step_txt] - summary[_second_step_txt]
        shifted_charge_capacity_column = {
            self.headers_summary.shifted_charge_capacity: individual_edge_movement.cumsum(),
        }
        summary = summary.assign(**shifted_charge_capacity_column)

        shifted_discharge_capacity_column = {
            self.headers_summary.shifted_discharge_capacity: summary[
                self.headers_summary.shifted_charge_capacity
            ]
            + summary[_first_step_txt],
        }
        summary = summary.assign(**shifted_discharge_capacity_column)
        ric = (summary[_first_step_txt].shift(1) - summary[_second_step_txt]) / summary[
            _second_step_txt
        ].shift(1)
        ric_column = {self.headers_summary.cumulated_ric: ric.cumsum()}
        summary = summary.assign(**ric_column)
        summary[self.headers_summary.cumulated_ric] = ric.cumsum()
        ric_sei = (
            summary[_first_step_txt] - summary[_second_step_txt].shift(1)
        ) / summary[_second_step_txt].shift(1)
        ric_sei_column = {self.headers_summary.cumulated_ric_sei: ric_sei.cumsum()}
        summary = summary.assign(**ric_sei_column)
        ric_disconnect = (
            summary[_second_step_txt].shift(1) - summary[_second_step_txt]
        ) / summary[_second_step_txt].shift(1)
        ric_disconnect_column = {
            self.headers_summary.cumulated_ric_disconnect: ric_disconnect.cumsum()
        }
        data.summary = summary.assign(**ric_disconnect_column)

        return data

    def _generate_specific_summary_columns(
        self, data: Data, mode: str, specific_columns: Sequence
    ) -> Data:
        specific_converter = self.get_converter_to_specific(dataset=data, mode=mode)
        summary = data.summary
        for col in specific_columns:
            logging.debug(f"generating specific column {col}_{mode}")
            summary[f"{col}_{mode}"] = specific_converter * summary[col]
        data.summary = summary
        return data

    def _c_rates_to_summary(self, data: Data) -> Data:
        logging.debug("Extracting C-rates")
        summary = data.summary
        steps = self.data.steps

        charge_steps = steps.loc[
            steps.type == "charge",
            [self.headers_step_table.cycle, self.headers_step_table.rate_avr],
        ].rename(
            columns={
                self.headers_step_table.rate_avr: self.headers_summary.charge_c_rate
            }
        )
        summary = summary.merge(
            charge_steps.drop_duplicates(
                subset=[self.headers_step_table.cycle], keep="first"
            ),
            left_on=self.headers_summary.cycle_index,
            right_on=self.headers_step_table.cycle,
            how="left",
        ).drop(columns=self.headers_step_table.cycle)

        discharge_steps = steps.loc[
            steps.type == "discharge",
            [self.headers_step_table.cycle, self.headers_step_table.rate_avr],
        ].rename(
            columns={
                self.headers_step_table.rate_avr: self.headers_summary.discharge_c_rate
            }
        )
        summary = summary.merge(
            discharge_steps.drop_duplicates(
                subset=[self.headers_step_table.cycle], keep="first"
            ),
            left_on=self.headers_summary.cycle_index,
            right_on=self.headers_step_table.cycle,
            how="left",
        ).drop(columns=self.headers_step_table.cycle)
        data.summary = summary
        return data

    def _equivalent_cycles_to_summary(
        self,
        data: Data,
        _first_step_txt: str,
        _second_step_txt: str,
        nom_cap: float,
        normalization_cycles: Union[Sequence, int, None],
    ) -> Data:
        # The method currently uses the charge capacity for calculating equivalent cycles. This
        # can be easily extended to also allow for choosing the discharge capacity later on if
        # it turns out that to be needed.

        summary = data.summary

        if normalization_cycles is not None:
            logging.info(
                f"Using these cycles for finding the nominal capacity: {normalization_cycles}"
            )
            if not isinstance(normalization_cycles, (list, tuple)):
                normalization_cycles = [normalization_cycles]

            cap_ref = summary.loc[
                summary[self.headers_normal.cycle_index_txt].isin(normalization_cycles),
                _first_step_txt,
            ]
            if not cap_ref.empty:
                nom_cap = cap_ref.mean()
            else:
                logging.info(f"Empty reference cycle(s)")

        normalized_cycle_index_column = {
            self.headers_summary.normalized_cycle_index: summary[
                self.headers_summary.cumulated_charge_capacity
            ]
            / nom_cap
        }
        summary = summary.assign(**normalized_cycle_index_column)
        data.summary = summary
        return data

    def _ir_to_summary(self, data):
        # should check:  test.charge_steps = None,
        # test.discharge_steps = None
        # THIS DOES NOT WORK PROPERLY!!!!
        # Found a file where it writes IR for cycle n on cycle n+1
        # This only picks out the data on the last IR step before
        summary = data.summary
        raw = data.raw

        logging.debug("finding ir")
        only_zeros = summary[self.headers_normal.discharge_capacity_txt] * 0.0
        discharge_steps = self.get_step_numbers(
            steptype="discharge",
            allctypes=False,
        )
        charge_steps = self.get_step_numbers(
            steptype="charge",
            allctypes=False,
        )
        ir_indexes = []
        ir_values = []
        ir_values2 = []
        for i in summary.index:
            # selecting the appropriate cycle
            cycle = summary.iloc[i][self.headers_normal.cycle_index_txt]
            step = discharge_steps[cycle]
            if step[0]:
                ir = raw.loc[
                    (raw[self.headers_normal.cycle_index_txt] == cycle)
                    & (data.raw[self.headers_normal.step_index_txt] == step[0]),
                    self.headers_normal.internal_resistance_txt,
                ]
                # This will not work if there are more than one item in step
                ir = ir.values[0]
            else:
                ir = 0
            step2 = charge_steps[cycle]
            if step2[0]:
                ir2 = raw[
                    (raw[self.headers_normal.cycle_index_txt] == cycle)
                    & (data.raw[self.headers_normal.step_index_txt] == step2[0])
                ][self.headers_normal.internal_resistance_txt].values[0]
            else:
                ir2 = 0
            ir_indexes.append(i)
            ir_values.append(ir)
            ir_values2.append(ir2)
        ir_frame = only_zeros + ir_values
        ir_frame2 = only_zeros + ir_values2
        summary.insert(0, column=self.headers_summary.ir_discharge, value=ir_frame)
        summary.insert(0, column=self.headers_summary.ir_charge, value=ir_frame2)
        data.summary = summary
        return data

    def _end_voltage_to_summary(self, data):
        # needs to be fixed so that end-voltage also can be extracted
        # from the summary
        ev_t0 = time.time()
        raw = data.raw
        summary = data.summary

        logging.debug("finding end-voltage")
        logging.debug(f"dt: {time.time() - ev_t0}")
        only_zeros_discharge = summary[self.headers_normal.discharge_capacity_txt] * 0.0
        only_zeros_charge = summary[self.headers_normal.charge_capacity_txt] * 0.0
        logging.debug("need to collect discharge steps")
        discharge_steps = self.get_step_numbers(steptype="discharge", allctypes=False)
        logging.debug(f"dt: {time.time() - ev_t0}")
        logging.debug("need to collect charge steps")
        charge_steps = self.get_step_numbers(steptype="charge", allctypes=False)
        logging.debug(f"dt: {time.time() - ev_t0}")
        endv_indexes = []
        endv_values_dc = []
        endv_values_c = []
        logging.debug("starting iterating through the index")
        for i in summary.index:
            cycle = summary.iloc[i][self.headers_normal.cycle_index_txt]
            step = discharge_steps[cycle]

            # finding end voltage for discharge
            if step[-1]:  # selecting last
                end_voltage_dc = raw[
                    (raw[self.headers_normal.cycle_index_txt] == cycle)
                    & (data.raw[self.headers_normal.step_index_txt] == step[-1])
                ][self.headers_normal.voltage_txt]
                # This will not work if there are more than one item in step
                end_voltage_dc = end_voltage_dc.values[-1]  # selecting
            else:
                end_voltage_dc = 0  # could also use numpy.nan

            # finding end voltage for charge
            step2 = charge_steps[cycle]
            if step2[-1]:
                end_voltage_c = raw[
                    (raw[self.headers_normal.cycle_index_txt] == cycle)
                    & (data.raw[self.headers_normal.step_index_txt] == step2[-1])
                ][self.headers_normal.voltage_txt]
                end_voltage_c = end_voltage_c.values[-1]
            else:
                end_voltage_c = 0
            endv_indexes.append(i)
            endv_values_dc.append(end_voltage_dc)
            endv_values_c.append(end_voltage_c)

        ir_frame_dc = only_zeros_discharge + endv_values_dc
        ir_frame_c = only_zeros_charge + endv_values_c
        data.summary.insert(
            0, column=self.headers_summary.end_voltage_discharge, value=ir_frame_dc
        )
        data.summary.insert(
            0, column=self.headers_summary.end_voltage_charge, value=ir_frame_c
        )

        return data

    def inspect_nominal_capacity(self, cycles=None):
        """Method for estimating the nominal capacity

        Args:
            cycles (list of ints): the cycles where it is assumed that the data reaches nominal capacity.

        Returns:
            Nominal capacity (float).
        """
        logging.debug("inspecting: nominal capacity")
        print("Sorry! This method is still under development.")
        print("Maybe you can plot your data and find the nominal capacity yourself?")
        if cycles is None:
            cycles = [1, 2, 3]

        summary = self.data.summary

        try:
            nc = summary.loc[
                summary[self.headers_normal.cycle_index_txt].isin(cycles),
                self.headers_summary.discharge_capacity,
            ].mean()
            print("All I can say for now is that the average discharge capacity")
            print(f"for the cycles {cycles} is {nc:0.2f}")
            nc = float(nc)

        except ZeroDivisionError:
            print("zero division error")
            nc = None

        return nc

    # ================ Experimental features =================

    # ---------------------- update --------------------------

    def dev_update(self, file_names=None, **kwargs):
        """Experimental method for updating a cellpy-file with new raw-files."""

        print("NOT FINISHED YET!")
        if len(self.data.raw_data_files) != 1:
            logging.warning("Merged data. But can only update based on the last file")
            print(self.data.raw_data_files)
            for fid in self.data.raw_data_files:
                print(fid)
        last = self.data.raw_data_files[0].last_data_point

        self.dev_update_from_raw(
            file_names=file_names, data_points=[last, None], **kwargs
        )
        print("lets try to merge")
        self.data = self.dev_update_merge()
        print("now it is time to update the step table")
        self.dev_update_make_steps()
        print("and finally, lets update the summary")
        self.dev_update_make_summary()

    def dev_update_loadcell(
        self,
        raw_files,
        cellpy_file=None,
        mass=None,
        summary_on_raw=False,
        summary_ir=True,
        summary_end_v=True,
        force_raw=False,
        use_cellpy_stat_file=None,
        nom_cap=None,
        selector=None,
        **kwargs,
    ):
        """Load cell from raw-files or cellpy-file.

        This is an experimental method. It is not finished yet and the logics
        in this method will most likely be moved to other methods since
        new versions of cellpy is now based on the get method (it implements
        similar logic as loadcell, but is more flexible and easier to use).



        """
        logging.info("Started cellpy.cellreader.dev_update_loadcell")

        if cellpy_file is None or force_raw:
            similar = None
        else:
            similar = self.check_file_ids(raw_files, cellpy_file, detailed=True)

        logging.debug("checked if the files were similar")

        if similar is None:
            # forcing to load only raw_files
            self.from_raw(raw_files, **kwargs)
            if self._validate_cell():
                if mass:
                    self.set_mass(mass)
                if summary_on_raw:
                    self.make_summary(
                        find_ir=summary_ir,
                        find_end_voltage=summary_end_v,
                        use_cellpy_stat_file=use_cellpy_stat_file,
                        nom_cap=nom_cap,
                    )
            else:
                logging.warning("Empty run!")
            return self

        self.load(cellpy_file, selector=selector)
        if mass:
            self.set_mass(mass)

        if all(similar.values()):
            logging.info("Everything is up to date")
            return

        start_file = True
        for i, f in enumerate(raw_files):
            # TODO: -> OtherPath?
            f = Path(f)
            if not similar[f.name] and start_file:
                try:
                    last_data_point = self.data.raw_data_files[i].last_data_point
                except IndexError:
                    last_data_point = 0

                self.dev_update_from_raw(
                    file_names=f, data_points=[last_data_point, None]
                )
                self.data = self.dev_update_merge()

            elif not similar[f.name]:
                try:
                    last_data_point = self.data.raw_data_files[i].last_data_point
                except IndexError:
                    last_data_point = 0

                self.dev_update_from_raw(
                    file_names=f, data_points=[last_data_point, None]
                )
                self.merge()

            start_file = False

        self.dev_update_make_steps()
        self.dev_update_make_summary(
            # all_tests=False,
            # find_ocv=summary_ocv,
            find_ir=summary_ir,
            find_end_voltage=summary_end_v,
            use_cellpy_stat_file=use_cellpy_stat_file,
        )
        return self

    # TODO @jepe (v.1.0.0): update this to use single data instances (i.e. to cell from cells)
    def dev_update_merge(self, t1, t2):
        print("NOT FINISHED YET - but very close")

        if t1.raw.empty:
            logging.debug("OBS! the first dataset is empty")

        if t2.raw.empty:
            logging.debug("the second dataset was empty")
            logging.debug(" -> merged contains only first")
            return t1
        test = t1

        cycle_index_header = self.headers_normal.cycle_index_txt

        if not t1.raw.empty:
            t1.raw = t1.raw.iloc[:-1]
            raw2 = pd.concat([t1.raw, t2.raw], ignore_index=True)
            test.raw = raw2
        else:
            test = t2
        logging.debug(" -> merged with new dataset")

        return test

    def dev_update_make_steps(self, **kwargs):
        old_steps = self.data.steps.iloc[:-1]
        # Note! hard-coding header name (might fail if changing default headers)
        from_data_point = self.data.steps.iloc[-1].point_first
        new_steps = self.make_step_table(from_data_point=from_data_point, **kwargs)
        merged_steps = pd.concat([old_steps, new_steps]).reset_index(drop=True)
        self.data.steps = merged_steps

    def dev_update_make_summary(self, **kwargs):
        print("NOT FINISHED YET - but not critical")
        # Update not implemented yet, running full summary calculations for now.
        # For later:
        # old_summary = self.data.summary.iloc[:-1]
        cycle_index_header = self.headers_summary.cycle_index
        from_cycle = self.data.summary.iloc[-1][cycle_index_header]
        self.make_summary(from_cycle=from_cycle, **kwargs)
        # For later:
        # (Remark! need to solve how to merge cumulated columns)
        # new_summary = self.make_summary(from_cycle=from_cycle)
        # merged_summary = pd.concat([old_summary, new_summary]).reset_index(drop=True)
        # self.data.summary = merged_summary

    def dev_update_from_raw(self, file_names=None, data_points=None, **kwargs):
        """This method is under development. Using this to develop updating files
        with only new data.
        """
        print("NOT FINISHED YET")
        # TODO @jepe: implement changes from original from_raw method introduced after this one was last edited.
        if file_names:
            self.file_names = file_names

        if file_names is None:
            logging.info(
                "No filename given and no stored in the file_names "
                "attribute. Returning None"
            )
            return None

        if not isinstance(self.file_names, (list, tuple)):
            self.file_names = [file_names]

        raw_file_loader = self.loader

        set_number = 0
        cell = None

        logging.debug("start iterating through file(s)")

        for f in self.file_names:
            logging.debug("loading raw file:")
            logging.debug(f"{f}")

            # get a list of cellpy.readers.core.Data objects
            # cell = raw_file_loader(f, data_points=data_points, **kwargs)
            # remark that the bounds are included (i.e. the first datapoint
            # is 5000.

            logging.debug(
                "added the data set - merging file info  - oh no; I am not implemented yet"
            )

            # raw_data_file = copy.deepcopy(test[set_number].raw_data_files[0])
            # file_size = test[set_number].raw_data_files_length[0]

            # test[set_number].raw_data_files.append(raw_data_file)
            # test[set_number].raw_data_files_length.append(file_size)
            # return test
        # cell[set_number].raw_units = self._set_raw_units()
        # self.cells.append(cell[set_number])
        # self.status_dataset = self._validate_cell()
        # self._invent_a_cell_name()
        return self


def get(
    filename=None,
    instrument=None,
    instrument_file=None,
    cellpy_file=None,
    cycle_mode=None,
    mass: Union[str, numbers.Number] = None,
    nominal_capacity: Union[str, numbers.Number] = None,
    nom_cap_specifics=None,
    loading=None,
    area: Union[str, numbers.Number] = None,
    estimate_area=True,
    logging_mode=None,
    auto_pick_cellpy_format=True,
    auto_summary=True,
    units=None,
    step_kwargs=None,
    summary_kwargs=None,
    selector=None,
    testing=False,
    refuse_copying=False,
    initialize=False,
    debug=False,
    **kwargs,
):
    """Create a CellpyCell object

    Args:
        filename (str, os.PathLike, OtherPath, or list of raw-file names): path to file(s) to load
        instrument (str): instrument to use (defaults to the one in your cellpy config file)
        instrument_file (str or path): yaml file for custom file type
        cellpy_file (str, os.PathLike, or OtherPath): if both filename (a raw-file) and cellpy_file (a cellpy file)
            is provided, cellpy will try to check if the raw-file is has been updated since the
            creation of the cellpy-file and select this instead of the raw file if cellpy thinks
            they are similar (use with care!).
        logging_mode (str): "INFO" or "DEBUG"
        cycle_mode (str): the cycle mode (e.g. "anode" or "full_cell")
        mass (float): mass of active material (mg) (defaults to mass given in cellpy-file or 1.0)
        nominal_capacity (float): nominal capacity for the cell (e.g. used for finding C-rates)
        nom_cap_specifics (str): either "gravimetric" (pr mass), or "areal" (per area).
            ("volumetric" is not fully implemented yet - let us know if you need it).
        loading (float): loading in units [mass] / [area]
        area (float): active electrode area (e.g. used for finding the areal capacity)
        estimate_area (bool): calculate area from loading if given (defaults to True)
        auto_pick_cellpy_format (bool): decide if it is a cellpy-file based on suffix.
        auto_summary (bool): (re-) create summary.
        units (dict): update cellpy units (used after the file is loaded, e.g. when creating summary).
        step_kwargs (dict): sent to make_steps
        summary_kwargs (dict): sent to make_summary
        selector (dict): passed to load (when loading cellpy-files).
        testing (bool): set to True if testing (will for example prevent making .log files)
        refuse_copying (bool): set to True if you do not want to copy the raw-file before loading.
        initialize (bool): set to True if you want to initialize the CellpyCell object (probably only
            useful if you want to return a cellpy-file with no data in it)
        debug (bool): set to True if you want to debug the loader.
        **kwargs: sent to the loader

    Keyword args ("arbin_res"):
        bad_steps (list of tuples): (c, s) tuples of steps s (in cycle c) to skip loading [arbin_res].
        dataset_number (int): the data set number ('Test-ID') to select if you are dealing
            with arbin files with more than one data-set. Defaults to selecting all data-sets and merging them.
        data_points (tuple of ints): load only data from data_point[0] to
                data_point[1] (use None for infinite).
        increment_cycle_index (bool): increment the cycle index if merging several datasets (default True).

    Keyword args ("maccor_txt", "neware_txt", "local_instrument", "custom"):
        sep (str): separator used in the file.
        skip_rows (int): number of rows to skip in the beginning of the file.
        header (int): row number of the header.
        encoding (str): encoding of the file.
        decimal (str): decimal separator.
        thousand (str): thousand separator.
        pre_processor_hook (callable): pre-processors to use.

    Keyword args ("pec_csv"):
        bad_steps (list): separator used in the file (not implemented yet).

    Returns:
        CellpyCell object (if successful, None if not)

    Examples:
        >>> # read an arbin .res file and create a cellpy object with
        >>> # populated summary and step-table:
        >>> c = cellpy.get("my_data.res", instrument="arbin_res", mass=1.14, area=2.12, loading=1.2, nom_cap=155.2)
        >>>
        >>> # load a cellpy-file:
        >>> c = cellpy.get("my_cellpy_file.clp")
        >>>
        >>> # load a txt-file exported from Maccor:
        >>> c = cellpy.get("my_data.txt", instrument="maccor_txt", model="one")
        >>>
        >>> # load a raw-file if it is newer than the corresponding cellpy-file,
        >>> # if not, load the cellpy-file:
        >>> c = cellpy.get("my_data.res", cellpy_file="my_data.clp")
        >>>
        >>> # load a file with a custom file-description:
        >>> c = cellpy.get("my_file.csv", instrument_file="my_instrument.yaml")
        >>>
        >>> # load three subsequent raw-files (of one cell) and merge them:
        >>> c = cellpy.get(["my_data_01.res", "my_data_02.res", "my_data_03.res"])
        >>>
        >>> # load a data set and get the summary charge and discharge capacities
        >>> # in Ah/g:
        >>> c = cellpy.get("my_data.res", units=dict(capacity="Ah"))
        >>>
        >>> # get an empty CellpyCell instance:
        >>> c = cellpy.get()  # or c = cellpy.get(initialize=True) if you want to initialize it.

    """

    # TODO: implement volume as parameter and 'volumetric' as nom_cap_specifics option.

    from cellpy import log

    db_readers = ["arbin_sql", "arbin_sql_7"]
    instruments_with_colliding_file_suffix = ["arbin_sql_h5"]

    step_kwargs = step_kwargs or {}
    summary_kwargs = summary_kwargs or {}
    load_cellpy_file = False
    logging_mode = "DEBUG" if testing else logging_mode
    log.setup_logging(default_level=logging_mode, testing=testing)
    logging.debug("-------running-get--------")
    cellpy_instance = CellpyCell(debug=debug, initialize=initialize)
    logging.debug(f"created CellpyCell instance")

    logging.debug(f"{cellpy_file=}")
    logging.debug(f"{filename=}")

    # used if all you want is an empty CellpyCell object
    if filename is None:
        if cellpy_file is None:
            logging.info("Running cellpy.get without a filename")
            logging.info("Returning an empty CellpyCell object.")
            cellpy_instance = _update_meta(
                cellpy_instance,
                cycle_mode=cycle_mode,
                mass=mass,
                nominal_capacity=nominal_capacity,
                nom_cap_specifics=nom_cap_specifics,
                area=area,
                loading=loading,
                estimate_area=estimate_area,
                units=units,
            )
            return cellpy_instance

        else:
            load_cellpy_file = True
            filename = OtherPath(cellpy_file)

    if isinstance(filename, (list, tuple)):
        logging.debug("got a list or tuple of names")
        load_cellpy_file = False
    else:
        logging.debug("got a single name")
        logging.debug(f"{filename=}")
        filename = OtherPath(filename)
        if (
            auto_pick_cellpy_format
            and instrument not in instruments_with_colliding_file_suffix
            and filename.suffix in [".h5", ".hdf5", ".cellpy", ".cpy"]
        ):
            load_cellpy_file = True

    if filename and cellpy_file and not load_cellpy_file:
        try:
            similar = cellpy_instance.check_file_ids(filename, cellpy_file)
            logging.debug(f"checked if the files were similar")
            if similar:
                load_cellpy_file = True
                filename = OtherPath(cellpy_file)
        except Exception as e:
            logging.debug(f"Error during checking if similar: {e}")
            logging.debug("Setting load_cellpy_file to False")

    if load_cellpy_file:
        logging.info(f"Loading cellpy-file: {filename}")
        if kwargs.pop("post_processor_hook", None) is not None:
            logging.warning(
                "post_processor_hook is not allowed when loading cellpy-files"
            )
        cellpy_instance.load(filename, selector=selector, **kwargs)
        return cellpy_instance

    logging.debug("Prepare for loading raw-file(s)")
    logging.debug(f"checking instrument and instrument_file")
    if instrument_file is not None:
        logging.debug(f"got instrument file {instrument_file=}")
        cellpy_instance.set_instrument(
            instrument="custom", instrument_file=instrument_file
        )

    elif instrument is not None:
        logging.debug(f"got instrument in stead of instrument file, {instrument=}")
        model = kwargs.pop("model", None)
        cellpy_instance.set_instrument(instrument=instrument, model=model, **kwargs)

    is_a_file = True
    if cellpy_instance.tester in db_readers:
        is_a_file = False

    logging.info(f"Loading raw-file: {filename}")
    cellpy_instance.from_raw(
        filename, is_a_file=is_a_file, refuse_copying=refuse_copying, **kwargs
    )

    if not cellpy_instance:
        print("Could not load file: check log!")
        print("Returning None")
        return

    # fix for allowing for setting nom_cap_specifics the "old" way:
    if nom_cap_specifics is None:
        nom_cap_specifics = summary_kwargs.pop("nom_cap_specifics", None)
        if nom_cap_specifics is None:
            nom_cap_specifics = step_kwargs.pop("nom_cap_specifics", None)

    cellpy_instance = _update_meta(
        cellpy_instance,
        cycle_mode=cycle_mode,
        mass=mass,
        nominal_capacity=nominal_capacity,
        nom_cap_specifics=nom_cap_specifics,
        area=area,
        loading=loading,
        estimate_area=estimate_area,
        units=units,
    )

    if auto_summary:
        logging.info("Creating step table")
        cellpy_instance.make_step_table(**step_kwargs)
        logging.info("Creating summary data")
        cellpy_instance.make_summary(**summary_kwargs)

    logging.info("Created CellpyCell object")
    return cellpy_instance


def _update_meta(
    cellpy_instance,
    cycle_mode=None,
    mass=None,
    nominal_capacity=None,
    nom_cap_specifics=None,
    area=None,
    loading=None,
    estimate_area=None,
    units=None,
    volume=None,  # not implemented yet
):
    """Used by get to update metadata in the CellpyCell object."""
    # Note: this is a bit messy, but it is a quick fix for now.
    #       I will clean it up later.
    # Note: if you want to add more metadata or similar for use by the get function,
    #       please also add a property to the CellpyCell class (e.g. don't update
    #       the data object directly, especially if handling units).

    if cycle_mode is not None:
        logging.debug("Setting cycle mode")
        cellpy_instance.cycle_mode = cycle_mode

    if nom_cap_specifics is not None:
        logging.debug("Setting nom_cap_specifics as given")
        cellpy_instance.nom_cap_specifics = nom_cap_specifics

    if units is not None:
        logging.debug(f"Updating units: {units}")
        cellpy_instance.cellpy_units.update(units)

    if mass is not None:
        logging.info(f"Setting mass: {mass}")
        cellpy_instance.mass = mass

    if nominal_capacity is not None:
        logging.info(f"Setting nominal capacity: {nominal_capacity}")
        if nom_cap_specifics is not None and not isinstance(
            nominal_capacity, numbers.Number
        ):
            logging.info(
                "Providing nominal capacity as string might override the given nom_cap_specifics"
            )
        cellpy_instance.nom_cap = nominal_capacity

    if area is not None:
        logging.debug(f"got area: {area}")
        cellpy_instance.active_electrode_area = area

    elif loading and estimate_area:
        logging.debug("-------------AREA-CALC----------------")
        logging.debug(f"got loading: {logging}")
        area = cellpy_instance.data.mass / loading
        logging.debug(
            f"calculating area from loading ({loading}) and mass ({cellpy_instance.data.mass}): {area}"
        )
        cellpy_instance.active_electrode_area = area

    else:
        logging.debug("using default area")

    if volume is not None:
        logging.debug(f"got volume: {volume}")
        logging.critical("Volume not implemented yet")

    return cellpy_instance


# ============== Internal tests =================


def check_raw():
    import cellpy
    from cellpy.utils import example_data

    cellpy_data_instance = CellpyCell()
    res_file_path = example_data.arbin_file_path()
    cellpy.get(res_file_path)

    data_point = 2283
    step_time = 1500.05
    sum_discharge_time = 362198.12
    my_test = cellpy_data_instance.data

    summary = my_test.summary
    raw = my_test.raw
    print(summary.columns)
    print(summary.index)
    print(summary.head())
    print(summary.iloc[1, 1])
    print(summary.loc[:, "Data_Point"])
    print(summary.loc[1, "Data_Point"])

    print(raw.columns)
    # assert my_test.summary.loc["1", "data_point"] == data_point


def check_cellpy_file():
    print("running", end=" ")
    print(sys.argv[0])

    from cellpy import log

    log.setup_logging(default_level="DEBUG")

    from cellpy.utils import example_data

    f = example_data.cellpy_file_path()
    print(f)
    print(f.is_file())
    c = CellpyCell()
    c.dev_load(f, accept_old=True)
    c.make_step_table()
    c.make_summary()
    print("Here we have it")
    print(c.data.summary.columns)
    print(c.data.steps.columns)
    print(c.data.raw.columns)


def save_and_load_cellpy_file():
    # check to see if updating to new cellpy file version works
    """
    # How to update the cellpy file version

    ## Top level

    ## Metadata

    ## Summary, Raw, and Step headers

    update_headers.py

    """

    f00 = Path("../../testdata/hdf5/20160805_test001_45_cc.h5")
    f04 = Path("../../testdata/hdf5/20160805_test001_45_cc_v4.h5")
    f05 = Path("../../testdata/hdf5/20160805_test001_45_cc_v5.h5")
    f06 = Path("../../testdata/hdf5/20160805_test001_45_cc_v6.h5")
    f07 = Path("../../testdata/hdf5/20160805_test001_45_cc_v7.h5")
    f08 = Path("../../testdata/hdf5/20160805_test001_45_cc_v8.h5")
    f_tmp = Path("../../tmp/v1.h5")

    old = f08
    out = f_tmp

    print("LOADING ORIGINAL".center(80, "*"))
    c = get(old)
    for a in dir(c.data):
        if not a.startswith("__"):
            if a not in ["raw", "summary", "steps"]:
                v = getattr(c.data, a)
                print(f"{a}: {v}")

    print("SAVING".center(80, "*"))
    c.save(out)

    print("LOADING NEW".center(80, "*"))
    c = get(out)  # , logging_mode="DEBUG"
    meta_test_dependent = c.data.meta_test_dependent
    meta_common = c.data.meta_common
    print(f"{meta_test_dependent=}")
    print(f"{meta_common=}")
    print(f"{c.data.raw_limits=}")
    print(f"{c.data.raw_units=}")

    for a in dir(c.data):
        if not a.startswith("__"):
            if a not in ["raw", "summary", "steps"]:
                v = getattr(c.data, a)
                print(f"{a}: {v}")
    # print("Here we have it")
    # print(c.data.summary.columns)
    # print(c.data.steps.columns)
    # print(c.data.raw.columns)


def load_and_save_to_excel():
    from pathlib import Path
    from pprint import pprint

    print(" loading cellpy file and saving to excel ".center(80, "="))

    raw_file = Path("../../testdata/data/20160805_test001_45_cc_01.res")
    cellpy_file = Path("../../tmp/20160805_test001_45_cc.h5")
    # excel_file1 = Path("../../tmp/01_gravimetric_old_20160805_test001_45_cc.xlsx")
    excel_file2 = Path("../../tmp/02_gravimetric_20160805_test001_45_cc.xlsx")
    # excel_file3 = Path("../../tmp/03_areal_20160805_test001_45_cc.xlsx")
    # excel_file4 = Path("../../tmp/04_areal_20160805_test001_45_cc.xlsx")
    # excel_file5 = Path("../../tmp/05_areal_20160805_test001_45_cc.xlsx")

    # c = get(raw_file, area=1.55, cycle_mode="anode", nominal_capacity=2.0, summary_kwargs={"nom_cap_specifics": "areal"}, debug=True)
    # c.to_excel(excel_file1, cycles=True)
    c = get(
        raw_file,
        mass=1.55,
        cycle_mode="anode",
        nominal_capacity="3579 mAh/g",
        debug=True,
    )
    c.to_excel(excel_file2, cycles=True)
    # c = get(raw_file, area=1.55, cycle_mode="anode", nominal_capacity=2.0, nom_cap_specifics="areal", debug=True)
    # c.to_excel(excel_file3, cycles=True)
    # c = get(raw_file, area="1.55 cm**2", cycle_mode="anode", nominal_capacity="2.0 mAh/cm**2", nom_cap_specifics="areal", debug=True)
    # c.to_excel(excel_file4, cycles=True)
    # print("saved ...")
    #
    # c.save(cellpy_file)
    # c2 = get(cellpy_file)
    # pprint(c2.cellpy_units)
    # pprint(c2.data.meta_common)
    # print("loaded again ...")
    # c2.to_excel(excel_file5, raw=True, cycles=True)
    # print("saved again ...")


def check_excel():
    import openpyxl
    from openpyxl.styles import Border, Side
    import pandas as pd

    from pathlib import Path

    print(" checking excel ".center(80, "="))
    excel_file = Path("../../tmp/nothing.xlsx")
    to_excel_method_kwargs = {"index": True, "header": True}

    df = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
    n_rows, n_cols = df.shape
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="first", **to_excel_method_kwargs)
        ws = writer.sheets["first"]
        border = Border()
        face_color = "00EEEEEE"
        fill = openpyxl.styles.PatternFill(
            start_color=face_color, end_color=face_color, fill_type="solid"
        )

        for cell in ws["A"]:
            print(cell)
            cell.border = border
        for cell in ws[1]:
            print(cell)
            cell.border = border
            cell.fill = fill

    print("done")


def check_new_dot_get_methods():
    from pathlib import Path
    from pprint import pprint
    import numpy as np

    print(" loading file and checking the .get methods ".center(80, "="))
    raw_file = Path("../../testdata/data/20160805_test001_45_cc_01.res")
    c = get(
        raw_file,
        mass=1.55,
        cycle_mode="anode",
        nominal_capacity="3579 mAh/g",
        debug=True,
    )
    # pprint(c.headers_normal)
    cycles_a = [1, 2, 3]
    cycles_b = np.array([1, 2, 3])
    cycles_c = [1, 2, 3]
    a = c.get_timestamp(cycles_a, with_index=True, units="raw")
    print(f"{cycles_a=} raw".center(80, "-"))
    pprint(a)

    a = c.get_timestamp(cycles_a, units="seconds")
    print(f"{cycles_a=} seconds".center(80, "-"))
    pprint(a)

    a = c.get_timestamp(cycles_a, units="minutes")
    print(f"{cycles_a=} minutes".center(80, "-"))
    pprint(a)

    a = c.get_timestamp(cycles_a, units="hours")
    print(f"{cycles_a=} hours".center(80, "-"))
    pprint(a)

    a = c.get_timestamp(cycles_a, in_minutes=True, units="hours")
    print(f"{cycles_a=} hours".center(80, "-"))
    pprint(a)

    #
    #
    # print(f"{cycles_b=}".center(80, "-"))
    # b = c.get_timestamp(cycles_b)
    # pprint(b)
    # print(f"{cycles_c=}".center(80, "-"))
    # c = c.get_timestamp(cycles_c, with_index=False, as_frame=False)
    # pprint(c)


if __name__ == "__main__":
    check_new_dot_get_methods()
